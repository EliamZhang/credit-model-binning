# -*- coding: utf-8 -*-
"""
模型分数分箱与策略阈值分析（优化版）

核心流程：
1. 数据加载，并按时间切分 Train / OOT；
2. 在 Train 内再切分 Development / Validation，OOT 不参与合箱调参；
3. Development 学习 20 等频初始边界，边界复用到 Validation、Train、OOT；
4. 对相邻初始箱执行：小箱清理 -> 主风险指标单调合并 -> 候选档位压缩；
5. 使用 Development + Validation 选择 6~8 档最终方案；
6. OOT 只用于最终单调性、PSI、AUC、KS 和策略分段验证；
7. 基于完整 Train 生成自动通过 / 人工审核 / 拒绝阈值；
8. 输出完整 Excel 报告。

运行方式：
    python binning.py

输入目录：res/
输出文件：out/策略报告_优化版.xlsx
"""

import ast
import math
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ============================================================
# 0. 配置
# ============================================================

DATA_DIR = Path("res")
OUT_DIR = Path("out")
REPORT_PATH = OUT_DIR / "策略报告_优化版.xlsx"

SAMPLE_FILE = "sample.csv"
APPLICATION_FILE = "application_info.csv"
SCORE_FILE = "aus_old_risk_bid_mltmodel_v1_2_20260325_lgb_score.csv"

RAW_SCORE_COL = "aus_old_risk_bid_mltmodel_v1_2_v20260325_lgb_score"
SCORE_COL = "score_mlt"

TRAIN_END_MONTH = "2025-10"
OOT_START_MONTH = "2025-11"

# Train 内最后若干个月作为合箱 Validation；OOT 保持完全独立。
VALIDATION_MONTH_COUNT = 3
MIN_DEVELOPMENT_MONTH_COUNT = 3

INITIAL_BIN_COUNT = 20
INITIAL_BIN_COL = "score_mlt_bin20"
FINAL_BIN_COL = "score_mlt_final_bin"

# 当前模型按“高分高风险”处理。
HIGH_SCORE_HIGH_RISK = True

# 是否启用优化后的自动合箱；False 时使用手工兜底方案。
AUTO_SELECT_MERGE_RANGES = True

# 最终风险档位数量。
MIN_FINAL_BIN_COUNT = 6
MAX_FINAL_BIN_COUNT = 8
TARGET_FINAL_BIN_COUNT = 7

# 合箱主指标：3M30+ 笔数逾期率。
PRIMARY_RATE_COL = "3m30p_cnt_bad_rate"
PRIMARY_MATURE_COL = "3m30p_cnt_mature"
PRIMARY_BAD_COL = "3m30p_cnt_bad"
PRIMARY_GOOD_COL = "3m30p_cnt_good"

# 最终箱约束。尾部箱允许更小，但仍需满足成熟量和好坏样本量。
MIN_MIDDLE_BIN_SAMPLE_PCT = 0.05
MIN_TAIL_BIN_SAMPLE_PCT = 0.025
MIN_FINAL_BIN_MATURE_COUNT = 1000
MIN_FINAL_BIN_BAD_COUNT = 20
MIN_FINAL_BIN_GOOD_COUNT = 200

# 单调与相邻差异控制。
DEVELOPMENT_INVERSION_TOLERANCE = 0.0
VALIDATION_INVERSION_TOLERANCE = 0.003
ADJACENT_PVALUE_TO_MERGE = 0.10
MIN_ADJACENT_ABS_RATE_DIFF = 0.003

# Development / Validation 分布稳定性，仅用于候选方案选择；OOT 不参与选箱。
PREFERRED_MAX_VALIDATION_PSI = 0.05
MAX_ACCEPTABLE_VALIDATION_PSI = 0.10

# 策略关键边界保护。强制处理小箱或倒挂时仍允许跨越保护边界。
PROTECT_STRATEGY_BOUNDARIES = True
PROTECT_LARGEST_RISK_JUMPS = 1
PROTECTED_BOUNDARY_PENALTY = 100.0

# 手工兜底方案，仅在实际初始箱数为 20 时直接使用；否则自动生成连续等宽兜底范围。
FINAL_BIN_RANGES: List[Tuple[int, int]] = [
    (1, 4),
    (5, 8),
    (9, 12),
    (13, 16),
    (17, 20),
]

# 默认策略的风险约束。
STRATEGY_CONFIG = {
    "strategy_name": "默认策略",
    "objective": "平衡通过率、整体风险和边际风险",
    "auto_constraints": {
        "max_cum_1m30p_cnt_bad_rate": 0.0090,
        "max_cum_3m30p_cnt_bad_rate": 0.0550,
        "max_marginal_3m30p_cnt_bad_rate": 0.0900,
    },
    "accept_constraints": {
        "max_cum_1m30p_cnt_bad_rate": 0.0130,
        "max_cum_3m30p_cnt_bad_rate": 0.0750,
        "max_marginal_3m30p_cnt_bad_rate": 0.1700,
    },
}

RISK_NUMERIC_COLS = [
    SCORE_COL,
    "duedate_1m_30",
    "duedate_3m_30",
    "principal",
    "estimate_principal_remaining_mob1",
    "estimate_principal_remaining_mob3",
    "dpd_days_ever_mob1",
    "dpd_days_ever_mob3",
]

REQUIRED_ANALYSIS_COLS = [
    "application_id",
    "user_id",
    "application_time",
    "application_month",
    SCORE_COL,
    "duedate_1m_30",
    "duedate_3m_30",
    "principal",
    "estimate_principal_remaining_mob1",
    "estimate_principal_remaining_mob3",
    "dpd_days_ever_mob1",
    "dpd_days_ever_mob3",
]


# ============================================================
# 1. 通用工具
# ============================================================

def clean_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """清理 UTF-8 BOM 和少数 CSV 表头乱码。"""
    result = frame.copy()
    result.columns = [str(col).lstrip("\ufeff").lstrip("ï»¿") for col in result.columns]
    return result


def read_csv_clean(path: Path) -> pd.DataFrame:
    """读取 CSV，并统一清理字段名。"""
    if not path.exists():
        raise FileNotFoundError(f"输入文件不存在: {path}")
    return clean_columns(pd.read_csv(path, low_memory=False))


def require_columns(frame: pd.DataFrame, columns: Iterable[str], context: str) -> None:
    """校验 DataFrame 是否包含必要字段。"""
    missing = [col for col in columns if col not in frame.columns]
    if missing:
        raise ValueError(f"{context} 缺少必要字段: {missing}")


def safe_div(numerator, denominator):
    """安全除法；分母为 0 时返回 NaN。"""
    num = np.asarray(numerator, dtype="float64")
    den = np.asarray(denominator, dtype="float64")
    result = np.full(np.broadcast(num, den).shape, np.nan, dtype="float64")
    np.divide(num, den, out=result, where=den != 0)

    if np.ndim(result) == 0:
        return float(result)
    if isinstance(numerator, pd.Series):
        return pd.Series(result, index=numerator.index)
    if isinstance(denominator, pd.Series):
        return pd.Series(result, index=denominator.index)
    return result


def flatten_dict(prefix: str, values: Dict[str, float]) -> Dict[str, float]:
    """将策略约束字典展开为平面字段。"""
    return {f"{prefix}_{key}": value for key, value in values.items()}


# ============================================================
# 2. 数据加载与样本切分
# ============================================================

def load_analysis_data() -> pd.DataFrame:
    """加载首版分箱真正需要的数据，删除未使用的其他模型表和交易特征表。"""
    print("加载数据 ...")

    sample = read_csv_clean(DATA_DIR / SAMPLE_FILE)
    application = read_csv_clean(DATA_DIR / APPLICATION_FILE)
    score = read_csv_clean(DATA_DIR / SCORE_FILE)

    require_columns(sample, ["application_id", "user_id"], "sample")
    require_columns(application, ["application_id", "user_id"], "application_info")
    require_columns(score, ["application_id", RAW_SCORE_COL], "score")

    # application_info 只补充 sample 中不存在的字段，避免出现 _x / _y。
    join_keys = ["application_id", "user_id"]
    application_extra_cols = [
        col for col in application.columns
        if col in join_keys or col not in sample.columns
    ]
    application_dedup = application[application_extra_cols].drop_duplicates(
        subset=join_keys,
        keep="first",
    )

    data = sample.merge(application_dedup, on=join_keys, how="left")

    score_dedup = (
        score[["application_id", RAW_SCORE_COL]]
        .drop_duplicates(subset="application_id", keep="first")
        .rename(columns={RAW_SCORE_COL: SCORE_COL})
    )
    data = data.merge(score_dedup, on="application_id", how="left")

    # application_month 缺失时，根据 application_time 补充。
    if "application_time" in data.columns:
        data["application_time"] = pd.to_datetime(data["application_time"], errors="coerce")

    if "application_month" not in data.columns:
        data["application_month"] = pd.Series(pd.NA, index=data.index, dtype="string")
    else:
        data["application_month"] = data["application_month"].astype("string").str.slice(0, 7)

    if "application_time" in data.columns:
        month_from_time = data["application_time"].dt.to_period("M").astype("string")
        data["application_month"] = data["application_month"].fillna(month_from_time)

    require_columns(data, REQUIRED_ANALYSIS_COLS, "拼接后的分析数据")

    for col in RISK_NUMERIC_COLS:
        data[col] = pd.to_numeric(data[col], errors="coerce")

    # 分箱分析只使用存在模型分的样本；缺失比例会在总览中单独展示。
    source_row_count = len(data)
    score_missing_count = int(data[SCORE_COL].isna().sum())
    data.attrs["source_row_count"] = source_row_count
    data.attrs["score_missing_count"] = score_missing_count

    data = data.loc[data[SCORE_COL].notna()].copy()
    if data.empty:
        raise ValueError(f"{SCORE_COL} 全为空，无法进行分箱")

    print(
        f"   原始 {source_row_count:,} 行；有效模型分 {len(data):,} 行；"
        f"模型分缺失 {score_missing_count:,} 行"
    )
    return data


def split_train_oot(data: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """按申请月份切分 Train 和 OOT。"""
    result = data.copy()
    month = result["application_month"].astype("string")

    train_mask = month.notna() & month.le(TRAIN_END_MONTH)
    oot_mask = month.notna() & month.ge(OOT_START_MONTH)

    result["sample_group"] = np.select(
        [
            train_mask.to_numpy(dtype=bool, na_value=False),
            oot_mask.to_numpy(dtype=bool, na_value=False),
        ],
        ["train", "oot"],
        default="gap_or_unknown",
    )

    train = result.loc[result["sample_group"].eq("train")].copy()
    oot = result.loc[result["sample_group"].eq("oot")].copy()

    if train.empty:
        raise ValueError("Train 样本为空，请检查 TRAIN_END_MONTH 和 application_month")
    if oot.empty:
        raise ValueError("OOT 样本为空，请检查 OOT_START_MONTH 和 application_month")

    print(f"样本切分完成：Train {len(train):,} 行，OOT {len(oot):,} 行")
    return result, train, oot


def split_development_validation(
    train: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[str]]:
    """
    在 Train 内按月份切分 Development / Validation。

    优先使用最后 VALIDATION_MONTH_COUNT 个完整月份作为 Validation；
    如果月份数量不足，则按 application_time 的时间顺序切出最后 20%。
    OOT 不参与任何合箱方案选择。
    """
    work = train.copy()
    available_months = sorted(
        work["application_month"].dropna().astype(str).unique().tolist()
    )

    max_validation_months = max(
        1,
        len(available_months) - MIN_DEVELOPMENT_MONTH_COUNT,
    )
    validation_month_count = min(VALIDATION_MONTH_COUNT, max_validation_months)

    if len(available_months) >= MIN_DEVELOPMENT_MONTH_COUNT + 1:
        validation_months = available_months[-validation_month_count:]
        validation_mask = work["application_month"].astype(str).isin(validation_months)
        development = work.loc[~validation_mask].copy()
        validation = work.loc[validation_mask].copy()
    else:
        ordered = work.sort_values(["application_time", "application_id"]).copy()
        split_at = max(1, int(len(ordered) * 0.80))
        split_at = min(split_at, len(ordered) - 1)
        development = ordered.iloc[:split_at].copy()
        validation = ordered.iloc[split_at:].copy()
        validation_months = sorted(
            validation["application_month"].dropna().astype(str).unique().tolist()
        )

    if development.empty or validation.empty:
        raise ValueError(
            "Development 或 Validation 为空，请检查 Train 时间范围和 VALIDATION_MONTH_COUNT"
        )

    development["merge_sample_group"] = "development"
    validation["merge_sample_group"] = "validation"

    print(
        f"Train 内部切分完成：Development {len(development):,} 行，"
        f"Validation {len(validation):,} 行；Validation 月份={validation_months}"
    )
    return development, validation, validation_months


# ============================================================
# 3. 等频初分与相邻箱合并
# ============================================================

def learn_equal_freq_edges(
    data: pd.DataFrame,
    score_col: str,
    n_bins: int,
) -> np.ndarray:
    """仅在 Train 上学习等频边界，并将首尾扩展为无穷。"""
    score = pd.to_numeric(data[score_col], errors="coerce").dropna()
    if score.empty:
        raise ValueError(f"{score_col} 全为空，无法分箱")

    _, raw_edges = pd.qcut(score, q=n_bins, retbins=True, duplicates="drop")
    edges = np.unique(np.asarray(raw_edges, dtype="float64"))
    if len(edges) < 2:
        raise ValueError(f"{score_col} 唯一值不足，无法形成有效分箱")

    edges[0] = -np.inf
    edges[-1] = np.inf
    return edges


def build_initial_edge_table(edges: np.ndarray) -> pd.DataFrame:
    """生成初始分箱边界配置表。"""
    rows = []
    for idx in range(len(edges) - 1):
        order = idx + 1
        rows.append(
            {
                "bin_order": order,
                INITIAL_BIN_COL: f"B{order:02d}",
                "score_left": edges[idx],
                "score_right": edges[idx + 1],
                "interval_rule": "(left, right]",
            }
        )
    return pd.DataFrame(rows)


def apply_edges(
    data: pd.DataFrame,
    score_col: str,
    edges: np.ndarray,
    bin_col: str,
) -> pd.DataFrame:
    """将 Train 学到的边界复用到任意样本。"""
    result = data.copy()
    labels = list(range(1, len(edges)))
    cut_result = pd.cut(
        pd.to_numeric(result[score_col], errors="coerce"),
        bins=edges,
        labels=labels,
        include_lowest=True,
        right=True,
    )
    result["initial_bin_order"] = cut_result.astype("Int64")
    result[bin_col] = result["initial_bin_order"].map(
        {order: f"B{order:02d}" for order in labels}
    )
    return result


def validate_merge_ranges(
    ranges: Sequence[Tuple[int, int]],
    initial_bin_count: int,
) -> None:
    """检查合箱范围是否连续、无重叠，并覆盖所有初始箱。"""
    flattened: List[int] = []
    for start, end in ranges:
        if start > end:
            raise ValueError(f"无效合箱范围: ({start}, {end})")
        flattened.extend(range(start, end + 1))

    expected = list(range(1, initial_bin_count + 1))
    if flattened != expected:
        raise ValueError(
            "FINAL_BIN_RANGES 必须连续且完整覆盖所有初始箱。"
            f"当前实际初始箱数={initial_bin_count}，覆盖结果={flattened}"
        )


def build_merge_map(
    ranges: Sequence[Tuple[int, int]],
    initial_bin_count: int,
) -> pd.DataFrame:
    """生成初始箱到最终风险等级的映射。"""
    validate_merge_ranges(ranges, initial_bin_count)

    rows = []
    for final_order, (start, end) in enumerate(ranges, start=1):
        final_label = chr(ord("A") + final_order - 1)
        merged_from = f"B{start:02d}-B{end:02d}" if start != end else f"B{start:02d}"
        for initial_order in range(start, end + 1):
            rows.append(
                {
                    "initial_bin_order": initial_order,
                    INITIAL_BIN_COL: f"B{initial_order:02d}",
                    "final_bin_order": final_order,
                    FINAL_BIN_COL: final_label,
                    "merged_from": merged_from,
                }
            )
    return pd.DataFrame(rows)


def apply_merge_map(data: pd.DataFrame, merge_map: pd.DataFrame) -> pd.DataFrame:
    """将初始箱映射到最终风险等级。"""
    result = data.merge(
        merge_map[[INITIAL_BIN_COL, "final_bin_order", FINAL_BIN_COL]],
        on=INITIAL_BIN_COL,
        how="left",
    )
    result["bin_order"] = result["final_bin_order"].astype("Int64")
    return result


def build_final_edge_table(
    initial_edges: pd.DataFrame,
    merge_map: pd.DataFrame,
) -> pd.DataFrame:
    """生成最终风险等级的上线边界表。"""
    merged = initial_edges.merge(
        merge_map[[INITIAL_BIN_COL, "final_bin_order", FINAL_BIN_COL, "merged_from"]],
        on=INITIAL_BIN_COL,
        how="left",
    ).sort_values("bin_order")

    final_edges = (
        merged.groupby(
            ["final_bin_order", FINAL_BIN_COL, "merged_from"],
            observed=True,
        )
        .agg(
            score_left=("score_left", "first"),
            score_right=("score_right", "last"),
            source_bin_start=("bin_order", "min"),
            source_bin_end=("bin_order", "max"),
        )
        .reset_index()
        .sort_values("final_bin_order")
        .reset_index(drop=True)
    )
    final_edges["interval_rule"] = "(left, right]"
    return final_edges


# ============================================================
# 4. 风险指标计算
# ============================================================

def add_risk_helper_columns(data: pd.DataFrame) -> pd.DataFrame:
    """生成分箱统计所需的成熟、逾期和金额字段。"""
    work = data.copy()

    work["_principal"] = work["principal"].fillna(0)

    work["_m1_mature_cnt"] = work["duedate_1m_30"].isin([0, 1])
    work["_m1_bad_cnt"] = work["duedate_1m_30"].eq(1)
    work["_m3_mature_cnt"] = work["duedate_3m_30"].isin([0, 1])
    work["_m3_bad_cnt"] = work["duedate_3m_30"].eq(1)

    m1_mature_amt = work["dpd_days_ever_mob1"].notna()
    m3_mature_amt = work["dpd_days_ever_mob3"].notna()
    m1_bad_amt = m1_mature_amt & work["dpd_days_ever_mob1"].ge(30)
    m3_bad_amt = m3_mature_amt & work["dpd_days_ever_mob3"].ge(30)

    work["_m1_amt_exposure"] = np.where(m1_mature_amt, work["_principal"], 0)
    work["_m3_amt_exposure"] = np.where(m3_mature_amt, work["_principal"], 0)
    work["_m1_amt_bad"] = np.where(
        m1_bad_amt,
        work["estimate_principal_remaining_mob1"].fillna(0),
        0,
    )
    work["_m3_amt_bad"] = np.where(
        m3_bad_amt,
        work["estimate_principal_remaining_mob3"].fillna(0),
        0,
    )
    return work


def calc_bin_stats(
    data: pd.DataFrame,
    bin_col: str,
    order_col: str,
    score_col: str = SCORE_COL,
) -> pd.DataFrame:
    """
    按分箱计算核心指标。

    每个率均保留可在 Excel 中复算的分子和分母：
    - 笔数逾期率 = bad / mature；
    - 金额逾期率 = bad amount / exposure；
    - 样本占比 = n / total_n。
    """
    required = [
        "application_id",
        bin_col,
        order_col,
        score_col,
        *RISK_NUMERIC_COLS[1:],
    ]
    require_columns(data, required, "calc_bin_stats")

    work = add_risk_helper_columns(data)
    stats = (
        work.groupby([bin_col, order_col], dropna=False, observed=True)
        .agg(
            n=("application_id", "count"),
            application_id_nunique=("application_id", "nunique"),
            principal_amt=("_principal", "sum"),
            score_min=(score_col, "min"),
            score_max=(score_col, "max"),
            score_mean=(score_col, "mean"),
            **{
                "1m30p_cnt_mature": ("_m1_mature_cnt", "sum"),
                "1m30p_cnt_bad": ("_m1_bad_cnt", "sum"),
                "3m30p_cnt_mature": ("_m3_mature_cnt", "sum"),
                "3m30p_cnt_bad": ("_m3_bad_cnt", "sum"),
                "1m30p_amt_exposure": ("_m1_amt_exposure", "sum"),
                "1m30p_amt_bad": ("_m1_amt_bad", "sum"),
                "3m30p_amt_exposure": ("_m3_amt_exposure", "sum"),
                "3m30p_amt_bad": ("_m3_amt_bad", "sum"),
            },
        )
        .reset_index()
        .rename(columns={order_col: "bin_order"})
        .sort_values("bin_order")
        .reset_index(drop=True)
    )

    total_n = stats["n"].sum()
    stats["total_n"] = total_n
    stats["sample_pct"] = safe_div(stats["n"], total_n)

    for prefix in ["1m30p", "3m30p"]:
        stats[f"{prefix}_cnt_good"] = (
            stats[f"{prefix}_cnt_mature"] - stats[f"{prefix}_cnt_bad"]
        )
        stats[f"{prefix}_cnt_bad_rate"] = safe_div(
            stats[f"{prefix}_cnt_bad"],
            stats[f"{prefix}_cnt_mature"],
        )
        stats[f"{prefix}_amt_bad_rate"] = safe_div(
            stats[f"{prefix}_amt_bad"],
            stats[f"{prefix}_amt_exposure"],
        )

    # 按低风险到高风险累计，直接用于阈值分析。
    stats["cum_n"] = stats["n"].cumsum()
    stats["cum_pass_rate"] = safe_div(stats["cum_n"], total_n)

    for prefix in ["1m30p", "3m30p"]:
        stats[f"cum_{prefix}_cnt_mature"] = stats[f"{prefix}_cnt_mature"].cumsum()
        stats[f"cum_{prefix}_cnt_bad"] = stats[f"{prefix}_cnt_bad"].cumsum()
        stats[f"cum_{prefix}_cnt_bad_rate"] = safe_div(
            stats[f"cum_{prefix}_cnt_bad"],
            stats[f"cum_{prefix}_cnt_mature"],
        )

        stats[f"cum_{prefix}_amt_exposure"] = stats[f"{prefix}_amt_exposure"].cumsum()
        stats[f"cum_{prefix}_amt_bad"] = stats[f"{prefix}_amt_bad"].cumsum()
        stats[f"cum_{prefix}_amt_bad_rate"] = safe_div(
            stats[f"cum_{prefix}_amt_bad"],
            stats[f"cum_{prefix}_amt_exposure"],
        )

    return stats


def check_monotonicity(
    stats: pd.DataFrame,
    rate_cols: Sequence[str],
    sample_group: str,
) -> pd.DataFrame:
    """检查风险率是否随风险等级非递减。"""
    ordered = stats.sort_values("bin_order").reset_index(drop=True)
    rows = []

    for rate_col in rate_cols:
        diff = ordered[rate_col].diff()
        violation = diff.lt(0).fillna(False)
        rows.append(
            {
                "sample_group": sample_group,
                "metric": rate_col,
                "is_monotonic_non_decreasing": not bool(violation.any()),
                "violation_cnt": int(violation.sum()),
                "violation_bins": ",".join(
                    ordered.loc[violation, "bin_order"].astype(str).tolist()
                ),
            }
        )
    return pd.DataFrame(rows)


def format_merge_ranges(ranges: Sequence[Tuple[int, int]]) -> str:
    """将合箱范围格式化为便于报告阅读和复用的字符串。"""
    return "[" + ", ".join(f"({start},{end})" for start, end in ranges) + "]"


def parse_merge_ranges(text: str) -> List[Tuple[int, int]]:
    """将候选表中的范围字符串还原为整数区间。"""
    parsed = ast.literal_eval(str(text))
    return [(int(start), int(end)) for start, end in parsed]


def make_equal_contiguous_ranges(
    initial_bin_count: int,
    final_bin_count: int,
) -> List[Tuple[int, int]]:
    """生成连续、尽量等宽的手工兜底合箱范围。"""
    final_bin_count = max(1, min(final_bin_count, initial_bin_count))
    groups = np.array_split(np.arange(1, initial_bin_count + 1), final_bin_count)
    return [(int(group[0]), int(group[-1])) for group in groups if len(group) > 0]


def calc_complete_initial_stats(
    data: pd.DataFrame,
    initial_edges: pd.DataFrame,
) -> pd.DataFrame:
    """
    计算完整的初始箱统计表。

    即使 Validation 某个初始箱没有样本，也保留该箱，避免候选范围聚合错位。
    """
    stats = calc_bin_stats(
        data,
        bin_col=INITIAL_BIN_COL,
        order_col="initial_bin_order",
    )
    stats = stats.loc[stats[INITIAL_BIN_COL].notna()].copy()

    edge_cols = [
        "bin_order",
        INITIAL_BIN_COL,
        "score_left",
        "score_right",
        "interval_rule",
    ]
    result = initial_edges[edge_cols].merge(
        stats.drop(columns=["score_left", "score_right", "interval_rule"], errors="ignore"),
        on=["bin_order", INITIAL_BIN_COL],
        how="left",
    )

    zero_cols = [
        "n",
        "application_id_nunique",
        "principal_amt",
        "1m30p_cnt_mature",
        "1m30p_cnt_bad",
        "1m30p_cnt_good",
        "3m30p_cnt_mature",
        "3m30p_cnt_bad",
        "3m30p_cnt_good",
        "1m30p_amt_exposure",
        "1m30p_amt_bad",
        "3m30p_amt_exposure",
        "3m30p_amt_bad",
    ]
    for col in zero_cols:
        if col not in result.columns:
            result[col] = 0.0
        result[col] = pd.to_numeric(result[col], errors="coerce").fillna(0)

    total_n = float(result["n"].sum())
    result["total_n"] = total_n
    result["sample_pct"] = safe_div(result["n"], total_n)

    for prefix in ["1m30p", "3m30p"]:
        result[f"{prefix}_cnt_good"] = (
            result[f"{prefix}_cnt_mature"] - result[f"{prefix}_cnt_bad"]
        )
        result[f"{prefix}_cnt_bad_rate"] = safe_div(
            result[f"{prefix}_cnt_bad"],
            result[f"{prefix}_cnt_mature"],
        )
        result[f"{prefix}_amt_bad_rate"] = safe_div(
            result[f"{prefix}_amt_bad"],
            result[f"{prefix}_amt_exposure"],
        )

    result = result.sort_values("bin_order").reset_index(drop=True)
    result["cum_n"] = result["n"].cumsum()
    result["cum_pass_rate"] = safe_div(result["cum_n"], total_n)

    for prefix in ["1m30p", "3m30p"]:
        result[f"cum_{prefix}_cnt_mature"] = result[f"{prefix}_cnt_mature"].cumsum()
        result[f"cum_{prefix}_cnt_bad"] = result[f"{prefix}_cnt_bad"].cumsum()
        result[f"cum_{prefix}_cnt_bad_rate"] = safe_div(
            result[f"cum_{prefix}_cnt_bad"],
            result[f"cum_{prefix}_cnt_mature"],
        )
        result[f"cum_{prefix}_amt_exposure"] = result[f"{prefix}_amt_exposure"].cumsum()
        result[f"cum_{prefix}_amt_bad"] = result[f"{prefix}_amt_bad"].cumsum()
        result[f"cum_{prefix}_amt_bad_rate"] = safe_div(
            result[f"cum_{prefix}_amt_bad"],
            result[f"cum_{prefix}_amt_exposure"],
        )

    return result


def aggregate_initial_stats_by_ranges(
    initial_stats: pd.DataFrame,
    ranges: Sequence[Tuple[int, int]],
) -> pd.DataFrame:
    """按候选合箱范围聚合初始箱风险表现。"""
    stats = initial_stats.sort_values("bin_order").reset_index(drop=True)
    total_n = float(stats["n"].sum())

    rows = []
    for final_order, (start, end) in enumerate(ranges, start=1):
        part = stats.loc[stats["bin_order"].between(start, end, inclusive="both")]
        if part.empty:
            raise ValueError(f"候选合箱范围 ({start}, {end}) 未匹配任何初始箱")

        n = float(part["n"].sum())
        row = {
            "final_bin_order": final_order,
            FINAL_BIN_COL: chr(ord("A") + final_order - 1),
            "merged_from": f"B{start:02d}-B{end:02d}" if start != end else f"B{start:02d}",
            "source_bin_start": start,
            "source_bin_end": end,
            "n": n,
            "principal_amt": float(part["principal_amt"].sum()),
            "score_left": part["score_left"].iloc[0],
            "score_right": part["score_right"].iloc[-1],
            "score_min": part["score_min"].min(),
            "score_max": part["score_max"].max(),
            "score_mean": safe_div((part["score_mean"] * part["n"]).sum(), n),
        }

        for prefix in ["1m30p", "3m30p"]:
            row[f"{prefix}_cnt_mature"] = float(part[f"{prefix}_cnt_mature"].sum())
            row[f"{prefix}_cnt_bad"] = float(part[f"{prefix}_cnt_bad"].sum())
            row[f"{prefix}_cnt_good"] = (
                row[f"{prefix}_cnt_mature"] - row[f"{prefix}_cnt_bad"]
            )
            row[f"{prefix}_amt_exposure"] = float(part[f"{prefix}_amt_exposure"].sum())
            row[f"{prefix}_amt_bad"] = float(part[f"{prefix}_amt_bad"].sum())
            row[f"{prefix}_cnt_bad_rate"] = safe_div(
                row[f"{prefix}_cnt_bad"],
                row[f"{prefix}_cnt_mature"],
            )
            row[f"{prefix}_amt_bad_rate"] = safe_div(
                row[f"{prefix}_amt_bad"],
                row[f"{prefix}_amt_exposure"],
            )

        row["sample_pct"] = safe_div(n, total_n)
        rows.append(row)

    result = pd.DataFrame(rows)
    result["cum_n"] = result["n"].cumsum()
    result["cum_pass_rate"] = safe_div(result["cum_n"], total_n)

    for prefix in ["1m30p", "3m30p"]:
        result[f"cum_{prefix}_cnt_mature"] = result[f"{prefix}_cnt_mature"].cumsum()
        result[f"cum_{prefix}_cnt_bad"] = result[f"{prefix}_cnt_bad"].cumsum()
        result[f"cum_{prefix}_cnt_bad_rate"] = safe_div(
            result[f"cum_{prefix}_cnt_bad"],
            result[f"cum_{prefix}_cnt_mature"],
        )
        result[f"cum_{prefix}_amt_exposure"] = result[f"{prefix}_amt_exposure"].cumsum()
        result[f"cum_{prefix}_amt_bad"] = result[f"{prefix}_amt_bad"].cumsum()
        result[f"cum_{prefix}_amt_bad_rate"] = safe_div(
            result[f"cum_{prefix}_amt_bad"],
            result[f"cum_{prefix}_amt_exposure"],
        )

    return result


def oriented_rate(values: pd.Series) -> pd.Series:
    """统一转换为随风险等级应非递减的方向。"""
    numeric = pd.to_numeric(values, errors="coerce")
    return numeric if HIGH_SCORE_HIGH_RISK else -numeric


def count_rate_inversions(
    stats: pd.DataFrame,
    rate_cols: Sequence[str],
    tolerance: float = 0.0,
) -> int:
    """统计风险率相邻显著倒挂次数。"""
    total = 0
    ordered = stats.sort_values("final_bin_order")
    for rate_col in rate_cols:
        diff = oriented_rate(ordered[rate_col]).diff()
        total += int(diff.lt(-tolerance).fillna(False).sum())
    return total


def max_rate_drop(
    stats: pd.DataFrame,
    rate_cols: Sequence[str],
) -> float:
    """计算所有风险率中最大的相邻倒挂幅度。"""
    ordered = stats.sort_values("final_bin_order")
    drops = []
    for rate_col in rate_cols:
        diff = oriented_rate(ordered[rate_col]).diff()
        value = (-diff).clip(lower=0).max()
        drops.append(float(value) if pd.notna(value) else 0.0)
    return max(drops) if drops else 0.0


def calc_psi_from_bin_stats(
    base_stats: pd.DataFrame,
    compare_stats: pd.DataFrame,
    eps: float = 1e-6,
) -> float:
    """用同一候选箱的样本占比计算 PSI。"""
    base_pct = base_stats["sample_pct"].fillna(0).to_numpy(dtype="float64") + eps
    compare_pct = compare_stats["sample_pct"].fillna(0).to_numpy(dtype="float64") + eps
    return float(((compare_pct - base_pct) * np.log(compare_pct / base_pct)).sum())


def calc_iv_from_stats(
    stats: pd.DataFrame,
    bad_col: str = PRIMARY_BAD_COL,
    good_col: str = PRIMARY_GOOD_COL,
    eps: float = 0.5,
) -> float:
    """使用箱级好坏样本量计算 IV。"""
    bad = pd.to_numeric(stats[bad_col], errors="coerce").fillna(0).to_numpy(float)
    good = pd.to_numeric(stats[good_col], errors="coerce").fillna(0).to_numpy(float)
    if bad.sum() <= 0 or good.sum() <= 0:
        return np.nan

    bad_dist = (bad + eps) / (bad.sum() + eps * len(bad))
    good_dist = (good + eps) / (good.sum() + eps * len(good))
    return float(np.sum((bad_dist - good_dist) * np.log(bad_dist / good_dist)))


def two_proportion_pvalue(
    bad_1: float,
    mature_1: float,
    bad_2: float,
    mature_2: float,
) -> float:
    """不依赖 scipy 的双侧两比例 Z 检验 p 值。"""
    if mature_1 <= 0 or mature_2 <= 0:
        return np.nan

    p1 = bad_1 / mature_1
    p2 = bad_2 / mature_2
    pooled = (bad_1 + bad_2) / (mature_1 + mature_2)
    variance = pooled * (1 - pooled) * (1 / mature_1 + 1 / mature_2)
    if variance <= 0:
        return 1.0 if math.isclose(p1, p2) else 0.0

    z_value = abs(p1 - p2) / math.sqrt(variance)
    normal_cdf = 0.5 * (1 + math.erf(z_value / math.sqrt(2)))
    return float(2 * (1 - normal_cdf))


def required_sample_pct(position: int, bin_count: int) -> float:
    """头尾风险箱使用较低样本占比要求，中间箱使用标准要求。"""
    if position in {0, bin_count - 1}:
        return MIN_TAIL_BIN_SAMPLE_PCT
    return MIN_MIDDLE_BIN_SAMPLE_PCT


def calc_bin_constraint_details(stats: pd.DataFrame) -> pd.DataFrame:
    """计算每个最终箱的样本、成熟量和好坏样本约束。"""
    rows = []
    ordered = stats.sort_values("final_bin_order").reset_index(drop=True)
    for position, row in ordered.iterrows():
        min_sample_pct = required_sample_pct(position, len(ordered))
        checks = {
            "sample_ok": row["sample_pct"] >= min_sample_pct,
            "mature_ok": row[PRIMARY_MATURE_COL] >= MIN_FINAL_BIN_MATURE_COUNT,
            "bad_ok": row[PRIMARY_BAD_COL] >= MIN_FINAL_BIN_BAD_COUNT,
            "good_ok": row[PRIMARY_GOOD_COL] >= MIN_FINAL_BIN_GOOD_COUNT,
        }
        severity = 0.0
        severity += max(0.0, 1 - safe_div(row["sample_pct"], min_sample_pct))
        severity += max(0.0, 1 - safe_div(row[PRIMARY_MATURE_COL], MIN_FINAL_BIN_MATURE_COUNT))
        severity += max(0.0, 1 - safe_div(row[PRIMARY_BAD_COL], MIN_FINAL_BIN_BAD_COUNT))
        severity += max(0.0, 1 - safe_div(row[PRIMARY_GOOD_COL], MIN_FINAL_BIN_GOOD_COUNT))

        rows.append(
            {
                "final_bin_order": int(row["final_bin_order"]),
                FINAL_BIN_COL: row[FINAL_BIN_COL],
                "required_sample_pct": min_sample_pct,
                **checks,
                "all_constraints_ok": all(checks.values()),
                "violation_severity": severity,
            }
        )
    return pd.DataFrame(rows)


def identify_protected_boundaries(
    initial_stats: pd.DataFrame,
    config: Dict,
) -> Set[int]:
    """
    找出应尽量保留的初始箱边界。

    边界编号 k 代表 Bk 与 B(k+1) 之间的切点。
    """
    if not PROTECT_STRATEGY_BOUNDARIES:
        return set()

    ordered = initial_stats.sort_values("bin_order").reset_index(drop=True)
    boundaries: Set[int] = set()
    max_boundary = len(ordered) - 1

    for constraint_group in ["auto_constraints", "accept_constraints"]:
        constraints = config[constraint_group]

        cum_limit = constraints.get("max_cum_3m30p_cnt_bad_rate")
        if cum_limit is not None:
            eligible = ordered.loc[ordered["cum_3m30p_cnt_bad_rate"].le(cum_limit)]
            if not eligible.empty:
                boundary = int(eligible["bin_order"].max())
                if 1 <= boundary <= max_boundary:
                    boundaries.add(boundary)

        marginal_limit = constraints.get("max_marginal_3m30p_cnt_bad_rate")
        if marginal_limit is not None:
            above = ordered.loc[ordered[PRIMARY_RATE_COL].gt(marginal_limit)]
            if not above.empty:
                boundary = int(above["bin_order"].min()) - 1
                if 1 <= boundary <= max_boundary:
                    boundaries.add(boundary)

    if PROTECT_LARGEST_RISK_JUMPS > 0:
        oriented = oriented_rate(ordered[PRIMARY_RATE_COL])
        jumps = oriented.diff().dropna().sort_values(ascending=False)
        for idx in jumps.head(PROTECT_LARGEST_RISK_JUMPS).index:
            boundary = int(ordered.loc[idx, "bin_order"]) - 1
            if 1 <= boundary <= max_boundary:
                boundaries.add(boundary)

    return boundaries


def merge_ranges_at(
    ranges: Sequence[Tuple[int, int]],
    pair_index: int,
) -> List[Tuple[int, int]]:
    """合并 ranges[pair_index] 与其右侧相邻范围。"""
    if pair_index < 0 or pair_index >= len(ranges) - 1:
        raise IndexError(f"无效相邻合箱位置: {pair_index}")
    result = list(ranges)
    left = result[pair_index]
    right = result[pair_index + 1]
    result[pair_index:pair_index + 2] = [(left[0], right[1])]
    return result


def pair_merge_diagnostics(
    current_stats: pd.DataFrame,
    ranges: Sequence[Tuple[int, int]],
    pair_index: int,
    initial_stats: pd.DataFrame,
    protected_boundaries: Set[int],
    ignore_protection: bool = False,
) -> Dict[str, float]:
    """计算合并某对相邻箱的风险差异、显著性、IV 损失和综合代价。"""
    left = current_stats.iloc[pair_index]
    right = current_stats.iloc[pair_index + 1]

    left_rate = left[PRIMARY_RATE_COL]
    right_rate = right[PRIMARY_RATE_COL]
    if pd.isna(left_rate) or pd.isna(right_rate):
        rate_gap = 0.0
    else:
        rate_gap = abs(float(left_rate) - float(right_rate))

    p_value = two_proportion_pvalue(
        left[PRIMARY_BAD_COL],
        left[PRIMARY_MATURE_COL],
        right[PRIMARY_BAD_COL],
        right[PRIMARY_MATURE_COL],
    )
    p_for_cost = 0.0 if pd.isna(p_value) else p_value

    current_iv = calc_iv_from_stats(current_stats)
    merged_ranges = merge_ranges_at(ranges, pair_index)
    merged_stats = aggregate_initial_stats_by_ranges(initial_stats, merged_ranges)
    merged_iv = calc_iv_from_stats(merged_stats)
    iv_loss = 0.0
    if pd.notna(current_iv) and pd.notna(merged_iv):
        iv_loss = max(0.0, float(current_iv - merged_iv))

    boundary = int(ranges[pair_index][1])
    is_protected = boundary in protected_boundaries
    protection_penalty = (
        0.0
        if ignore_protection or not is_protected
        else PROTECTED_BOUNDARY_PENALTY
    )

    # 风险越接近、差异越不显著、IV 损失越小，越优先合并。
    cost = (
        rate_gap * 100
        + (1 - p_for_cost)
        + iv_loss * 10
        + protection_penalty
    )
    return {
        "pair_index": pair_index,
        "boundary": boundary,
        "left_rate": left_rate,
        "right_rate": right_rate,
        "abs_rate_diff": rate_gap,
        "p_value": p_value,
        "iv_loss": iv_loss,
        "is_protected_boundary": is_protected,
        "merge_cost": cost,
    }


def choose_best_adjacent_pair(
    ranges: Sequence[Tuple[int, int]],
    initial_stats: pd.DataFrame,
    protected_boundaries: Set[int],
    allowed_pair_indices: Optional[Sequence[int]] = None,
    ignore_protection: bool = False,
) -> Dict[str, float]:
    """从允许的相邻箱中选择综合代价最低的一对。"""
    current_stats = aggregate_initial_stats_by_ranges(initial_stats, ranges)
    pair_indices = (
        list(allowed_pair_indices)
        if allowed_pair_indices is not None
        else list(range(len(ranges) - 1))
    )
    if not pair_indices:
        raise ValueError("没有可合并的相邻箱")

    diagnostics = [
        pair_merge_diagnostics(
            current_stats,
            ranges,
            pair_index,
            initial_stats,
            protected_boundaries,
            ignore_protection=ignore_protection,
        )
        for pair_index in pair_indices
    ]
    return min(diagnostics, key=lambda item: item["merge_cost"])


def primary_inversion_pair_indices(stats: pd.DataFrame) -> List[int]:
    """返回主风险指标发生倒挂的相邻箱左侧位置。"""
    ordered = stats.sort_values("final_bin_order").reset_index(drop=True)
    diff = oriented_rate(ordered[PRIMARY_RATE_COL]).diff()
    violation_rows = diff.lt(-DEVELOPMENT_INVERSION_TOLERANCE).fillna(False)
    return [int(row_index - 1) for row_index in ordered.index[violation_rows] if row_index > 0]


def evaluate_merge_candidate(
    development_initial_stats: pd.DataFrame,
    validation_initial_stats: pd.DataFrame,
    ranges: Sequence[Tuple[int, int]],
    initial_iv: float,
    step_no: int,
    stage: str,
    merge_reason: str,
) -> Dict[str, object]:
    """计算一个候选合箱方案的完整评分指标。"""
    development_stats = aggregate_initial_stats_by_ranges(development_initial_stats, ranges)
    validation_stats = aggregate_initial_stats_by_ranges(validation_initial_stats, ranges)

    rate_cols = [
        "1m30p_cnt_bad_rate",
        "3m30p_cnt_bad_rate",
        "1m30p_amt_bad_rate",
        "3m30p_amt_bad_rate",
    ]
    development_primary_inversions = count_rate_inversions(
        development_stats,
        [PRIMARY_RATE_COL],
        tolerance=DEVELOPMENT_INVERSION_TOLERANCE,
    )
    validation_primary_inversions = count_rate_inversions(
        validation_stats,
        [PRIMARY_RATE_COL],
        tolerance=VALIDATION_INVERSION_TOLERANCE,
    )
    validation_all_inversions = count_rate_inversions(
        validation_stats,
        rate_cols,
        tolerance=VALIDATION_INVERSION_TOLERANCE,
    )

    constraint_details = calc_bin_constraint_details(development_stats)
    constraint_violation_count = int((~constraint_details["all_constraints_ok"]).sum())

    validation_psi = calc_psi_from_bin_stats(development_stats, validation_stats)
    final_iv = calc_iv_from_stats(development_stats)
    iv_retention = safe_div(final_iv, initial_iv)

    adjacent_diffs = oriented_rate(development_stats[PRIMARY_RATE_COL]).diff().dropna()
    min_adjacent_rate_diff = (
        float(adjacent_diffs.min()) if not adjacent_diffs.empty else np.nan
    )

    compare = development_stats[["final_bin_order", PRIMARY_RATE_COL]].merge(
        validation_stats[["final_bin_order", PRIMARY_RATE_COL]],
        on="final_bin_order",
        suffixes=("_development", "_validation"),
    ).dropna()
    rank_correlation = (
        compare[f"{PRIMARY_RATE_COL}_development"].corr(
            compare[f"{PRIMARY_RATE_COL}_validation"],
            method="spearman",
        )
        if len(compare) >= 2
        else np.nan
    )

    strategy_metrics = summarize_strategy_from_candidate_stats(
        development_stats,
        STRATEGY_CONFIG,
    )

    final_bin_count = len(ranges)
    eligible_bin_count = MIN_FINAL_BIN_COUNT <= final_bin_count <= MAX_FINAL_BIN_COUNT
    hard_constraints_ok = all(
        [
            eligible_bin_count,
            development_primary_inversions == 0,
            constraint_violation_count == 0,
        ]
    )
    validation_psi_ok = validation_psi <= MAX_ACCEPTABLE_VALIDATION_PSI

    rank_value = 0.0 if pd.isna(rank_correlation) else float(rank_correlation)
    iv_value = 0.0 if pd.isna(iv_retention) else float(np.clip(iv_retention, 0, 1.5))
    min_sep_value = 0.0 if pd.isna(min_adjacent_rate_diff) else max(0.0, min_adjacent_rate_diff)
    accepted_rate = strategy_metrics.get("accepted_rate", np.nan)
    accepted_rate_value = 0.0 if pd.isna(accepted_rate) else float(accepted_rate)

    candidate_score = (
        100.0 * int(hard_constraints_ok)
        - 30.0 * development_primary_inversions
        - 12.0 * validation_primary_inversions
        - 4.0 * validation_all_inversions
        - 15.0 * constraint_violation_count
        - 150.0 * max(0.0, validation_psi - PREFERRED_MAX_VALIDATION_PSI)
        + 12.0 * iv_value
        + 5.0 * rank_value
        + 100.0 * min_sep_value
        + 2.0 * accepted_rate_value
        - 1.5 * abs(final_bin_count - TARGET_FINAL_BIN_COUNT)
    )

    return {
        "selected": False,
        "step_no": step_no,
        "stage": stage,
        "merge_reason": merge_reason,
        "hard_constraints_ok": hard_constraints_ok,
        "eligible_bin_count": eligible_bin_count,
        "final_bin_count": final_bin_count,
        "ranges": format_merge_ranges(ranges),
        "development_primary_inversion_cnt": development_primary_inversions,
        "validation_primary_inversion_cnt": validation_primary_inversions,
        "validation_all_inversion_cnt": validation_all_inversions,
        "validation_max_rate_drop": max_rate_drop(validation_stats, rate_cols),
        "constraint_violation_count": constraint_violation_count,
        "min_development_sample_pct": float(development_stats["sample_pct"].min()),
        "min_development_mature_count": float(development_stats[PRIMARY_MATURE_COL].min()),
        "min_development_bad_count": float(development_stats[PRIMARY_BAD_COL].min()),
        "min_development_good_count": float(development_stats[PRIMARY_GOOD_COL].min()),
        "validation_psi": validation_psi,
        "preferred_validation_psi_ok": validation_psi <= PREFERRED_MAX_VALIDATION_PSI,
        "validation_psi_acceptable_ok": validation_psi_ok,
        "primary_iv": final_iv,
        "primary_iv_retention": iv_retention,
        "development_validation_rank_corr": rank_correlation,
        "min_adjacent_primary_rate_diff": min_adjacent_rate_diff,
        "candidate_score": candidate_score,
        **strategy_metrics,
    }


def summarize_strategy_from_candidate_stats(
    final_stats: pd.DataFrame,
    config: Dict,
) -> Dict[str, float]:
    """在候选合箱累计曲线上快速评估策略通过率和风险。"""
    curve = final_stats.sort_values("final_bin_order").reset_index(drop=True).copy()
    curve["threshold_order"] = curve["final_bin_order"]
    curve["marginal_sample_pct"] = curve["sample_pct"]
    curve["marginal_3m30p_cnt_bad_rate"] = curve["3m30p_cnt_bad_rate"]

    def choose(constraints: Dict[str, float]) -> Optional[pd.Series]:
        eligible = curve.copy()
        for constraint_name, maximum in constraints.items():
            metric = constraint_name.removeprefix("max_")
            if metric not in eligible.columns:
                continue
            eligible = eligible.loc[eligible[metric].le(maximum)]
        if eligible.empty:
            return None
        return eligible.sort_values(
            ["cum_pass_rate", "threshold_order"],
            ascending=[False, False],
        ).iloc[0]

    auto_row = choose(config["auto_constraints"])
    accept_row = choose(config["accept_constraints"])
    if auto_row is None or accept_row is None:
        return {
            "strategy_status": "无满足约束的阈值",
            "auto_pass_rate": np.nan,
            "accepted_rate": np.nan,
            "manual_review_rate": np.nan,
            "reject_rate": np.nan,
            "accepted_1m30p_cnt_bad_rate": np.nan,
            "accepted_3m30p_cnt_bad_rate": np.nan,
            "last_accepted_marginal_3m30p_cnt_bad_rate": np.nan,
        }

    if accept_row["threshold_order"] < auto_row["threshold_order"]:
        accept_row = auto_row

    return {
        "strategy_status": "OK",
        "auto_pass_rate": auto_row["cum_pass_rate"],
        "accepted_rate": accept_row["cum_pass_rate"],
        "manual_review_rate": accept_row["cum_pass_rate"] - auto_row["cum_pass_rate"],
        "reject_rate": 1 - accept_row["cum_pass_rate"],
        "accepted_1m30p_cnt_bad_rate": accept_row["cum_1m30p_cnt_bad_rate"],
        "accepted_3m30p_cnt_bad_rate": accept_row["cum_3m30p_cnt_bad_rate"],
        "last_accepted_marginal_3m30p_cnt_bad_rate": accept_row[
            "marginal_3m30p_cnt_bad_rate"
        ],
    }


def build_merge_candidate_score_table(
    development_initial_stats: pd.DataFrame,
    validation_initial_stats: pd.DataFrame,
    initial_bin_count: int,
    config: Dict,
) -> Tuple[pd.DataFrame, pd.DataFrame, Set[int]]:
    """
    按可解释的分阶段流程生成候选合箱方案。

    阶段一：清理样本、成熟量、坏样本或好样本不足的箱；
    阶段二：使用主指标 3M30+ 执行 PAVA 风格单调合并；
    阶段三：按相邻风险差异、显著性、IV 损失和策略边界保护压缩到 6~8 档；
    阶段四：继续生成 8、7、6 档候选，由 Development + Validation 评分选择。
    """
    ranges: List[Tuple[int, int]] = [(idx, idx) for idx in range(1, initial_bin_count + 1)]
    protected_boundaries = identify_protected_boundaries(development_initial_stats, config)
    initial_iv = calc_iv_from_stats(development_initial_stats)

    candidate_rows: List[Dict[str, object]] = []
    step_rows: List[Dict[str, object]] = []
    step_no = 0

    def perform_merge(
        pair_index: int,
        stage: str,
        reason: str,
        diagnostics: Dict[str, float],
    ) -> None:
        nonlocal ranges, step_no
        before_ranges = list(ranges)
        left_range = ranges[pair_index]
        right_range = ranges[pair_index + 1]
        ranges = merge_ranges_at(ranges, pair_index)
        step_no += 1

        step_rows.append(
            {
                "step_no": step_no,
                "stage": stage,
                "merge_reason": reason,
                "left_range": str(left_range),
                "right_range": str(right_range),
                "merged_range": str(ranges[pair_index]),
                "boundary": diagnostics.get("boundary"),
                "is_protected_boundary": diagnostics.get("is_protected_boundary"),
                "left_primary_rate": diagnostics.get("left_rate"),
                "right_primary_rate": diagnostics.get("right_rate"),
                "abs_primary_rate_diff": diagnostics.get("abs_rate_diff"),
                "two_proportion_p_value": diagnostics.get("p_value"),
                "primary_iv_loss": diagnostics.get("iv_loss"),
                "before_ranges": format_merge_ranges(before_ranges),
                "after_ranges": format_merge_ranges(ranges),
                "after_bin_count": len(ranges),
            }
        )
        candidate_rows.append(
            evaluate_merge_candidate(
                development_initial_stats,
                validation_initial_stats,
                ranges,
                initial_iv,
                step_no,
                stage,
                reason,
            )
        )

    # 0. 初始状态仅用于过程记录。
    candidate_rows.append(
        evaluate_merge_candidate(
            development_initial_stats,
            validation_initial_stats,
            ranges,
            initial_iv,
            step_no,
            "initial",
            "20 等频初始箱",
        )
    )

    # 1. 小箱清理。
    while len(ranges) > MIN_FINAL_BIN_COUNT:
        current_stats = aggregate_initial_stats_by_ranges(development_initial_stats, ranges)
        constraints = calc_bin_constraint_details(current_stats)
        violating = constraints.loc[~constraints["all_constraints_ok"]]
        if violating.empty:
            break

        target_order = int(
            violating.sort_values("violation_severity", ascending=False).iloc[0][
                "final_bin_order"
            ]
        )
        target_index = target_order - 1
        allowed_pairs = []
        if target_index > 0:
            allowed_pairs.append(target_index - 1)
        if target_index < len(ranges) - 1:
            allowed_pairs.append(target_index)

        diagnostics = choose_best_adjacent_pair(
            ranges,
            development_initial_stats,
            protected_boundaries,
            allowed_pair_indices=allowed_pairs,
            ignore_protection=True,
        )
        perform_merge(
            int(diagnostics["pair_index"]),
            "small_bin_cleanup",
            "样本占比、成熟量或好坏样本量不足",
            diagnostics,
        )

    # 2. PAVA 风格主指标单调合并。
    while len(ranges) > MIN_FINAL_BIN_COUNT:
        current_stats = aggregate_initial_stats_by_ranges(development_initial_stats, ranges)
        inversion_pairs = primary_inversion_pair_indices(current_stats)
        if not inversion_pairs:
            break

        # 从倒挂最严重的一对开始处理。
        oriented = oriented_rate(current_stats[PRIMARY_RATE_COL])
        pair_index = min(
            inversion_pairs,
            key=lambda idx: oriented.iloc[idx + 1] - oriented.iloc[idx],
        )
        diagnostics = choose_best_adjacent_pair(
            ranges,
            development_initial_stats,
            protected_boundaries,
            allowed_pair_indices=[pair_index],
            ignore_protection=True,
        )
        perform_merge(
            pair_index,
            "pava_monotonic_merge",
            "主指标 3M30+ 出现相邻倒挂",
            diagnostics,
        )

    # 3. 如果档位仍多于上限，强制压缩到 MAX_FINAL_BIN_COUNT。
    while len(ranges) > MAX_FINAL_BIN_COUNT:
        diagnostics = choose_best_adjacent_pair(
            ranges,
            development_initial_stats,
            protected_boundaries,
        )
        perform_merge(
            int(diagnostics["pair_index"]),
            "granularity_reduction",
            "档位数量超过上限，选择信息损失最小的相邻箱",
            diagnostics,
        )

    # 4. 继续生成 7 档和 6 档候选。
    while len(ranges) > MIN_FINAL_BIN_COUNT:
        current_stats = aggregate_initial_stats_by_ranges(development_initial_stats, ranges)
        all_diagnostics = [
            pair_merge_diagnostics(
                current_stats,
                ranges,
                pair_index,
                development_initial_stats,
                protected_boundaries,
            )
            for pair_index in range(len(ranges) - 1)
        ]

        statistically_similar = [
            item
            for item in all_diagnostics
            if (
                (pd.notna(item["p_value"]) and item["p_value"] >= ADJACENT_PVALUE_TO_MERGE)
                or item["abs_rate_diff"] <= MIN_ADJACENT_ABS_RATE_DIFF
            )
        ]
        diagnostics = min(
            statistically_similar or all_diagnostics,
            key=lambda item: item["merge_cost"],
        )
        reason = (
            "相邻风险差异不显著或风险率接近"
            if statistically_similar
            else "生成更精简候选档位，选择信息损失最小的相邻箱"
        )
        perform_merge(
            int(diagnostics["pair_index"]),
            "candidate_reduction",
            reason,
            diagnostics,
        )

    candidates = pd.DataFrame(candidate_rows).drop_duplicates(subset=["ranges"], keep="last")
    candidates["target_bin_distance"] = (
        candidates["final_bin_count"] - TARGET_FINAL_BIN_COUNT
    ).abs()

    eligible = candidates.loc[candidates["eligible_bin_count"]].copy()
    selection_pool = eligible if not eligible.empty else candidates.copy()
    selection_pool = selection_pool.sort_values(
        [
            "hard_constraints_ok",
            "development_primary_inversion_cnt",
            "constraint_violation_count",
            "validation_primary_inversion_cnt",
            "validation_psi_acceptable_ok",
            "preferred_validation_psi_ok",
            "primary_iv_retention",
            "development_validation_rank_corr",
            "target_bin_distance",
            "candidate_score",
        ],
        ascending=[False, True, True, True, False, False, False, False, True, False],
        na_position="last",
    )

    if selection_pool.empty:
        raise ValueError("未生成任何可用合箱候选方案")

    selected_ranges_text = selection_pool.iloc[0]["ranges"]
    candidates.loc[candidates["ranges"].eq(selected_ranges_text), "selected"] = True
    candidates = candidates.sort_values(
        ["selected", "hard_constraints_ok", "candidate_score", "final_bin_count"],
        ascending=[False, False, False, False],
    ).reset_index(drop=True)

    steps = pd.DataFrame(step_rows)
    return candidates, steps, protected_boundaries


def selected_ranges_from_candidate_table(
    candidates: pd.DataFrame,
    fallback_ranges: Sequence[Tuple[int, int]],
) -> List[Tuple[int, int]]:
    """从候选评分表读取最终方案；无结果时使用手工兜底。"""
    if candidates.empty:
        return list(fallback_ranges)
    selected = candidates.loc[candidates["selected"].eq(True)]
    if selected.empty:
        return list(fallback_ranges)
    return parse_merge_ranges(str(selected.iloc[0]["ranges"]))


# ============================================================
# 5. OOT 基础验证
# ============================================================

def calc_population_psi(
    train: pd.DataFrame,
    oot: pd.DataFrame,
    bin_col: str,
    final_edges: pd.DataFrame,
    eps: float = 1e-6,
) -> pd.DataFrame:
    """计算 Train 与 OOT 的最终箱分布 PSI。"""
    base = final_edges[["final_bin_order", bin_col]].drop_duplicates()
    train_count = train[bin_col].value_counts().rename("train_n")
    oot_count = oot[bin_col].value_counts().rename("oot_n")

    psi = (
        base.merge(train_count, left_on=bin_col, right_index=True, how="left")
        .merge(oot_count, left_on=bin_col, right_index=True, how="left")
        .fillna({"train_n": 0, "oot_n": 0})
        .sort_values("final_bin_order")
        .reset_index(drop=True)
    )

    psi["train_pct"] = safe_div(psi["train_n"], psi["train_n"].sum())
    psi["oot_pct"] = safe_div(psi["oot_n"], psi["oot_n"].sum())

    train_pct = psi["train_pct"].clip(lower=eps)
    oot_pct = psi["oot_pct"].clip(lower=eps)
    psi["psi_component"] = (oot_pct - train_pct) * np.log(oot_pct / train_pct)
    psi["psi_total"] = psi["psi_component"].sum()
    return psi


def calc_auc_ks(
    data: pd.DataFrame,
    score_col: str,
    label_col: str,
) -> pd.Series:
    """直接计算二分类 AUC 和 KS，不依赖 sklearn。"""
    work = data[[score_col, label_col]].copy()
    work[score_col] = pd.to_numeric(work[score_col], errors="coerce")
    work[label_col] = pd.to_numeric(work[label_col], errors="coerce")
    work = work.loc[
        work[score_col].notna() & work[label_col].isin([0, 1])
    ].copy()

    n = len(work)
    bad_count = int(work[label_col].eq(1).sum())
    good_count = int(work[label_col].eq(0).sum())
    bad_rate = safe_div(bad_count, n)

    if n == 0 or bad_count == 0 or good_count == 0:
        return pd.Series(
            {
                "n": n,
                "bad_cnt": bad_count,
                "good_cnt": good_count,
                "bad_rate": bad_rate,
                "auc": np.nan,
                "ks": np.nan,
            }
        )

    risk_score = work[score_col] if HIGH_SCORE_HIGH_RISK else -work[score_col]
    ranks = risk_score.rank(method="average")
    bad_rank_sum = ranks.loc[work[label_col].eq(1)].sum()
    auc = (
        bad_rank_sum - bad_count * (bad_count + 1) / 2
    ) / (bad_count * good_count)

    ordered = work.assign(_risk_score=risk_score).sort_values(
        "_risk_score",
        ascending=False,
    )
    cum_bad = ordered[label_col].eq(1).cumsum() / bad_count
    cum_good = ordered[label_col].eq(0).cumsum() / good_count
    ks = (cum_bad - cum_good).abs().max()

    return pd.Series(
        {
            "n": n,
            "bad_cnt": bad_count,
            "good_cnt": good_count,
            "bad_rate": bad_rate,
            "auc": auc,
            "ks": ks,
        }
    )


def calc_performance_table(data: pd.DataFrame) -> pd.DataFrame:
    """按 Train / OOT 计算 1M30+ 和 3M30+ 的 AUC、KS。"""
    rows = []
    for sample_group, group_data in data.groupby("sample_group", observed=True):
        if sample_group not in {"train", "oot"}:
            continue
        for label_col in ["duedate_1m_30", "duedate_3m_30"]:
            metrics = calc_auc_ks(group_data, SCORE_COL, label_col).to_dict()
            metrics.update(
                {
                    "sample_group": sample_group,
                    "label": label_col,
                }
            )
            rows.append(metrics)

    return pd.DataFrame(rows)[
        ["sample_group", "label", "n", "bad_cnt", "good_cnt", "bad_rate", "auc", "ks"]
    ]


# ============================================================
# 6. 阈值曲线与策略结果
# ============================================================

def calc_portfolio_metrics(data: pd.DataFrame) -> Dict[str, float]:
    """计算一组样本的核心风险指标。"""
    work = add_risk_helper_columns(data)

    result: Dict[str, float] = {
        "n": len(work),
        "principal": float(work["_principal"].sum()),
    }

    for prefix, mature_col, bad_col, exposure_col, bad_amt_col in [
        ("1m30p", "_m1_mature_cnt", "_m1_bad_cnt", "_m1_amt_exposure", "_m1_amt_bad"),
        ("3m30p", "_m3_mature_cnt", "_m3_bad_cnt", "_m3_amt_exposure", "_m3_amt_bad"),
    ]:
        mature = int(work[mature_col].sum())
        bad = int(work[bad_col].sum())
        exposure = float(work[exposure_col].sum())
        bad_amount = float(work[bad_amt_col].sum())

        result[f"{prefix}_cnt_mature"] = mature
        result[f"{prefix}_cnt_bad"] = bad
        result[f"{prefix}_cnt_bad_rate"] = safe_div(bad, mature)
        result[f"{prefix}_amt_exposure"] = exposure
        result[f"{prefix}_amt_bad"] = bad_amount
        result[f"{prefix}_amt_bad_rate"] = safe_div(bad_amount, exposure)

    return result


def prefix_metrics(metrics: Dict[str, float], prefix: str) -> Dict[str, float]:
    """给指标字典统一增加前缀。"""
    return {f"{prefix}_{key}": value for key, value in metrics.items()}


def build_threshold_curve(
    train: pd.DataFrame,
    final_edges: pd.DataFrame,
) -> pd.DataFrame:
    """使用最终箱右边界生成可上线的候选阈值曲线。"""
    max_score = train[SCORE_COL].max()
    threshold_table = final_edges[
        ["final_bin_order", FINAL_BIN_COL, "score_right", "merged_from"]
    ].copy()
    threshold_table["threshold"] = threshold_table["score_right"].replace(np.inf, max_score)
    threshold_table = threshold_table.loc[threshold_table["threshold"].notna()].copy()
    threshold_table = threshold_table.sort_values("final_bin_order").reset_index(drop=True)

    total_n = len(train)
    total_principal = train["principal"].fillna(0).sum()
    score = train[SCORE_COL]

    rows = []
    previous_threshold: Optional[float] = None

    for threshold_order, threshold_row in threshold_table.iterrows():
        threshold = float(threshold_row["threshold"])

        if HIGH_SCORE_HIGH_RISK:
            cumulative_mask = score.le(threshold)
            marginal_mask = (
                cumulative_mask
                if previous_threshold is None
                else score.gt(previous_threshold) & score.le(threshold)
            )
        else:
            cumulative_mask = score.ge(threshold)
            marginal_mask = (
                cumulative_mask
                if previous_threshold is None
                else score.lt(previous_threshold) & score.ge(threshold)
            )

        cumulative_metrics = calc_portfolio_metrics(train.loc[cumulative_mask])
        marginal_metrics = calc_portfolio_metrics(train.loc[marginal_mask])

        row = {
            "threshold_order": threshold_order + 1,
            "threshold": threshold,
            "prev_threshold": previous_threshold,
            "final_bin_order": threshold_row["final_bin_order"],
            FINAL_BIN_COL: threshold_row[FINAL_BIN_COL],
            "merged_from": threshold_row["merged_from"],
        }
        row.update(prefix_metrics(cumulative_metrics, "cum"))
        row.update(prefix_metrics(marginal_metrics, "marginal"))
        row["cum_pass_rate"] = safe_div(cumulative_metrics["n"], total_n)
        row["cum_principal_pct"] = safe_div(
            cumulative_metrics["principal"],
            total_principal,
        )
        row["marginal_sample_pct"] = safe_div(marginal_metrics["n"], total_n)
        row["marginal_principal_pct"] = safe_div(
            marginal_metrics["principal"],
            total_principal,
        )
        rows.append(row)
        previous_threshold = threshold

    return pd.DataFrame(rows)


def select_threshold_under_constraints(
    curve: pd.DataFrame,
    constraints: Dict[str, float],
) -> Optional[pd.Series]:
    """选择满足风险约束且累计通过率最高的阈值。"""
    eligible = curve.copy()
    for constraint_name, maximum in constraints.items():
        metric = constraint_name.removeprefix("max_")
        require_columns(eligible, [metric], "策略阈值曲线")
        eligible = eligible.loc[eligible[metric].le(maximum)]

    if eligible.empty:
        return None

    return eligible.sort_values(
        ["cum_pass_rate", "threshold_order"],
        ascending=[False, False],
    ).iloc[0]


def build_strategy_plan(
    curve: pd.DataFrame,
    config: Dict,
) -> pd.DataFrame:
    """根据唯一一套配置生成自动通过、人工审核和拒绝阈值。"""
    auto_row = select_threshold_under_constraints(
        curve,
        config["auto_constraints"],
    )
    accept_row = select_threshold_under_constraints(
        curve,
        config["accept_constraints"],
    )

    base = {
        "strategy_name": config["strategy_name"],
        "objective": config["objective"],
    }

    if auto_row is None or accept_row is None:
        return pd.DataFrame([{**base, "status": "无满足约束的阈值"}])

    # 接纳阈值不能比自动通过阈值更严格。
    if accept_row["threshold_order"] < auto_row["threshold_order"]:
        accept_row = auto_row

    result = {
        **base,
        "status": "OK",
        "auto_pass_threshold": auto_row["threshold"],
        "auto_pass_bin": auto_row[FINAL_BIN_COL],
        "reject_threshold": accept_row["threshold"],
        "manual_review_upper_bin": accept_row[FINAL_BIN_COL],
        "auto_pass_rate": auto_row["cum_pass_rate"],
        "accepted_rate": accept_row["cum_pass_rate"],
        "manual_review_rate": (
            accept_row["cum_pass_rate"] - auto_row["cum_pass_rate"]
        ),
        "reject_rate": 1 - accept_row["cum_pass_rate"],
        "accepted_1m30p_cnt_bad_rate": accept_row["cum_1m30p_cnt_bad_rate"],
        "accepted_3m30p_cnt_bad_rate": accept_row["cum_3m30p_cnt_bad_rate"],
        "accepted_1m30p_amt_bad_rate": accept_row["cum_1m30p_amt_bad_rate"],
        "accepted_3m30p_amt_bad_rate": accept_row["cum_3m30p_amt_bad_rate"],
        "last_accepted_marginal_3m30p_cnt_bad_rate": accept_row[
            "marginal_3m30p_cnt_bad_rate"
        ],
    }
    return pd.DataFrame([result])


def calc_segment_metrics(
    data: pd.DataFrame,
    lower_threshold: Optional[float],
    upper_threshold: Optional[float],
) -> Dict[str, float]:
    """计算一个策略分数区间的样本和风险指标。"""
    score = data[SCORE_COL]
    mask = score.notna()

    if HIGH_SCORE_HIGH_RISK:
        if lower_threshold is not None:
            mask &= score.gt(lower_threshold)
        if upper_threshold is not None:
            mask &= score.le(upper_threshold)
    else:
        if lower_threshold is not None:
            mask &= score.lt(lower_threshold)
        if upper_threshold is not None:
            mask &= score.ge(upper_threshold)

    segment = data.loc[mask]
    metrics = calc_portfolio_metrics(segment)
    metrics["sample_pct"] = safe_div(len(segment), len(data))
    metrics["principal_pct"] = safe_div(
        metrics["principal"],
        data["principal"].fillna(0).sum(),
    )
    return metrics


def build_strategy_segment_report(
    train: pd.DataFrame,
    oot: pd.DataFrame,
    strategy_plan: pd.DataFrame,
) -> pd.DataFrame:
    """验证唯一策略在 Train 和 OOT 中的三段表现。"""
    valid = strategy_plan.loc[strategy_plan["status"].eq("OK")]
    if valid.empty:
        return pd.DataFrame()

    strategy = valid.iloc[0]
    auto_threshold = float(strategy["auto_pass_threshold"])
    reject_threshold = float(strategy["reject_threshold"])
    segments = [
        ("自动通过", None, auto_threshold),
        ("人工审核", auto_threshold, reject_threshold),
        ("拒绝", reject_threshold, None),
    ]

    rows = []
    for sample_group, data in [("train", train), ("oot", oot)]:
        for decision, lower, upper in segments:
            metrics = calc_segment_metrics(data, lower, upper)
            rows.append(
                {
                    "sample_group": sample_group,
                    "strategy_name": strategy["strategy_name"],
                    "decision": decision,
                    "lower_threshold_exclusive": lower,
                    "upper_threshold_inclusive": upper,
                    **metrics,
                }
            )

    return pd.DataFrame(rows)




def build_binning_process_table(
    initial_stats: pd.DataFrame,
    merge_map: pd.DataFrame,
) -> pd.DataFrame:
    """
    汇总初始分箱、风险表现和最终合箱映射。

    这张表用于回答三个问题：
    1. 每个初始箱的样本量和风险表现如何；
    2. 相邻箱之间是否出现风险倒挂；
    3. 每个初始箱最终被合并到哪个风险等级。
    """
    process = initial_stats.copy().rename(columns={"bin_order": "initial_bin_order"})
    process = process.merge(
        merge_map[
            [
                "initial_bin_order",
                INITIAL_BIN_COL,
                "final_bin_order",
                FINAL_BIN_COL,
                "merged_from",
            ]
        ],
        on=["initial_bin_order", INITIAL_BIN_COL],
        how="left",
    )
    process = process.sort_values("initial_bin_order").reset_index(drop=True)

    for prefix in ["1m30p", "3m30p"]:
        rate_col = f"{prefix}_cnt_bad_rate"
        diff_col = f"{prefix}_rate_diff_prev"
        inversion_col = f"{prefix}_inversion_flag"
        process[diff_col] = process[rate_col].diff()
        process[inversion_col] = process[diff_col].lt(0).fillna(False)

    process["merge_action"] = np.where(
        process["merged_from"].astype(str).str.contains("-", regex=False),
        "相邻箱合并",
        "单箱保留",
    )

    key_columns = [
        "initial_bin_order",
        INITIAL_BIN_COL,
        "score_left",
        "score_right",
        "score_min",
        "score_max",
        "score_mean",
        "n",
        "sample_pct",
        "1m30p_cnt_mature",
        "1m30p_cnt_bad",
        "1m30p_cnt_bad_rate",
        "1m30p_rate_diff_prev",
        "1m30p_inversion_flag",
        "3m30p_cnt_mature",
        "3m30p_cnt_bad",
        "3m30p_cnt_bad_rate",
        "3m30p_rate_diff_prev",
        "3m30p_inversion_flag",
        "1m30p_amt_exposure",
        "1m30p_amt_bad",
        "1m30p_amt_bad_rate",
        "3m30p_amt_exposure",
        "3m30p_amt_bad",
        "3m30p_amt_bad_rate",
        "cum_pass_rate",
        "cum_1m30p_cnt_bad_rate",
        "cum_3m30p_cnt_bad_rate",
        "final_bin_order",
        FINAL_BIN_COL,
        "merged_from",
        "merge_action",
    ]
    return process[[col for col in key_columns if col in process.columns]]


def build_threshold_selection_table(
    threshold_curve: pd.DataFrame,
    strategy_plan: pd.DataFrame,
    config: Dict,
) -> pd.DataFrame:
    """
    在阈值曲线上补充约束检查结果和最终阈值标记。

    Excel 中可以直接看到：
    - 每个候选阈值的累计通过率、累计风险和边际风险；
    - 是否满足自动通过约束；
    - 是否满足整体接纳约束；
    - 哪一行最终被选为自动通过阈值或人工审核上限。
    """
    result = threshold_curve.copy()

    for group_name, constraints in [
        ("auto", config["auto_constraints"]),
        ("accept", config["accept_constraints"]),
    ]:
        check_columns = []
        for constraint_name, limit in constraints.items():
            metric = constraint_name.removeprefix("max_")
            check_col = f"{group_name}_check_{metric}"
            limit_col = f"{group_name}_limit_{metric}"
            result[limit_col] = limit
            result[check_col] = result[metric].le(limit).fillna(False)
            check_columns.append(check_col)
        result[f"{group_name}_all_constraints_ok"] = result[check_columns].all(axis=1)

    result["selected_role"] = ""
    result["selection_reason"] = ""

    valid = strategy_plan.loc[strategy_plan["status"].eq("OK")]
    if not valid.empty:
        strategy = valid.iloc[0]
        auto_threshold = float(strategy["auto_pass_threshold"])
        reject_threshold = float(strategy["reject_threshold"])

        auto_mask = np.isclose(result["threshold"].astype(float), auto_threshold)
        reject_mask = np.isclose(result["threshold"].astype(float), reject_threshold)

        result.loc[auto_mask, "selected_role"] = "自动通过阈值"
        result.loc[auto_mask, "selection_reason"] = (
            "满足自动通过全部风险约束，且累计通过率最高"
        )

        same_threshold = auto_mask & reject_mask
        result.loc[reject_mask & ~same_threshold, "selected_role"] = "人工审核上限/拒绝阈值"
        result.loc[reject_mask & ~same_threshold, "selection_reason"] = (
            "满足整体接纳全部风险约束，且累计接纳率最高"
        )
        result.loc[same_threshold, "selected_role"] = "自动通过阈值及拒绝阈值"
        result.loc[same_threshold, "selection_reason"] = (
            "自动通过与整体接纳最终选择了同一阈值"
        )

    first_columns = [
        "selected_role",
        "selection_reason",
        "threshold_order",
        "threshold",
        "prev_threshold",
        "final_bin_order",
        FINAL_BIN_COL,
        "merged_from",
        "cum_pass_rate",
        "cum_n",
        "cum_principal_pct",
        "cum_1m30p_cnt_mature",
        "cum_1m30p_cnt_bad",
        "cum_1m30p_cnt_bad_rate",
        "cum_3m30p_cnt_mature",
        "cum_3m30p_cnt_bad",
        "cum_3m30p_cnt_bad_rate",
        "cum_1m30p_amt_exposure",
        "cum_1m30p_amt_bad",
        "cum_1m30p_amt_bad_rate",
        "cum_3m30p_amt_exposure",
        "cum_3m30p_amt_bad",
        "cum_3m30p_amt_bad_rate",
        "marginal_sample_pct",
        "marginal_n",
        "marginal_3m30p_cnt_mature",
        "marginal_3m30p_cnt_bad",
        "marginal_3m30p_cnt_bad_rate",
        "auto_all_constraints_ok",
        "accept_all_constraints_ok",
    ]
    remaining_columns = [col for col in result.columns if col not in first_columns]
    return result[[col for col in first_columns if col in result.columns] + remaining_columns]


def build_metric_dictionary() -> pd.DataFrame:
    """输出 Excel 核心字段和计算口径说明。"""
    rows = [
        ("通用", "n", "箱内或区间内的申请样本量", "COUNT(application_id)"),
        ("通用", "sample_pct", "箱内样本占全部样本的比例", "n / total_n"),
        ("通用", "principal_amt", "箱内样本本金合计", "SUM(principal)"),
        ("分箱", "score_left / score_right", "分箱的模型分左右边界", "(score_left, score_right]"),
        ("分箱", "merged_from", "最终风险档位由哪些初始箱合并而来", "例如 B06-B08"),
        ("笔数风险", "1m30p_cnt_mature", "1M30+ 已成熟样本量", "duedate_1m_30 IN (0, 1)"),
        ("笔数风险", "1m30p_cnt_bad", "1M30+ 逾期样本量", "duedate_1m_30 = 1"),
        ("笔数风险", "1m30p_cnt_bad_rate", "1M30+ 笔数逾期率", "1m30p_cnt_bad / 1m30p_cnt_mature"),
        ("笔数风险", "3m30p_cnt_mature", "3M30+ 已成熟样本量", "duedate_3m_30 IN (0, 1)"),
        ("笔数风险", "3m30p_cnt_bad", "3M30+ 逾期样本量", "duedate_3m_30 = 1"),
        ("笔数风险", "3m30p_cnt_bad_rate", "3M30+ 笔数逾期率", "3m30p_cnt_bad / 3m30p_cnt_mature"),
        ("金额风险", "1m30p_amt_exposure", "1M30+ 已成熟样本的本金敞口", "SUM(principal) WHERE MOB1 已成熟"),
        ("金额风险", "1m30p_amt_bad", "MOB1 30+ 样本的剩余本金", "SUM(estimate_principal_remaining_mob1)"),
        ("金额风险", "1m30p_amt_bad_rate", "1M30+ 金额逾期率", "1m30p_amt_bad / 1m30p_amt_exposure"),
        ("金额风险", "3m30p_amt_exposure", "3M30+ 已成熟样本的本金敞口", "SUM(principal) WHERE MOB3 已成熟"),
        ("金额风险", "3m30p_amt_bad", "MOB3 30+ 样本的剩余本金", "SUM(estimate_principal_remaining_mob3)"),
        ("金额风险", "3m30p_amt_bad_rate", "3M30+ 金额逾期率", "3m30p_amt_bad / 3m30p_amt_exposure"),
        ("阈值", "cum_pass_rate", "从低风险端累计到当前阈值的通过率", "cum_n / total_n"),
        ("阈值", "marginal_sample_pct", "当前档位新增样本占比", "marginal_n / total_n"),
        ("阈值", "marginal_3m30p_cnt_bad_rate", "当前新增档位自身的 3M30+ 风险", "marginal_bad / marginal_mature"),
        ("阈值", "auto_all_constraints_ok", "该候选阈值是否满足自动通过全部约束", "全部自动通过检查项均为 True"),
        ("阈值", "accept_all_constraints_ok", "该候选阈值是否满足整体接纳全部约束", "全部整体接纳检查项均为 True"),
        ("验证", "PSI", "Train 与 OOT 的分箱分布稳定性", "SUM((OOT%-Train%) * LN(OOT%/Train%))"),
        ("验证", "AUC / KS", "模型风险区分能力指标", "分别衡量排序能力和好坏样本累计差异"),
    ]
    return pd.DataFrame(rows, columns=["category", "field", "definition", "calculation"])


def build_monthly_bin_stability(data: pd.DataFrame) -> pd.DataFrame:
    """按月份、样本组和最终风险档输出箱级稳定性指标。"""
    rows = []
    valid = data.loc[
        data["sample_group"].isin(["train", "oot"])
        & data["application_month"].notna()
        & data[FINAL_BIN_COL].notna()
    ].copy()

    for (sample_group, application_month), month_data in valid.groupby(
        ["sample_group", "application_month"],
        observed=True,
    ):
        stats = calc_bin_stats(
            month_data,
            bin_col=FINAL_BIN_COL,
            order_col="bin_order",
        )
        stats.insert(0, "application_month", application_month)
        stats.insert(0, "sample_group", sample_group)
        rows.append(stats)

    if not rows:
        return pd.DataFrame()

    result = pd.concat(rows, ignore_index=True)
    result["primary_rate_diff_prev"] = result.groupby(
        ["sample_group", "application_month"],
        observed=True,
    )[PRIMARY_RATE_COL].diff()
    if not HIGH_SCORE_HIGH_RISK:
        result["primary_rate_diff_prev"] = -result["primary_rate_diff_prev"]
    result["primary_inversion_flag"] = (
        result["primary_rate_diff_prev"] < -VALIDATION_INVERSION_TOLERANCE
    )
    return result


def build_monthly_stability_summary(monthly_stats: pd.DataFrame) -> pd.DataFrame:
    """汇总每个月的主风险指标单调性和样本表现。"""
    if monthly_stats.empty:
        return pd.DataFrame()

    return (
        monthly_stats.groupby(["sample_group", "application_month"], observed=True)
        .agg(
            n=("n", "sum"),
            mature_count=(PRIMARY_MATURE_COL, "sum"),
            bad_count=(PRIMARY_BAD_COL, "sum"),
            bin_count=(FINAL_BIN_COL, "nunique"),
            primary_inversion_count=("primary_inversion_flag", "sum"),
            max_primary_rate_drop=("primary_rate_diff_prev", lambda s: float((-s).clip(lower=0).max())),
        )
        .reset_index()
        .assign(
            primary_bad_rate=lambda frame: safe_div(
                frame["bad_count"], frame["mature_count"]
            ),
            primary_monotonic_ok=lambda frame: frame["primary_inversion_count"].eq(0),
        )
    )


# ============================================================
# 7. 报告数据整理与输出
# ============================================================

def build_train_oot_compare(
    train_stats: pd.DataFrame,
    oot_stats: pd.DataFrame,
    final_edges: pd.DataFrame,
) -> pd.DataFrame:
    """生成最终箱 Train / OOT 对比表。"""
    key_cols = [FINAL_BIN_COL]
    compare_cols = [
        FINAL_BIN_COL,
        "n",
        "sample_pct",
        "1m30p_cnt_mature",
        "1m30p_cnt_bad",
        "1m30p_cnt_bad_rate",
        "3m30p_cnt_mature",
        "3m30p_cnt_bad",
        "3m30p_cnt_bad_rate",
        "1m30p_amt_exposure",
        "1m30p_amt_bad",
        "1m30p_amt_bad_rate",
        "3m30p_amt_exposure",
        "3m30p_amt_bad",
        "3m30p_amt_bad_rate",
    ]

    comparison = train_stats[compare_cols].merge(
        oot_stats[compare_cols],
        on=key_cols,
        how="outer",
        suffixes=("_train", "_oot"),
    )

    return final_edges[
        ["final_bin_order", FINAL_BIN_COL, "merged_from", "score_left", "score_right"]
    ].merge(comparison, on=FINAL_BIN_COL, how="left")


def build_overview(
    data: pd.DataFrame,
    train: pd.DataFrame,
    development: pd.DataFrame,
    validation: pd.DataFrame,
    oot: pd.DataFrame,
    validation_months: Sequence[str],
    initial_bin_count: int,
    final_bin_count: int,
    selected_merge_ranges: Sequence[Tuple[int, int]],
    selected_candidate: Optional[pd.Series],
    protected_boundaries: Set[int],
    psi: pd.DataFrame,
    performance: pd.DataFrame,
    monotonicity: pd.DataFrame,
    strategy_plan: pd.DataFrame,
) -> pd.DataFrame:
    """整理报告首页的核心结论，并按模块分组展示。"""
    source_row_count = int(data.attrs.get("source_row_count", len(data)))
    score_missing_count = int(data.attrs.get("score_missing_count", 0))

    rows = [
        ("样本", "原始样本量", source_row_count),
        ("样本", "有效模型分样本量", len(data)),
        ("样本", "模型分缺失量", score_missing_count),
        ("样本", "模型分缺失率", safe_div(score_missing_count, source_row_count)),
        ("样本", "Train 样本量", len(train)),
        ("样本", "Development 样本量", len(development)),
        ("样本", "Validation 样本量", len(validation)),
        ("样本", "OOT 样本量", len(oot)),
        ("时间切分", "Train 截止月份", TRAIN_END_MONTH),
        ("时间切分", "Validation 月份", ",".join(validation_months)),
        ("时间切分", "OOT 起始月份", OOT_START_MONTH),
        ("分箱", "初始箱数量", initial_bin_count),
        ("分箱", "最终箱数量", final_bin_count),
        ("分箱", "自动合箱", AUTO_SELECT_MERGE_RANGES),
        ("分箱", "合箱主指标", PRIMARY_RATE_COL),
        ("分箱", "最终采用合箱方案", format_merge_ranges(selected_merge_ranges)),
        ("分箱", "受保护初始边界", ",".join(map(str, sorted(protected_boundaries)))),
        ("稳定性", "最终箱 Train/OOT PSI", psi["psi_total"].iloc[0]),
    ]

    if selected_candidate is not None:
        rows.extend(
            [
                ("候选评分", "Development 主指标倒挂数", selected_candidate.get("development_primary_inversion_cnt")),
                ("候选评分", "Validation 主指标倒挂数", selected_candidate.get("validation_primary_inversion_cnt")),
                ("候选评分", "Development/Validation PSI", selected_candidate.get("validation_psi")),
                ("候选评分", "主指标 IV 保留率", selected_candidate.get("primary_iv_retention")),
                ("候选评分", "跨样本风险排序相关性", selected_candidate.get("development_validation_rank_corr")),
                ("候选评分", "候选综合得分", selected_candidate.get("candidate_score")),
            ]
        )

    for _, perf in performance.iterrows():
        prefix = f"{perf['sample_group']}_{perf['label']}"
        rows.extend(
            [
                ("模型效果", f"{prefix}_bad_rate", perf["bad_rate"]),
                ("模型效果", f"{prefix}_auc", perf["auc"]),
                ("模型效果", f"{prefix}_ks", perf["ks"]),
            ]
        )

    for sample_group in ["development", "validation", "train", "oot"]:
        sample_check = monotonicity.loc[monotonicity["sample_group"].eq(sample_group)]
        if sample_check.empty:
            continue
        rows.append(
            (
                "单调性",
                f"{sample_group}_最终箱全部单调",
                bool(sample_check["is_monotonic_non_decreasing"].all()),
            )
        )

    valid_strategy = strategy_plan.loc[strategy_plan["status"].eq("OK")]
    if not valid_strategy.empty:
        row = valid_strategy.iloc[0]
        rows.extend(
            [
                ("策略", "策略名称", row["strategy_name"]),
                ("策略", "自动通过阈值", row["auto_pass_threshold"]),
                ("策略", "自动通过截止风险档", row["auto_pass_bin"]),
                ("策略", "人工审核上限/拒绝阈值", row["reject_threshold"]),
                ("策略", "人工审核截止风险档", row["manual_review_upper_bin"]),
                ("策略", "自动通过率", row["auto_pass_rate"]),
                ("策略", "人工审核率", row["manual_review_rate"]),
                ("策略", "拒绝率", row["reject_rate"]),
                ("策略风险", "接纳人群1M30+笔数逾期率", row["accepted_1m30p_cnt_bad_rate"]),
                ("策略风险", "接纳人群3M30+笔数逾期率", row["accepted_3m30p_cnt_bad_rate"]),
                ("策略风险", "接纳人群1M30+金额逾期率", row["accepted_1m30p_amt_bad_rate"]),
                ("策略风险", "接纳人群3M30+金额逾期率", row["accepted_3m30p_amt_bad_rate"]),
                ("策略风险", "最后接纳档边际3M30+", row["last_accepted_marginal_3m30p_cnt_bad_rate"]),
            ]
        )
    else:
        status = strategy_plan.iloc[0]["status"] if not strategy_plan.empty else "未生成"
        rows.append(("策略", "策略状态", status))

    return pd.DataFrame(rows, columns=["section", "metric", "value"])


def build_config_table(
    selected_merge_ranges: Sequence[Tuple[int, int]],
    validation_months: Sequence[str],
    protected_boundaries: Set[int],
) -> pd.DataFrame:
    """输出便于后续修改和版本管理的参数表。"""
    rows = [
        {"config_group": "基础配置", "config_name": "DATA_DIR", "config_value": str(DATA_DIR)},
        {"config_group": "基础配置", "config_name": "TRAIN_END_MONTH", "config_value": TRAIN_END_MONTH},
        {"config_group": "基础配置", "config_name": "OOT_START_MONTH", "config_value": OOT_START_MONTH},
        {"config_group": "基础配置", "config_name": "VALIDATION_MONTH_COUNT", "config_value": VALIDATION_MONTH_COUNT},
        {"config_group": "基础配置", "config_name": "ACTUAL_VALIDATION_MONTHS", "config_value": ",".join(validation_months)},
        {"config_group": "基础配置", "config_name": "INITIAL_BIN_COUNT", "config_value": INITIAL_BIN_COUNT},
        {"config_group": "基础配置", "config_name": "HIGH_SCORE_HIGH_RISK", "config_value": HIGH_SCORE_HIGH_RISK},
        {"config_group": "合箱配置", "config_name": "AUTO_SELECT_MERGE_RANGES", "config_value": AUTO_SELECT_MERGE_RANGES},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_COUNT", "config_value": MIN_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "MAX_FINAL_BIN_COUNT", "config_value": MAX_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "TARGET_FINAL_BIN_COUNT", "config_value": TARGET_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "PRIMARY_RATE_COL", "config_value": PRIMARY_RATE_COL},
        {"config_group": "合箱配置", "config_name": "MIN_MIDDLE_BIN_SAMPLE_PCT", "config_value": MIN_MIDDLE_BIN_SAMPLE_PCT},
        {"config_group": "合箱配置", "config_name": "MIN_TAIL_BIN_SAMPLE_PCT", "config_value": MIN_TAIL_BIN_SAMPLE_PCT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_MATURE_COUNT", "config_value": MIN_FINAL_BIN_MATURE_COUNT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_BAD_COUNT", "config_value": MIN_FINAL_BIN_BAD_COUNT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_GOOD_COUNT", "config_value": MIN_FINAL_BIN_GOOD_COUNT},
        {"config_group": "合箱配置", "config_name": "VALIDATION_INVERSION_TOLERANCE", "config_value": VALIDATION_INVERSION_TOLERANCE},
        {"config_group": "合箱配置", "config_name": "ADJACENT_PVALUE_TO_MERGE", "config_value": ADJACENT_PVALUE_TO_MERGE},
        {"config_group": "合箱配置", "config_name": "MIN_ADJACENT_ABS_RATE_DIFF", "config_value": MIN_ADJACENT_ABS_RATE_DIFF},
        {"config_group": "合箱配置", "config_name": "PREFERRED_MAX_VALIDATION_PSI", "config_value": PREFERRED_MAX_VALIDATION_PSI},
        {"config_group": "合箱配置", "config_name": "MAX_ACCEPTABLE_VALIDATION_PSI", "config_value": MAX_ACCEPTABLE_VALIDATION_PSI},
        {"config_group": "合箱配置", "config_name": "PROTECTED_BOUNDARIES", "config_value": ",".join(map(str, sorted(protected_boundaries)))},
        {"config_group": "合箱配置", "config_name": "MANUAL_FALLBACK_FINAL_BIN_RANGES", "config_value": str(FINAL_BIN_RANGES)},
        {"config_group": "合箱配置", "config_name": "SELECTED_FINAL_BIN_RANGES", "config_value": format_merge_ranges(selected_merge_ranges)},
    ]

    flattened = {
        **flatten_dict("auto", STRATEGY_CONFIG["auto_constraints"]),
        **flatten_dict("accept", STRATEGY_CONFIG["accept_constraints"]),
    }
    for name, value in flattened.items():
        rows.append(
            {
                "config_group": STRATEGY_CONFIG["strategy_name"],
                "config_name": name,
                "config_value": value,
            }
        )

    return pd.DataFrame(rows)


def format_excel_report(path: Path) -> None:
    """设置基础格式，并突出倒挂、约束失败、选中阈值和选中合箱方案。"""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    selected_fill = PatternFill("solid", fgColor="C6E0B4")
    selected_reject_fill = PatternFill("solid", fgColor="FCE4D6")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in sheet.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = 0
            for cell in column_cells:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 42)

        headers = {cell.column: str(cell.value) for cell in sheet[1]}
        header_to_col = {str(cell.value): cell.column for cell in sheet[1]}

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                header = headers.get(cell.column, "").lower()
                if cell.value is None:
                    continue
                if any(key in header for key in ["rate", "pct", "retention"]):
                    cell.number_format = "0.00%"
                elif any(key in header for key in ["auc", "ks", "psi", "corr", "p_value"]):
                    cell.number_format = "0.0000"
                elif any(key in header for key in ["threshold", "score_left", "score_right", "score_min", "score_max", "score_mean"]):
                    cell.number_format = "0.0000"
                elif isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        if sheet.title == "02_分箱过程":
            inversion_cols = [
                header_to_col.get("1m30p_inversion_flag"),
                header_to_col.get("3m30p_inversion_flag"),
            ]
            inversion_cols = [col for col in inversion_cols if col is not None]
            for row_idx in range(2, sheet.max_row + 1):
                if any(sheet.cell(row_idx, col).value is True for col in inversion_cols):
                    for cell in sheet[row_idx]:
                        cell.fill = warning_fill

        if sheet.title == "08_阈值选择过程":
            selected_col = header_to_col.get("selected_role")
            auto_ok_col = header_to_col.get("auto_all_constraints_ok")
            accept_ok_col = header_to_col.get("accept_all_constraints_ok")

            for row_idx in range(2, sheet.max_row + 1):
                selected_role = sheet.cell(row_idx, selected_col).value if selected_col else ""
                if selected_role and "自动通过" in str(selected_role):
                    for cell in sheet[row_idx]:
                        cell.fill = selected_fill
                if selected_role and "拒绝阈值" in str(selected_role):
                    for cell in sheet[row_idx]:
                        cell.fill = selected_reject_fill

                for check_col in [auto_ok_col, accept_ok_col]:
                    if check_col and sheet.cell(row_idx, check_col).value is False:
                        sheet.cell(row_idx, check_col).fill = fail_fill

        if sheet.title == "16_合箱候选评分":
            selected_col = header_to_col.get("selected")
            hard_ok_col = header_to_col.get("hard_constraints_ok")
            for row_idx in range(2, sheet.max_row + 1):
                if selected_col and sheet.cell(row_idx, selected_col).value is True:
                    for cell in sheet[row_idx]:
                        cell.fill = selected_fill
                elif hard_ok_col and sheet.cell(row_idx, hard_ok_col).value is False:
                    sheet.cell(row_idx, hard_ok_col).fill = fail_fill

        if sheet.title == "15_月度箱表现":
            inversion_col = header_to_col.get("primary_inversion_flag")
            if inversion_col:
                for row_idx in range(2, sheet.max_row + 1):
                    if sheet.cell(row_idx, inversion_col).value is True:
                        for cell in sheet[row_idx]:
                            cell.fill = warning_fill

    workbook.save(path)


def write_report(
    overview: pd.DataFrame,
    binning_process: pd.DataFrame,
    final_development_stats: pd.DataFrame,
    final_validation_stats: pd.DataFrame,
    final_train_stats: pd.DataFrame,
    final_oot_stats: pd.DataFrame,
    train_oot_compare: pd.DataFrame,
    threshold_selection: pd.DataFrame,
    strategy_plan: pd.DataFrame,
    strategy_segments: pd.DataFrame,
    performance: pd.DataFrame,
    psi: pd.DataFrame,
    monotonicity: pd.DataFrame,
    monthly_stability: pd.DataFrame,
    monthly_stability_summary: pd.DataFrame,
    merge_candidates: pd.DataFrame,
    merge_steps: pd.DataFrame,
    config_table: pd.DataFrame,
    metric_dictionary: pd.DataFrame,
) -> None:
    """输出优化版策略报告，并保留完整的合箱过程。"""
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(REPORT_PATH, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="01_总览", index=False)
        binning_process.to_excel(writer, sheet_name="02_分箱过程", index=False)
        final_development_stats.to_excel(writer, sheet_name="03_最终分箱_开发", index=False)
        final_validation_stats.to_excel(writer, sheet_name="04_最终分箱_验证", index=False)
        final_train_stats.to_excel(writer, sheet_name="05_最终分箱_Train", index=False)
        final_oot_stats.to_excel(writer, sheet_name="06_最终分箱_OOT", index=False)
        train_oot_compare.to_excel(writer, sheet_name="07_Train_OOT对比", index=False)
        threshold_selection.to_excel(writer, sheet_name="08_阈值选择过程", index=False)
        strategy_plan.to_excel(writer, sheet_name="09_策略结果", index=False)
        strategy_segments.to_excel(writer, sheet_name="10_策略分段验证", index=False)
        performance.to_excel(writer, sheet_name="11_AUC_KS", index=False)
        psi.to_excel(writer, sheet_name="12_PSI", index=False)
        monotonicity.to_excel(writer, sheet_name="13_单调性", index=False)
        monthly_stability_summary.to_excel(writer, sheet_name="14_月度稳定性汇总", index=False)
        monthly_stability.to_excel(writer, sheet_name="15_月度箱表现", index=False)
        merge_candidates.to_excel(writer, sheet_name="16_合箱候选评分", index=False)
        merge_steps.to_excel(writer, sheet_name="17_合箱步骤", index=False)
        config_table.to_excel(writer, sheet_name="18_配置参数", index=False)
        metric_dictionary.to_excel(writer, sheet_name="19_指标说明", index=False)

    format_excel_report(REPORT_PATH)


# ============================================================
# 8. 主流程
# ============================================================

def _log_step(label: str, t_prev: float) -> float:
    t_now = time.time()
    elapsed = t_now - t_prev
    print(f"  [{label}] 耗时 {elapsed:.1f}s | 累计 {t_now - _log_step._t0:.1f}s")
    return t_now


def main() -> None:
    _t = _log_step._t0 = time.time()

    data = load_analysis_data()
    source_row_count = data.attrs.get("source_row_count")
    score_missing_count = data.attrs.get("score_missing_count")

    all_data, train, oot = split_train_oot(data)
    development, validation, validation_months = split_development_validation(train)
    _t = _log_step("1/9 数据加载与时间切分", _t)

    all_data.attrs["source_row_count"] = source_row_count
    all_data.attrs["score_missing_count"] = score_missing_count

    # 1) Development 学习初始边界，复用到 Validation、完整 Train 和 OOT。
    edges = learn_equal_freq_edges(development, SCORE_COL, INITIAL_BIN_COUNT)
    actual_initial_bin_count = len(edges) - 1
    initial_edges = build_initial_edge_table(edges)

    if actual_initial_bin_count < MIN_FINAL_BIN_COUNT:
        raise ValueError(
            f"模型分唯一值不足，实际仅形成 {actual_initial_bin_count} 个初始箱，"
            f"小于最小最终箱数 {MIN_FINAL_BIN_COUNT}"
        )

    all_binned = apply_edges(all_data, SCORE_COL, edges, INITIAL_BIN_COL)
    train_binned = all_binned.loc[all_binned["sample_group"].eq("train")].copy()
    oot_binned = all_binned.loc[all_binned["sample_group"].eq("oot")].copy()
    development_binned = apply_edges(development, SCORE_COL, edges, INITIAL_BIN_COL)
    validation_binned = apply_edges(validation, SCORE_COL, edges, INITIAL_BIN_COL)

    development_initial_stats = calc_complete_initial_stats(
        development_binned,
        initial_edges,
    )
    validation_initial_stats = calc_complete_initial_stats(
        validation_binned,
        initial_edges,
    )
    _t = _log_step(f"2/9 Development 等频初分：{actual_initial_bin_count} 箱", _t)

    # 2) Development + Validation 自动选择合箱；OOT 不参与。
    fallback_ranges = (
        FINAL_BIN_RANGES
        if actual_initial_bin_count == INITIAL_BIN_COUNT
        else make_equal_contiguous_ranges(actual_initial_bin_count, TARGET_FINAL_BIN_COUNT)
    )

    if AUTO_SELECT_MERGE_RANGES:
        merge_candidates, merge_steps, protected_boundaries = build_merge_candidate_score_table(
            development_initial_stats,
            validation_initial_stats,
            actual_initial_bin_count,
            STRATEGY_CONFIG,
        )
        selected_merge_ranges = selected_ranges_from_candidate_table(
            merge_candidates,
            fallback_ranges,
        )
    else:
        selected_merge_ranges = list(fallback_ranges)
        protected_boundaries = set()
        manual_development_stats = aggregate_initial_stats_by_ranges(
            development_initial_stats,
            selected_merge_ranges,
        )
        merge_candidates = pd.DataFrame(
            [
                {
                    "selected": True,
                    "stage": "manual",
                    "merge_reason": "AUTO_SELECT_MERGE_RANGES=False，使用手工兜底方案",
                    "final_bin_count": len(selected_merge_ranges),
                    "ranges": format_merge_ranges(selected_merge_ranges),
                    "development_primary_inversion_cnt": count_rate_inversions(
                        manual_development_stats,
                        [PRIMARY_RATE_COL],
                    ),
                }
            ]
        )
        merge_steps = pd.DataFrame()

    merge_map = build_merge_map(selected_merge_ranges, actual_initial_bin_count)
    final_edges = build_final_edge_table(initial_edges, merge_map)
    _t = _log_step(
        f"3/9 自动合箱完成：{len(final_edges)} 档，方案={format_merge_ranges(selected_merge_ranges)}",
        _t,
    )

    # 3) 将最终合箱映射应用到所有样本。
    development_final = apply_merge_map(development_binned, merge_map)
    validation_final = apply_merge_map(validation_binned, merge_map)
    train_final = apply_merge_map(train_binned, merge_map)
    oot_final = apply_merge_map(oot_binned, merge_map)
    all_final = apply_merge_map(all_binned, merge_map)

    def final_stats(frame: pd.DataFrame) -> pd.DataFrame:
        return calc_bin_stats(
            frame,
            bin_col=FINAL_BIN_COL,
            order_col="bin_order",
        ).merge(
            final_edges,
            left_on=["bin_order", FINAL_BIN_COL],
            right_on=["final_bin_order", FINAL_BIN_COL],
            how="left",
        )

    final_development_stats = final_stats(development_final)
    final_validation_stats = final_stats(validation_final)
    final_train_stats = final_stats(train_final)
    final_oot_stats = final_stats(oot_final)
    _t = _log_step("4/9 生成 Development/Validation/Train/OOT 最终箱统计", _t)

    # 4) 最终验证。
    rate_cols = [
        "1m30p_cnt_bad_rate",
        "3m30p_cnt_bad_rate",
        "1m30p_amt_bad_rate",
        "3m30p_amt_bad_rate",
    ]
    monotonicity = pd.concat(
        [
            check_monotonicity(final_development_stats, rate_cols, "development"),
            check_monotonicity(final_validation_stats, rate_cols, "validation"),
            check_monotonicity(final_train_stats, rate_cols, "train"),
            check_monotonicity(final_oot_stats, rate_cols, "oot"),
        ],
        ignore_index=True,
    )

    psi = calc_population_psi(train_final, oot_final, FINAL_BIN_COL, final_edges)
    performance = calc_performance_table(all_final)
    train_oot_compare = build_train_oot_compare(
        final_train_stats,
        final_oot_stats,
        final_edges,
    )
    monthly_stability = build_monthly_bin_stability(all_final)
    monthly_stability_summary = build_monthly_stability_summary(monthly_stability)
    _t = _log_step(
        f"5/9 OOT 单调性/PSI/AUC/KS 验证：PSI={psi['psi_total'].iloc[0]:.4f}",
        _t,
    )

    # 5) 使用完整 Train 生成策略阈值。
    threshold_curve = build_threshold_curve(train_final, final_edges)
    strategy_plan = build_strategy_plan(threshold_curve, STRATEGY_CONFIG)
    strategy_segments = build_strategy_segment_report(
        train_final,
        oot_final,
        strategy_plan,
    )
    binning_process = build_binning_process_table(development_initial_stats, merge_map)
    threshold_selection = build_threshold_selection_table(
        threshold_curve,
        strategy_plan,
        STRATEGY_CONFIG,
    )
    _t = _log_step("6/9 阈值曲线与自动通过/人工审核/拒绝策略", _t)

    selected_candidate_rows = merge_candidates.loc[
        merge_candidates.get("selected", pd.Series(False, index=merge_candidates.index)).eq(True)
    ]
    selected_candidate = (
        selected_candidate_rows.iloc[0] if not selected_candidate_rows.empty else None
    )

    overview = build_overview(
        all_data,
        train_final,
        development_final,
        validation_final,
        oot_final,
        validation_months,
        actual_initial_bin_count,
        len(final_edges),
        selected_merge_ranges,
        selected_candidate,
        protected_boundaries,
        psi,
        performance,
        monotonicity,
        strategy_plan,
    )
    config_table = build_config_table(
        selected_merge_ranges,
        validation_months,
        protected_boundaries,
    )
    metric_dictionary = build_metric_dictionary()

    write_report(
        overview=overview,
        binning_process=binning_process,
        final_development_stats=final_development_stats,
        final_validation_stats=final_validation_stats,
        final_train_stats=final_train_stats,
        final_oot_stats=final_oot_stats,
        train_oot_compare=train_oot_compare,
        threshold_selection=threshold_selection,
        strategy_plan=strategy_plan,
        strategy_segments=strategy_segments,
        performance=performance,
        psi=psi,
        monotonicity=monotonicity,
        monthly_stability=monthly_stability,
        monthly_stability_summary=monthly_stability_summary,
        merge_candidates=merge_candidates,
        merge_steps=merge_steps,
        config_table=config_table,
        metric_dictionary=metric_dictionary,
    )

    _t = _log_step("7/9 写入 Excel 报告", _t)
    _t = _log_step("8/9 报告格式化完成", _t)
    _log_step(f"9/9 完成 => {REPORT_PATH}", _t)


if __name__ == "__main__":
    main()

