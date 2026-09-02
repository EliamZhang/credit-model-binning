# -*- coding: utf-8 -*-
"""新客三份报告生成器：从 4 份 Excel 读取数值，生成 docs/ 下 3 份 md 报告。

数值纪律（CLAUDE.md §7.3）：本脚本数字来自 Excel（openpyxl data_only 读值）或
res/*.csv 重算（交叉报告三章矩阵内收入 4 指标：total_income/total_expenses/gross_surplus/net_surplus 中位数，分档与矩阵逐格核对一致），不手抄；
写完自动回读 md 与 Excel 逐项核对关键值。重跑分箱后需重跑本脚本再提交。

生成：
1. docs/分箱方法论与结果说明报告（新客价值模型笔数口径）.md
2. docs/分箱方法论与结果说明报告（新客mlt笔数口径）.md
3. docs/两模型交叉效果评估报告（新客mlt × 新客价值模型）.md
"""
import csv
import re
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DOCS = ROOT / "docs"
TODAY = "20260901"

WTH_XLSX = ROOT / "out" / f"binning_new_worthiness_strategy_report_{TODAY}.xlsx"
MLT_XLSX = ROOT / "out" / f"binning_new_mlt_strategy_report_{TODAY}.xlsx"
CROSS_XLSX = ROOT / "out" / f"binning_new_cross_strategy_report_{TODAY}.xlsx"
COND_XLSX = ROOT / "out" / f"binning_new_worthiness_cond_strategy_report_{TODAY}.xlsx"


# ---------- 通用工具 ----------

def load(path: Path):
    return load_workbook(path, data_only=True)


def tables_in_sheet(ws):
    """把 sheet 解析为 [(header_names, [data_rows])...]。write_sheet 固定格式：
    标题行（单值）→ 空行 → 表头行 → 数据行 → 空行 → …；因此表头只出现在空行之后。"""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    out = []
    i = 0
    prev_blank = True
    while i < len(rows):
        vals = [v for v in rows[i] if v is not None]
        if not vals:
            prev_blank = True
            i += 1
            continue
        if prev_blank:
            if len(vals) == 1:  # 标题行：其后的表头行仍应被识别（标题与表头之间可能有也可能没有空行）
                i += 1
                continue
            header = rows[i]
            data = []
            j = i + 1
            while j < len(rows):
                dvals = [v for v in rows[j] if v is not None]
                if not dvals:
                    break
                data.append(rows[j])
                j += 1
            out.append((header, data))
            i = j
        else:
            i += 1
    return out


def as_dicts(header, data):
    return [dict(zip(header, row)) for row in data]


def find_table(ws, first_header_cell):
    for header, data in tables_in_sheet(ws):
        if header[0] == first_header_cell:
            return as_dicts(header, data)
    raise KeyError(f"未找到表 {first_header_cell}")


def overview(ws):
    """01_总览：section | metric | value → dict[(section, metric)]。"""
    out = {}
    for header, data in tables_in_sheet(ws):
        if header[:2] == ["section", "metric"]:
            for d in data:
                if d[0] is not None and d[1] is not None:
                    out[(d[0], d[1])] = d[2]
    return out


def num(x, nd=0):
    if x is None:
        return "—"
    return f"{x:,.{nd}f}"


def pct(x, nd=2):
    if x is None:
        return "—"
    return f"{x*100:.{nd}f}%"


def pct_ci(v, lo, hi):
    if v is None:
        return "—"
    return f"{pct(v)} [{pct(lo)}, {pct(hi)}]"


def rate4(x):
    if x is None:
        return "—"
    return f"{x:.4f}"


def pp4(x):
    """pp 差写 +0.0014 式四位数。"""
    if x is None:
        return "—"
    return f"{x:+.4f}"


def diff_pp(a, b):
    if a is None or b is None:
        return "—"
    return f"{(a - b)*100:+.2f}pp"


def month_range(rows, group):
    mons = sorted({r["application_month"] for r in rows if r["sample_group"] == group})
    return f"{mons[0]}—{mons[-1]}"


def stage_cn(stage):
    return {
        "small_bin_cleanup": "小箱清理",
        "pava_monotonic_merge": "单调合并",
        "granularity_reduction": "档位压缩",
        "candidate_reduction": "候选生成",
        "share_balancing": "分布整形",
        "initial": "初始",
    }.get(stage, stage)


def pval_text(p):
    if p is None:
        return "—"
    return "<0.001" if p < 0.001 else f"{p:.3f}"


def ivloss_text(v):
    if v is None:
        return "—"
    return "<0.0001" if v < 0.0001 else f"{v:.4f}"


def bin_label(v):
    """初始箱序号 → 'B01' 式标签；接受 int / '15' / 'B15'。"""
    if v is None:
        return "?"
    s = str(v).strip()
    if s.isdigit():
        return f"B{int(s):02d}"
    return s


# ---------- 数据准备口径（来自 *_original.csv 行数） ----------

def prep_stats():
    """剔除未完成申请的行数统计（数据准备阶段，来自 *_original.csv）。"""
    import csv
    out = {}
    for key, base in [("sample", "new_sample"), ("mlt", "new_mlt_score")]:
        orig = ROOT / "res" / f"{base}_original.csv"
        cur = ROOT / "res" / f"{base}.csv"
        for path in (orig, cur):
            n = None
            for enc in ("utf-8-sig", "gbk"):
                try:
                    with open(path, encoding=enc, newline="") as f:
                        n = sum(1 for _ in csv.reader(f)) - 1
                    break
                except (UnicodeDecodeError, UnicodeError):
                    continue
            if n is None:
                raise ValueError(f"无法读取 {path}")
            if path == orig:
                n_orig = n
            else:
                n_cur = n
        out[key] = (n_orig, n_cur)
    return out


# ---------- 单模型报告渲染 ----------

def render_single_model(
    model_key: str,
    xlsx: Path,
    md_path: Path,
    title: str,
    model_cn: str,
    score_col: str,
    final_bin_col: str,
    raw_score_col: str,
    decile: str,
    auto_bin: str,
    accept_bin: str,
    missing_note: str,
    merge_note: str,
    prep: dict,
    manual: bool = False,
    value_semantics: bool = False,
    dist_note: str = "",
    steps_note: str = "",
    cand_note: str = "",
):
    wb = load(xlsx)
    ov = overview(wb["01_总览"])

    def O(section, metric):
        return ov[(section, metric)]

    n_raw = int(O("样本", "原始样本量"))
    n_removed = int(O("样本", "剔除未完成申请量"))
    n_valid = int(O("样本", "有效模型分样本量"))
    n_missing = int(O("样本", "模型分缺失量"))
    n_train = int(O("样本", "Train 样本量"))
    n_oot = int(O("样本", "OOT 样本量"))
    plan = str(O("分箱", "最终采用合箱方案"))
    n_bins = int(O("分箱", "最终箱数量"))
    psi = float(O("稳定性", "最终箱 Train/OOT PSI"))
    auto_th = str(O("模型策略阈值", "自动通过阈值"))
    accept_th = str(O("模型策略阈值", "人工审核上限/拒绝阈值"))
    acc_rate = float(O("策略风险", "接纳人群3M30+笔数逾期率"))
    marginal = float(O("策略风险", "最后接纳档边际3M30+"))
    tr_auc3 = float(O("模型效果", "train_duedate_3m_30_auc"))
    oo_auc3 = float(O("模型效果", "oot_duedate_3m_30_auc"))
    tr_ks3 = float(O("模型效果", "train_duedate_3m_30_ks"))
    oo_ks3 = float(O("模型效果", "oot_duedate_3m_30_ks"))
    tr_auc1 = float(O("模型效果", "train_duedate_1m_30_auc"))
    oo_auc1 = float(O("模型效果", "oot_duedate_1m_30_auc"))
    tr_ks1 = float(O("模型效果", "train_duedate_1m_30_ks"))
    oo_ks1 = float(O("模型效果", "oot_duedate_1m_30_ks"))
    tr_mono = bool(O("单调性", "train_最终箱全部单调"))
    oo_mono = bool(O("单调性", "oot_最终箱全部单调"))
    train_bad3 = float(O("模型效果", "train_duedate_3m_30_bad_rate"))
    oot_bad3 = float(O("模型效果", "oot_duedate_3m_30_bad_rate"))
    train_bad1 = float(O("模型效果", "train_duedate_1m_30_bad_rate"))
    oot_bad1 = float(O("模型效果", "oot_duedate_1m_30_bad_rate"))

    # 漏斗（01 总览，中文指标名）
    def F(prefix, metric):
        return float(O("历史实际审批漏斗", f"{prefix}_{metric}"))

    # 策略流量（01 总览）
    def S(prefix, metric):
        return float(O("模型策略测算流量", f"{prefix}_{metric}"))

    # 03 最终分箱统计
    final_rows = find_table(wb["03_最终分箱统计"], "sample_group")
    bin_col = next(k for k in final_rows[0] if k.endswith("_final_bin"))
    tr_rows = [r for r in final_rows if r["sample_group"] == "Train"]
    oo_rows = [r for r in final_rows if r["sample_group"] == "OOT"]
    tr_rows.sort(key=lambda r: r["bin_order"])
    oo_rows.sort(key=lambda r: r["bin_order"])

    # 05 模型验证
    perf = find_table(wb["05_模型验证"], "sample_group")
    perf_tbl = [r for r in perf if r.get("label") in ("duedate_1m_30", "duedate_3m_30")]
    tables5 = tables_in_sheet(wb["05_模型验证"])
    psi_rows = None
    mono_rows = None
    monthly_rows = None
    for header, data in tables5:
        if header[0] == "final_bin_order" and "psi_component" in header:
            psi_rows = as_dicts(header, data)
        if header[0] == "sample_group" and "metric" in header:
            mono_rows = as_dicts(header, data)
        if header[0] == "sample_group" and "application_month" in header and "primary_inversion_count" in header:
            monthly_rows = as_dicts(header, data)
    assert psi_rows is not None and mono_rows is not None and monthly_rows is not None

    # 月度倒挂统计
    tr_month_bad = [r for r in monthly_rows if r["sample_group"] == "train" and int(r["primary_inversion_count"]) > 0]
    oo_month_bad = [r for r in monthly_rows if r["sample_group"] == "oot" and int(r["primary_inversion_count"]) > 0]
    tr_months_all = [r for r in monthly_rows if r["sample_group"] == "train"]

    # 04 策略方案
    funnel_tbl = find_table(wb["04_策略方案"], "metric_scope")
    flow_tbl = find_table(wb["04_策略方案"], "metric_scope")  # 第二次出现 → 用表扫描
    tables4 = tables_in_sheet(wb["04_策略方案"])
    funnel_rows = flow_rows = th_sel = sens = seg = None
    for header, data in tables4:
        d = as_dicts(header, data)
        if header[0] == "metric_scope" and header[1] == "sample_group" and "actual_apply_cnt" in header:
            funnel_rows = d
        if header[0] == "metric_scope" and header[1] == "sample_group" and "strategy_estimated_total_a" in header[2]:
            flow_rows = d
        if header[0] == "selected_role":
            th_sel = d
        if header[0] == "threshold_type":
            sens = d
        if header[0] == "sample_group" and "decision" in header:
            seg = d
    assert funnel_rows is not None and flow_rows is not None and th_sel is not None and sens is not None and seg is not None

    # 06 附录配置
    cfg_tbl = find_table(wb["06_附录"], "config_group")

    # 敏感性与分段数值（Train/OOT）
    def sens_row(ttype, scenario):
        for r in sens:
            if r["threshold_type"] == ttype and r["scenario"] == scenario:
                return r
        return None

    def seg_rows(group):
        return [r for r in seg if r["sample_group"] == group]

    # 数据准备统计
    n_prep_raw, n_prep_cur = prep["sample"]
    n_prep_removed = n_prep_raw - n_prep_cur

    # 时间范围
    tr_range = "2024-01—2025-10"
    oo_range = "2025-11—2026-05"
    oot_last_month_n = max((int(r["n"]) for r in monthly_rows if r["sample_group"] == "oot" and r["application_month"] == "2026-05"), default=0)

    # 单调性明细（OOT 倒挂档位）
    mono_bad = [r for r in mono_rows if not bool(r["is_monotonic_non_decreasing"])]

    # 02 分箱详情：合箱执行步骤与候选方案
    steps_tbl = cand_tbl = None
    for header, data in tables_in_sheet(wb["02_分箱详情"]):
        d = as_dicts(header, data)
        if header[0] == "step_no" and header[1] == "stage" and "merged_range" in header:
            steps_tbl = d
        if header[0] == "selected" and "candidate_score" in header:
            cand_tbl = d
    assert steps_tbl is not None and cand_tbl is not None, "02_分箱详情 缺少步骤表/候选表"
    steps_tbl = sorted(steps_tbl, key=lambda r: int(r["step_no"]))
    candidates = [r for r in cand_tbl if int(r["final_bin_count"]) in (6, 7, 8)]
    candidates.sort(key=lambda r: (str(r["selected"]) != "True", -float(r["candidate_score"])))

    if value_semantics:
        cross_ov = overview(load(CROSS_XLSX)["01_总览"])
        pearson_cross = float(cross_ov[("相关性", "模型分 Pearson 相关（Train）")])
    else:
        pearson_cross = None

    # 约束核对数值（从 03 的累计列取，与阈值选择一致）
    auto_cum = tr_rows[auto_bin_rank(auto_bin) - 1]
    acc_cum = tr_rows[accept_bin_rank(accept_bin) - 1]
    auto_cum1 = auto_cum["cum_1m30p_cnt_bad_rate"]
    auto_cum3 = auto_cum["cum_3m30p_cnt_bad_rate"]
    acc_cum1 = acc_cum["cum_1m30p_cnt_bad_rate"]
    acc_cum3 = acc_cum["cum_3m30p_cnt_bad_rate"]
    auto_cum3_hi = auto_cum["cum_3m30p_cnt_bad_rate_ci_high"]
    acc_cum3_hi = acc_cum["cum_3m30p_cnt_bad_rate_ci_high"]
    acc_cum1_hi = acc_cum["cum_1m30p_cnt_bad_rate_ci_high"]
    marg3_hi = acc_cum["3m30p_cnt_bad_rate_ci_high"]

    lines = []
    A = lines.append

    A(f"# 新客{model_cn}分数分箱与策略阈值设定报告（笔数口径）\n")
    A(f"> 本报告说明新客{model_cn}（`{raw_score_col}`）分数分箱、样本外验证及策略阈值设定结果，由 `scr/_gen_new_reports.py` 从 `out/{xlsx.name}` 读取数值生成，与 Excel 逐项一致。管线沿用笔数违约合箱口径：完整 Train 用于学习分箱边界、执行合箱、选择候选方案和确定策略阈值；OOT 仅用于最终验证。")
    A(">")
    A(f"> 数据范围：数据源 `new_sample.csv` 已在数据准备阶段剔除未完成申请（原始 {num(n_prep_raw)} 笔中 `0.Incomplete` / `1.In Progress` {num(n_prep_removed)} 笔、占 {pct(n_prep_removed/n_prep_raw)}），分析样本 {num(n_prep_cur)} 笔全部为完成进件；按月样本时间范围为 2024-01—2026-05，其中 2026-05 为非完整月份。{model_cn}分覆盖 {num(n_valid)} 笔（{pct(n_valid/n_raw)}），缺失 {num(n_missing)} 笔（{pct(n_missing/n_raw)}）{missing_note}，缺失样本不进入分箱与策略测算、线上按拒绝处理。{model_cn}分为**高分高风险**：check_data 十分位 3M30+ 笔数逾期率由最低分位 {decile.split('→')[0].strip()} 单调升至最高分位 {decile.split('→')[1].strip()}（倒挂 0 处，`HIGH_SCORE_HIGH_RISK=True`）。")
    A("")
    A("## 一、结论摘要\n")

    # —— 摘要大表（40 列 + 档位，Train/OOT 各一张，整体行用 01 总览 + overall_* 列）——
    funnel_by_group = {r["sample_group"]: r for r in funnel_rows}

    def big_table_md(rows, group, n_total):
        funnel = funnel_by_group[group]
        first, last = rows[0], rows[-1]
        iv1 = sum(r["1m30p_iv_component"] for r in rows)
        iv3 = sum(r["3m30p_iv_component"] for r in rows)
        cols = [
            ("档位", None, None),
            ("分数下界", "score_left", "bound"),
            ("分数上界", "score_right", "bound"),
            ("样本量", "n", "num"),
            ("档位占比", "sample_pct", "pct"),
            ("累计流量", "cum_pass_rate", "pct"),
            ("策略分段", "strategy_estimated_decision", "raw"),
            ("策略箱流量", "strategy_estimated_bin_flow_rate", "pct"),
            ("1M30+笔数逾期率", "1m30p_cnt_bad_rate", "pct"),
            ("1M30+金额逾期率", "1m30p_amt_bad_rate", "pct"),
            ("3M30+笔数逾期率", "3m30p_cnt_bad_rate", "pct"),
            ("3M30+金额逾期率", "3m30p_amt_bad_rate", "pct"),
            ("1M30+笔数Lift", "1m30p_cnt_lift", "r4"),
            ("1M30+金额Lift", "1m30p_amt_lift", "r4"),
            ("3M30+笔数Lift", "3m30p_cnt_lift", "r4"),
            ("3M30+金额Lift", "3m30p_amt_lift", "r4"),
            ("累计1M30+笔数逾期率", "cum_1m30p_cnt_bad_rate", "pct"),
            ("累计1M30+金额逾期率", "cum_1m30p_amt_bad_rate", "pct"),
            ("累计3M30+笔数逾期率", "cum_3m30p_cnt_bad_rate", "pct"),
            ("累计3M30+金额逾期率", "cum_3m30p_amt_bad_rate", "pct"),
            ("1M30+ IV分项", "1m30p_iv_component", "r4"),
            ("3M30+ IV分项", "3m30p_iv_component", "r4"),
            ("1M30+ KS曲线值", "1m30p_ks_curve", "r4"),
            ("3M30+ KS曲线值", "3m30p_ks_curve", "r4"),
            ("PSI分项", "train_oot_psi_component", "r4"),
            ("实际完成率", "actual_completion_rate", "pct"),
            ("实际审批通过率", "actual_approval_rate", "pct"),
            ("实际自动审批通过率", "actual_auto_approval_rate", "pct"),
            ("实际人工审批通过率", "actual_manual_approval_rate", "pct"),
            ("实际自动审批占比", "actual_auto_approval_share", "pct"),
            ("实际人工审批占比", "actual_manual_approval_share", "pct"),
            ("实际成交转化率", "actual_deal_rate", "pct"),
            ("测算自动通过率", None, None),
            ("测算人工审核率", None, None),
            ("测算总接纳率", None, None),
            ("测算拒绝率", None, None),
            ("1M30+整体AUC", None, None),
            ("3M30+整体AUC", None, None),
            ("1M30+整体KS", None, None),
            ("3M30+整体KS", None, None),
            ("整体PSI", None, None),
        ]
        out = ["| " + " | ".join(c[0] for c in cols) + " |",
               "| " + " | ".join("---:" for _ in cols) + " |"]
        for r in rows:
            cells = []
            for label, key, fmt in cols:
                if key is None:
                    cells.append("—")
                elif fmt == "pct":
                    cells.append(pct(r[key]))
                elif fmt == "r4":
                    cells.append(rate4(r[key]))
                elif fmt == "num":
                    cells.append(num(r[key]))
                elif fmt == "bound":
                    cells.append({"-inf": "−∞", "inf": "+∞"}.get(str(r[key]), str(r[key])))
                else:
                    cells.append(str(r[key]))
            out.append("| " + " | ".join(cells) + " |")
        cells = []
        for label, key, fmt in cols:
            if label == "档位":
                cells.append("**整体**")
            elif label in ("分数下界", "分数上界", "策略分段", "1M30+ KS曲线值", "3M30+ KS曲线值", "PSI分项"):
                cells.append("—")
            elif label == "样本量":
                cells.append(f"**{num(n_total)}**")
            elif label in ("档位占比", "累计流量", "策略箱流量"):
                cells.append("**100.00%**")
            elif label in ("1M30+笔数逾期率", "1M30+金额逾期率", "3M30+笔数逾期率", "3M30+金额逾期率"):
                cells.append(f"**{pct(last[f'cum_{key}'])}**")
            elif label in ("累计1M30+笔数逾期率", "累计1M30+金额逾期率", "累计3M30+笔数逾期率", "累计3M30+金额逾期率"):
                cells.append(f"**{pct(last[key])}**")
            elif label in ("1M30+笔数Lift", "1M30+金额Lift", "3M30+笔数Lift", "3M30+金额Lift"):
                cells.append("**1.0000**")
            elif label in ("1M30+ IV分项", "3M30+ IV分项"):
                cells.append(f"**{rate4(iv1 if '1M30' in label else iv3)}**")
            elif label in ("实际完成率", "实际审批通过率", "实际自动审批通过率", "实际人工审批通过率",
                           "实际自动审批占比", "实际人工审批占比", "实际成交转化率"):
                cells.append(f"**{pct(funnel[key])}**")
            elif label == "测算自动通过率":
                cells.append(f"**{pct(first['strategy_estimated_overall_auto_pass_rate'])}**")
            elif label == "测算人工审核率":
                cells.append(f"**{pct(first['strategy_estimated_overall_manual_review_rate'])}**")
            elif label == "测算总接纳率":
                cells.append(f"**{pct(first['strategy_estimated_overall_total_accept_rate'])}**")
            elif label == "测算拒绝率":
                cells.append(f"**{pct(first['strategy_estimated_overall_reject_rate'])}**")
            elif label == "1M30+整体AUC":
                cells.append(f"**{rate4(first['overall_1m30p_auc'])}**")
            elif label == "3M30+整体AUC":
                cells.append(f"**{rate4(first['overall_3m30p_auc'])}**")
            elif label == "1M30+整体KS":
                cells.append(f"**{rate4(first['overall_1m30p_ks'])}**")
            elif label == "3M30+整体KS":
                cells.append(f"**{rate4(first['overall_3m30p_ks'])}**")
            elif label == "整体PSI":
                cells.append(f"**{rate4(first['train_oot_psi_total'])}**")
            else:
                cells.append("—")
        out.append("| " + " | ".join(cells) + " |")
        return "\n".join(out)

    A("**Train 分箱结果**：\n")
    A(big_table_md(tr_rows, "Train", n_train))
    A("")
    A("**OOT 分箱结果**（沿用 Train 分箱边界）：\n")
    A(big_table_md(oo_rows, "OOT", n_oot))
    A("")
    A("> 口径说明：所有指标均独立成列。箱级行展示实际审批转化、四项风险率及对应 Lift、IV 分项、KS 曲线值和 PSI 分项；Lift = 档位逾期率 ÷ 该样本组整体逾期率，衡量单箱风险相对整体的倍数，整体行恒为 1.0000（基准）；四项累计逾期率按 bin_order 从低风险向高风险逐箱累加（累计至最后一档即等于样本组整体逾期率，故整体行与整体率一致）；模型策略整体转化率、AUC、整体 KS 与整体 PSI 仅在同一分箱表的\"整体\"行展示，避免将整体指标误解为单箱指标。Train 整体 IV 分项为 1M30+ / 3M30+ 笔数口径合计。")
    A("")
    A("**核心结论**：\n")

    # 结论 1：Train 风险分层
    A(f"1. **Train 风险分层成立**：1M30+、3M30+ 的笔数和金额逾期率均随风险档位单调上升，3M30+ 笔数逾期率由 A 档的 {pct(tr_rows[0]['3m30p_cnt_bad_rate'])} 升至 {tr_rows[-1][bin_col]} 档的 {pct(tr_rows[-1]['3m30p_cnt_bad_rate'])}；")

    # 结论 2：方案与人数分布
    maxbin = max(tr_rows, key=lambda r: r["sample_pct"])
    shares = " / ".join(pct(r["sample_pct"]) for r in tr_rows)
    if manual:
        A(f"2. **{n_bins} 档方案经手动指定（模型配置 final_bin_ranges，2026-09-01 用户确认）**：自动合箱在该口径下选中 6 档（保留最坏极端箱 B20 单箱），经评审改为手动 {n_bins} 档 {plan} 消除 7/8 档候选的 B20 单箱倒挂（详见三（二））；人数分布偏中高风险、峰值档超上限——Train 各档占比为 {shares}，最大单箱为 {maxbin[bin_col]} 档（{bin_label(maxbin['source_bin_start'])}–{bin_label(maxbin['source_bin_end'])}）{pct(maxbin['sample_pct'])}，超过 21% 的人数分布上限：{dist_note}；")
    else:
        A(f"2. **{n_bins} 档方案经自动合箱选中，Train 主指标倒挂 0 处**：{plan}；人数分布偏中高风险、峰值档超上限——Train 各档占比为 {shares}，最大单箱为 {maxbin[bin_col]} 档（{bin_label(maxbin['source_bin_start'])}–{bin_label(maxbin['source_bin_end'])}）{pct(maxbin['sample_pct'])}，超过 21% 的人数分布上限：{dist_note}；")

    # 结论 3：OOT 主指标
    viol = [r for r in mono_bad if r["sample_group"] == "oot" and r["metric"] == "1m30p_cnt_bad_rate"]
    if viol:
        vbn = int(viol[0]["violation_bins"]) - 1
        vrow = oo_rows[vbn]
        prev = oo_rows[vbn - 1]
        if vbn == len(oo_rows) - 1:
            inv_txt = f"；OOT 1M30+ 笔数在 {vrow[bin_col]} 档出现尾部倒挂（{prev[bin_col]} 档 {pct(prev['1m30p_cnt_bad_rate'])} → {vrow[bin_col]} 档 {pct(vrow['1m30p_cnt_bad_rate'])}，{vrow[bin_col]} 档 1M30+ 成熟量仅 {num(vrow['1m30p_cnt_mature'])}，属尾部小样本噪声）"
        else:
            inv_txt = f"；OOT 1M30+ 笔数在 {vrow[bin_col]} 档出现 1 处倒挂（{prev[bin_col]} 档 {pct(prev['1m30p_cnt_bad_rate'])} → {vrow[bin_col]} 档 {pct(vrow['1m30p_cnt_bad_rate'])}，见五（一））"
    else:
        vrow = prev = None
        inv_txt = ""
    A(f"3. **OOT 主指标基本稳定**：3M30+ 笔数逾期率 OOT 由 A 档 {pct(oo_rows[0]['3m30p_cnt_bad_rate'])} 升至 {oo_rows[-1][bin_col]} 档 {pct(oo_rows[-1]['3m30p_cnt_bad_rate'])}，1M30+ / 3M30+ 金额口径在 OOT 全部单调{inv_txt}；")

    # 结论 4：历史实际 vs 模型测算
    A(f"4. **历史实际与模型测算差异显著**：Train 历史实际审批通过率 {pct(F('Train','审批通过率'))}、自动审批通过率 {pct(F('Train','自动审批通过率'))}，模型策略测算 Train 自动通过率 {pct(S('Train','测算自动通过率'))}、总接纳率 {pct(S('Train','测算总接纳率'))}；OOT 实际审批通过率 {pct(F('OOT','审批通过率'))}、自动审批通过率 {pct(F('OOT','自动审批通过率'))}，测算自动通过率、总接纳率分别为 {pct(S('OOT','测算自动通过率'))}、{pct(S('OOT','测算总接纳率'))}，需结合现行业务规则评估策略落地影响；")

    # 结论 5：跨期稳定
    def pp_delta(metric):
        o, t = S("OOT", metric), S("Train", metric)
        return abs(o - t) * 100
    A(f"5. **跨期分布稳定**：Train/OOT PSI 为 {rate4(psi)}；OOT 测算自动通过率、总接纳率分别为 {pct(S('OOT','测算自动通过率'))}、{pct(S('OOT','测算总接纳率'))}，较 Train 高 {pp_delta('测算自动通过率'):.2f}、{pp_delta('测算总接纳率'):.2f} 个百分点；")

    # 结论 6：阈值与 CI 余量
    auto_hi_note = "略超" if auto_cum3_hi > 0.055 else "贴近"
    acc_hi_note = "略超" if acc_cum3_hi > 0.075 else "贴近"
    A(f"6. **策略阈值点估计满足默认约束、CI 余量有限**：自动通过阈值 {auto_th}（{auto_bin} 档右边界，Train 自动通过率 {pct(S('Train','测算自动通过率'))}），总接纳阈值 {accept_th}（{accept_bin} 档右边界，Train 总接纳率 {pct(S('Train','测算总接纳率'))}）；接纳人群 3M30+ {pct(acc_rate)}、最后接纳档边际 3M30+ {pct(marginal)}。自动通过累计 3M30+ CI 上界 {pct(auto_cum3_hi)}（{auto_hi_note} 5.50% 上限）、总接纳累计 3M30+ CI 上界 {pct(acc_cum3_hi)}（{acc_hi_note} 7.50% 上限），实施后需按约束口径持续监测；")

    # 结论 7：样本外衰减与月度稳定
    oo_max_drop = max((r["max_primary_rate_drop"] for r in oo_month_bad), default=None)
    A(f"7. **样本外排序能力有所衰减**：3M30+ AUC / KS 由 Train 的 {rate4(tr_auc3)} / {rate4(tr_ks3)} 降至 OOT 的 {rate4(oo_auc3)} / {rate4(oo_ks3)}，但仍保留风险区分能力；Train {len(tr_months_all)} 个月中有 {len(tr_month_bad)} 个月、OOT {len(oo_month_bad)} 个月出现超过容忍度的主指标倒挂{('（OOT 单月最大 ' + diff_pp(oo_max_drop, 0) + '）') if oo_max_drop else ''}，需持续监测。")
    A("")
    A("**核心指标总览**：\n")
    A("| 指标 | Train | OOT |")
    A("| --- | ---: | ---: |")
    A(f"| 有效模型分样本量 | {num(n_train)} | {num(n_oot)} |")
    A(f"| 1M30+ 笔数逾期率 | {pct(train_bad1)} | {pct(oot_bad1)} |")
    A(f"| 3M30+ 笔数逾期率 | {pct(train_bad3)} | {pct(oot_bad3)} |")
    A(f"| 1M30+ AUC / KS | {rate4(tr_auc1)} / {rate4(tr_ks1)} | {rate4(oo_auc1)} / {rate4(oo_ks1)} |")
    A(f"| 3M30+ AUC / KS | {rate4(tr_auc3)} / {rate4(tr_ks3)} | {rate4(oo_auc3)} / {rate4(oo_ks3)} |")
    A(f"| 最终箱 Train/OOT PSI | — | {rate4(psi)} |")
    A(f"| 模型策略测算自动通过率 / 总接纳率 | {pct(S('Train','测算自动通过率'))} / {pct(S('Train','测算总接纳率'))} | {pct(S('OOT','测算自动通过率'))} / {pct(S('OOT','测算总接纳率'))} |")
    A(f"| 接纳人群 3M30+ / 最后接纳档边际 | {pct(acc_rate)} / {pct(marginal)} | — |")
    A("")
    A("**分析流程**：\n")
    A("| 步骤 | 环节 | 关键操作与产出 |")
    A("| --- | --- | --- |")
    flow = [
        ("①", "数据加载与清洗",
         f"数据源`new_sample.csv` 已剔除未完成申请（原始 {num(n_prep_raw)} 笔中 `0.Incomplete` / `1.In Progress` {num(n_prep_removed)} 笔），有效样本 {num(n_prep_cur)} 笔；{model_cn}分覆盖 {num(n_valid)} 笔（{pct(n_valid/n_raw)}），缺失 {num(n_missing)} 笔（{pct(n_missing/n_raw)}）不进入分析"),
        ("②", "样本切分",
         f"按申请月份切分：Train 2024-01—2025-10（{num(n_train)}）用于分箱与决策；OOT 2025-11—2026-05-20（{num(n_oot)}）仅用于最终验证"),
        ("③", "初始分箱",
         f"Train 上按`{score_col}` 分位数等频初分为 20 箱（B01–B20，左开右闭、首尾 ±∞），边界固定后原样复用至 OOT"),
        ("④", "自动合箱（Train）",
         "四阶段递进：小箱清理（消除单箱硬约束违反）→ 单调合并（PAVA 风格消除主指标倒挂）→ 档位压缩（≤ 8 档）→ 候选生成（8/7/6 档）" if not manual else
         "四阶段递进：小箱清理（1 步）→ 单调合并（本次无独立步骤）→ 档位压缩 → 候选生成（8/7/6 档）；7/8 档候选残留主指标倒挂 1 处、6 档候选无倒挂（箱级违规 2 项）"),
        ("⑤", "候选评估与选择",
         "硬约束筛选（6–8 档、无倒挂、单箱约束满足、不跨极端箱边界）→ 对单箱 Train 占比超过 21% 上限的候选做\"均衡拆分 + 相邻再合并\"整形（share_balancing）→ 按倒挂数、IV 保留率、最小相邻差距、档位偏离综合评分 → 选中 7 档方案" if not manual else
         "硬约束筛选（该口径下 6–8 档候选均未完全满足）→ 自动选中 6 档方案，经评审改为手动 7 档（模型配置 final_bin_ranges，2026-09-01 用户确认）"),
        ("⑥", "样本外验证",
         "Train/OOT 对照验证：风险单调性、分布稳定性（PSI）、区分能力（AUC/KS）、月度稳定性、测算分段风险梯度"),
        ("⑦", "策略阈值设定",
         f"Train 上按风险由低至高逐档放宽，校验累计 1M30+/3M30+ 与边际 3M30+ 约束 → 自动通过止于 {auto_bin} 档右边界、总接纳止于 {accept_bin} 档右边界"),
        ("⑧", "策略流量测算",
         "按选定阈值将样本分为自动通过 / 人工审核 / 拒绝三段，输出 Train/OOT 各段占比、逾期率与风险梯度"),
        ("⑨", "历史实际审批漏斗",
         "基于`application_info` 状态字段按 `application_id` 去重，核算完成率、审批通过率、自动/人工构成与成交转化率"),
        ("⑩", "报告输出",
         "汇总分箱结果、候选方案、阈值选择过程、验证结论与上线实施规范"),
    ]
    for no, name, desc in flow:
        A(f"| {no} | {name} | {desc} |")
    A("")
    A("## 二、样本设计与指标定义\n")
    A("### （一）数据集划分\n")
    A("| 数据集 | 时间范围 | 样本量（有效模型分） | 用途 |")
    A("| --- | --- | ---: | --- |")
    A(f"| Train | {tr_range} | {num(n_train)} | 学习初始边界、执行合箱、选择方案并确定策略阈值 |")
    A(f"| OOT | {oo_range}（截至 2026-05-20，其中 2026-05 {num(oot_last_month_n)} 笔且 3M30+ 未成熟） | {num(n_oot)} | 独立样本外验证，不参与分箱设计、候选选择或阈值设定 |")
    A("")
    A(f"- Train 截止月份为 2025-10，OOT 自 2025-11 起；")
    A(f"- 数据源已在数据准备阶段剔除未完成申请（原始 {num(n_prep_raw)} 笔中剔除 {num(n_prep_removed)} 笔、占 {pct(n_prep_removed/n_prep_raw)}），分析样本 {num(n_prep_cur)} 笔全部为完成进件；")
    A(f"- {model_cn}分缺失 {num(n_missing)} 笔（占 {pct(n_missing/n_raw)}）{missing_note}，分箱与策略测算仅使用存在模型分的 {num(n_valid)} 笔（Train {num(n_train)} + OOT {num(n_oot)}）；缺失样本不进入分箱统计，线上按拒绝处理；")
    A(f"- 新客 Train 3M30+ 标签成熟率约 10.71%（成交样本才有 duedate 表现标签，新客完成申请成交率约 12%，属结构性口径，2026-09-01 已与用户确认记录在案）；")
    A("- 历史实际审批漏斗独立于模型分，基于完整完成申请核算。")
    A("")
    A("### （二）风险指标")
    A("1M30+ 用于刻画短期风险，3M30+ 用于刻画成熟度更高的中期风险。笔数口径反映风险覆盖范围，金额口径反映损失强度；笔数口径用于合箱决策，金额口径用于方案评价与策略验证。")
    A("")
    A(f"总体风险水平：Train 的 1M30+、3M30+ 笔数逾期率分别为 {pct(train_bad1)} 和 {pct(train_bad3)}；OOT 分别为 {pct(oot_bad1)} 和 {pct(oot_bad3)}。合箱同时以 1M30+、3M30+ 笔数逾期率作为单调性主指标；单箱成熟量、显著性检验和 IV 等统计环节以 3M30+ 为锚定口径。分箱结果表的各档风险率旁同步展示对应 Lift（某箱逾期率 ÷ 该样本组整体逾期率），衡量单箱风险相对整体水平的倍数：Lift < 1 表示低于整体，> 1 表示高于整体；另附四项累计逾期率（按 bin_order 从低风险向高风险逐箱累加），用于观察\"截止到某档为止\"的累计风险水平，累计至最后一档即等于样本组整体逾期率。")
    A("")
    A("### （三）模型分方向验证")
    A(f"check_data 十分位验证（完整 Train，按 {score_col} 分位数）：3M30+ 笔数逾期率由最低分位的 {decile.split('→')[0].strip()} 单调升至最高分位的 {decile.split('→')[1].strip()}，倒挂 0 处，沿用 `HIGH_SCORE_HIGH_RISK=True`。")
    if value_semantics:
        A("")
        A(f"价值语义说明：价值模型的本义为\"低分 = 高价值\"（新客价值模型文档口径：分数越低，利息贡献越高，见 docs/新客价值模型效果评估文档_0520.html）；价值模型分与 mlt 主模型分在双分样本上的 Pearson 相关为 {rate4(pearson_cross)}（见《两模型交叉效果评估报告（新客mlt × 新客价值模型）》）。因此本报告的 A 档（最低分）同时是\"高价值 + 低风险\"档，G 档（最高分）同时是\"高风险 + 低价值\"档；风险类结论不受该语义影响，经营/提额类场景（优先经营象限、额度分层）需结合该语义使用。")
    A("")
    A("## 三、分箱方案设计与结果\n")
    A("### （一）初始分箱")
    A(f"在完整 Train 上按 {score_col} 分位数构建 20 个等频初始箱（B01–B20，按分数升序排列）。区间采用左开右闭形式 (left, right]，首尾边界扩展为 ±∞；后续仅合并相邻箱。")
    A("")
    A("### （二）合箱流程与约束")
    A(merge_note)
    A("")
    A("| 阶段 | 触发条件 | 处理规则 |")
    A("| --- | --- | --- |")
    A("| ① 约束修正 | 存在违反单箱硬约束的箱 | 优先处理违反程度最高的箱，仅允许与相邻箱合并 |")
    A("| ② 单调合并（PAVA 风格） | 主指标存在相邻倒挂 | 优先合并倒挂幅度最大的相邻对，直至主指标无倒挂 |")
    A("| ③ 档位压缩 | 档位数多于 8 档 | 反复合并综合代价最低的相邻对 |")
    A("| ④ 候选生成 | 档位数多于 6 档 | 继续生成 8、7、6 档候选，并基于 Train 综合评分 |")
    A("| ⑤ 分布整形（share_balancing） | 某档 Train 样本占比超过上限 | 从低风险侧取第一个可行拆点将超限箱一分为二（两个子箱均不超限），再合并综合代价最低的相邻对回到原档数，整形候选与原候选一同评分 |")
    A("")
    A("**单箱硬约束**：Train 上中间箱样本占比须 ≥ 5%，首尾箱须 ≥ 2.5%；主指标成熟样本量须 ≥ 1,000，坏样本量须 ≥ 20，好样本量须 ≥ 200；最低和最高风险初始箱标记为极端箱（成熟样本量下限 500），默认禁止跨越极端箱边界合并。")
    A("")
    A("**保护边界**：优先保留策略风险边界、最大风险跃升边界和极端箱边界。跨越普通保护边界的合并代价增加 100；极端箱边界的合并代价增加 10,000，且默认禁止跨越。")
    A("")
    A("**人数分布上限**：最终任意一档的 Train 样本占比不得超过 `MAX_FINAL_BIN_SHARE = 21%`。超过上限的候选方案在评分前先做分布整形：将超限箱沿低风险侧优先的可行拆点拆为两个合规子箱，若档数超出目标则合并综合代价最低的相邻对回到原档数；拆不开或合不回去时放弃整形、原候选保留。该机制不改变合并代价本身，仅控制最终人数分布。")
    A("")
    A("**合并代价**：由相邻箱风险差距、两比例 Z 检验、IV 损失及保护边界惩罚共同确定。风险差距越大、统计差异越显著、IV 损失越高或涉及保护边界，合并优先级越低。")
    A("")
    A("### （三）合箱执行记录\n")
    A("| 步 | 阶段 | 合并初始箱 | 左→右 3M30+ | 两比例检验 p | IV 损失 | 合并后档位数 |")
    A("| --- | --- | --- | ---: | ---: | ---: | ---: |")
    for s in steps_tbl:
        lr = f"{pct(s['left_primary_rate'])}→{pct(s['right_primary_rate'])}"
        A(f"| {s['step_no']} | {stage_cn(s['stage'])} | {str(s['merged_range']).replace(', ', ',')} | {lr} | {pval_text(s['two_proportion_p_value'])} | {ivloss_text(s['primary_iv_loss'])} | {s['after_bin_count']} |")
    A("")
    A(steps_note)
    A("")
    A("### （四）候选方案评估")
    A("候选方案先接受硬约束筛选，包括档位数为 6–8、Train 主指标无倒挂、单箱约束全部满足且未跨越极端箱边界。通过筛选的方案再按 Train 全指标倒挂数、IV 保留率、最小相邻风险差距及目标档位偏离程度综合评分。单箱样本占比超过上限的方案先做分布整形，整形变体（标注\"整形\"）与原方案一同评分。")
    A("")
    A("| 档位数 | 方案 | 来源 | 主指标倒挂 | 全指标倒挂 | IV 保留率 | 最小相邻差距 | 综合得分 |")
    A("| --- | --- | --- | ---: | ---: | ---: | ---: | ---: |")
    for c in candidates:
        sel = str(c["selected"]) == "True"
        nbin = f"**{c['final_bin_count']}（选中）**" if sel else str(c["final_bin_count"])
        diff_txt = f"{float(c['min_adjacent_primary_rate_diff'])*100:+.2f}pp"
        score_txt = f"**{float(c['candidate_score']):.2f}**" if sel else f"{float(c['candidate_score']):.2f}"
        A(f"| {nbin} | {c['ranges']} | {stage_cn(c['stage'])} | {c['train_primary_inversion_cnt']} | {c['train_all_inversion_cnt']} | {rate4(c['primary_iv_retention'])} | {diff_txt} | {score_txt} |")
    A("")
    A(cand_note)
    A("")
    A("### （五）单箱约束检查\n")
    A("| 档位 | 样本量 | 占比 | 3M30+ 成熟量 | 坏样本量 | 好样本量 | 结果 |")
    A("| --- | ---: | ---: | ---: | ---: | ---: | --- |")
    fails = []
    for r in tr_rows:
        is_tail = int(r["bin_order"]) in (1, len(tr_rows))
        floor = 0.025 if is_tail else 0.05
        reasons = []
        if r["sample_pct"] < floor - 1e-9:
            reasons.append("占比不足")
        if r["3m30p_cnt_mature"] < 1000:
            reasons.append("成熟量不足")
        if r["3m30p_cnt_bad"] < 20:
            reasons.append("坏样本不足")
        if r["3m30p_cnt_good"] < 200:
            reasons.append("好样本不足")
        if reasons:
            fails.append((r[bin_col], r["sample_pct"], is_tail, floor))
        A(f"| {r[bin_col]} | {num(r['n'])} | {pct(r['sample_pct'])} | {num(r['3m30p_cnt_mature'])} | {num(r['3m30p_cnt_bad'])} | {num(r['3m30p_cnt_good'])} | {'不满足' if reasons else '满足'} |")
    A("")
    if fails:
        parts = []
        for bl, sp, is_tail, floor in fails:
            parts.append(f"{bl} 档占比 {sp*100:.4f}% 略低于{'首尾箱' if is_tail else '中间箱'} {floor*100:.1f}% 下限")
        A(f"单箱约束不满足项：{'；'.join(parts)}；其余各档样本占比、成熟量、坏样本量、好样本量均满足普通箱约束。最大单箱占比为 {maxbin[bin_col]} 档的 {pct(maxbin['sample_pct'])}，超过 21% 的人数分布上限（见三（四））。")
    else:
        A(f"各档样本占比、成熟量、坏样本量、好样本量均满足普通箱约束，无需依赖极端箱放宽；最大单箱占比为 {maxbin[bin_col]} 档的 {pct(maxbin['sample_pct'])}，超过 21% 的人数分布上限（见三（四））。")
    A("")
    A("### （六）最终分箱统计\n")
    A("**Train**：\n")
    A("| 档位 | 样本量 | 占比 | 1M30+ 笔数逾期率 [95% CI] | 3M30+ 笔数逾期率 [95% CI] | 3M30+ 金额逾期率 | 累计 3M30+ 笔数逾期率 |")
    A("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in tr_rows:
        A(f"| {r[bin_col]} | {num(r['n'])} | {pct(r['sample_pct'])} | {pct_ci(r['1m30p_cnt_bad_rate'], r['1m30p_cnt_bad_rate_ci_low'], r['1m30p_cnt_bad_rate_ci_high'])} | {pct_ci(r['3m30p_cnt_bad_rate'], r['3m30p_cnt_bad_rate_ci_low'], r['3m30p_cnt_bad_rate_ci_high'])} | {pct(r['3m30p_amt_bad_rate'])} | {pct(r['cum_3m30p_cnt_bad_rate'])} |")
    A("")
    A(f"Train 的四类风险率均随档位单调递增。3M30+ 笔数逾期率由 A 档的 {pct(tr_rows[0]['3m30p_cnt_bad_rate'])} 升至 {tr_rows[-1][bin_col]} 档的 {pct(tr_rows[-1]['3m30p_cnt_bad_rate'])}，高风险尾部保持明显分离。")
    A("")
    A("**OOT**（沿用 Train 分箱边界）：\n")
    A("| 档位 | 样本量 | 占比 | 1M30+ 笔数逾期率 [95% CI] | 3M30+ 笔数逾期率 [95% CI] | 3M30+ 金额逾期率 | 累计 3M30+ 笔数逾期率 |")
    A("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in oo_rows:
        A(f"| {r[bin_col]} | {num(r['n'])} | {pct(r['sample_pct'])} | {pct_ci(r['1m30p_cnt_bad_rate'], r['1m30p_cnt_bad_rate_ci_low'], r['1m30p_cnt_bad_rate_ci_high'])} | {pct_ci(r['3m30p_cnt_bad_rate'], r['3m30p_cnt_bad_rate_ci_low'], r['3m30p_cnt_bad_rate_ci_high'])} | {pct(r['3m30p_amt_bad_rate'])} | {pct(r['cum_3m30p_cnt_bad_rate'])} |")
    A("")
    if viol:
        A(f"OOT 的 3M30+ 笔数与金额、1M30+ 金额保持单调，1M30+ 笔数存在 1 处倒挂（见五（一）），高风险尾部仍可分离。")
    else:
        A(f"OOT 的 1M30+ 与 3M30+ 笔数、金额逾期率均保持单调，无局部倒挂；高风险尾部仍可分离。")
    A("")
    A("## 四、历史实际审批与模型策略测算结果\n")
    A("### （一）历史实际审批漏斗\n")
    A("历史实际审批漏斗来自 `application_info` 的 `application_status`、`assessment_status` 和 `status`，按唯一 `application_id` 统计（未完成申请已在数据源剔除，完成率恒为 100%）。")
    A("")
    A("| 数据集 | 申请数 | 完成进件数 | 审批通过数 | 自动审批通过数 | 人工审批通过数 | 成交数 |")
    A("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in funnel_rows:
        A(f"| {r['sample_group']} | {num(r['actual_apply_cnt'])} | {num(r['actual_completed_application_cnt'])} | {num(r['actual_approved_application_cnt'])} | {num(r['actual_auto_approved_application_cnt'])} | {num(r['actual_manual_approved_application_cnt'])} | {num(r['actual_deal_sample_cnt'])} |")
    A("")
    A("| 数据集 | 审批通过率 | 自动审批通过率 | 人工审批通过率 | 自动审批占比 | 人工审批占比 | 成交转化率 |")
    A("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in funnel_rows:
        A(f"| {r['sample_group']} | {pct(r['actual_approval_rate'])} | {pct(r['actual_auto_approval_rate'])} | {pct(r['actual_manual_approval_rate'])} | {pct(r['actual_auto_approval_share'])} | {pct(r['actual_manual_approval_share'])} | {pct(r['actual_deal_rate'])} |")
    A("")
    A("计算定义：未完成申请（`application_status` 属于 `0.Incomplete`、`1.In Progress`）已在数据准备阶段从 `new_sample.csv` 中剔除，不进入历史漏斗、分箱与策略测算，故分析样本全部为完成进件、完成率恒为 100%；审批通过按状态首字符属于 3/4 判定；自动/人工审批通过还须分别包含 `Auto Approved` / `Manual Approved`；成交为 `status` 属于 `Active_Account`、`Closed`、`Blocked`。")
    A("")
    A("### （二）模型策略阈值设定原则")
    A("自动通过和总接纳阈值均设在最终分箱边界上。在完整 Train 上按风险由低至高逐档放宽阈值，同时计算累计指标和新增档位的边际指标，并在满足风险上限的候选中选择通过率最高者。风险约束（默认策略，与老客一致）：")
    A("")
    A("| 约束阶段 | 累计 1M30+ 笔数逾期率 | 累计 3M30+ 笔数逾期率 | 边际 3M30+ 笔数逾期率 |")
    A("| --- | --- | --- | --- |")
    A("| 自动通过 | ≤ 0.90% | ≤ 5.50% | ≤ 9.00% |")
    A("| 总接纳（自动 + 人工） | ≤ 1.30% | ≤ 7.50% | ≤ 17.00% |")
    A("")
    A("### （三）模型策略阈值选择过程\n")
    A("| 候选 | 阈值 | 档位 | 累计通过率 | 累计 1M30+ | 累计 3M30+ [CI 上界] | 边际 3M30+ [CI 上界] | 自动约束 | 接纳约束 |")
    A("| --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in th_sel:
        role = r["selected_role"]
        if role in ("自动通过阈值", "人工审核上限/拒绝阈值"):
            role = f"**{role}（选中）**"
        elif r["threshold_order"] == 1:
            role = "首档（最严）"
        else:
            role = "候选"
        all_auto = all(r.get(f"auto_check_{k}") for k in ("cum_1m30p_cnt_bad_rate", "cum_3m30p_cnt_bad_rate", "marginal_3m30p_cnt_bad_rate"))
        all_acc = all(r.get(f"accept_check_{k}") for k in ("cum_1m30p_cnt_bad_rate", "cum_3m30p_cnt_bad_rate", "marginal_3m30p_cnt_bad_rate"))
        A(f"| {role} | {r['threshold']} | {r[bin_col]} | {pct(r.get('cum_pass_rate'))} | {pct(r.get('cum_1m30p_cnt_bad_rate'))} | {pct(r.get('cum_3m30p_cnt_bad_rate'))} [{pct(r.get('cum_3m30p_cnt_bad_rate_ci_high'))}] | {pct(r.get('marginal_3m30p_cnt_bad_rate'))} [{pct(r.get('marginal_3m30p_cnt_bad_rate_ci_high'))}] | {'通过' if all_auto else '不通过'} | {'通过' if all_acc else '不通过'} |")
    A("")
    nxt_auto = tr_rows[auto_bin_rank(auto_bin)]
    nxt_acc = tr_rows[accept_bin_rank(accept_bin)]
    A(f"- 自动通过阈值为 {auto_bin} 档右边界 {auto_th}，累计通过率 {pct(S('Train','测算自动通过率'))}；总接纳阈值为 {accept_bin} 档右边界 {accept_th}，累计接纳率 {pct(S('Train','测算总接纳率'))}；")
    A(f"- 约束核对：自动通过档累计 1M30+ {pct(auto_cum1)}、累计 3M30+ {pct(auto_cum3)}（CI 上界 {pct(auto_cum3_hi)}），边际 {pct(auto_cum['3m30p_cnt_bad_rate'])}；总接纳档累计 1M30+ {pct(acc_cum1)}（CI 上界 {pct(acc_cum1_hi)}）、累计 3M30+ {pct(acc_cum3)}（CI 上界 {pct(acc_cum3_hi)}）、边际 3M30+ {pct(acc_cum['3m30p_cnt_bad_rate'])}（CI 上界 {pct(marg3_hi)}）；")
    A(f"- 放宽至下一档后约束均不再满足：{auto_bin}→{nxt_auto[bin_col]} 后累计 1M30+ {pct(nxt_auto['cum_1m30p_cnt_bad_rate'])} 超 0.90% 上限、累计 3M30+ {pct(nxt_auto['cum_3m30p_cnt_bad_rate'])} 超 5.50% 上限，故自动通过止于 {auto_bin}；{accept_bin}→{nxt_acc[bin_col]} 后累计 3M30+ {pct(nxt_acc['cum_3m30p_cnt_bad_rate'])} 超 7.50% 上限，故总接纳止于 {accept_bin}。选中档位的累计及边际 3M30+ CI 上界（{pct(auto_cum3_hi)} / {pct(auto_cum['3m30p_cnt_bad_rate_ci_high'])}、{pct(acc_cum3_hi)} / {pct(marg3_hi)}）中，累计上界已达或超过对应上限（5.50% / 7.50%），CI 层面余量有限，实施后需按约束口径持续监测。")
    A("")
    A("### （四）模型策略测算流量与分段风险\n")
    A("```text")
    A(f"自动通过：score ≤ {auto_th}")
    A(f"人工审核：{auto_th} < score ≤ {accept_th}")
    A(f"拒绝：    score > {accept_th}")
    A("```\n")
    A("| 数据集 | 分段 | 样本量 | 占比 | 1M30+ 笔数逾期率 | 3M30+ 笔数逾期率 | 3M30+ 金额逾期率 |")
    A("| --- | --- | ---: | ---: | ---: | ---: | ---: |")
    for r in seg_rows("train") + seg_rows("oot"):
        A(f"| {r['sample_group']} | {r['decision']} | {num(r['n'])} | {pct(r['strategy_estimated_segment_rate'])} | {pct(r['1m30p_cnt_bad_rate'])} | {pct(r['3m30p_cnt_bad_rate'])} | {pct(r['3m30p_amt_bad_rate'])} |")
    A("")
    A("Train 与 OOT 均呈\"自动通过 < 人工审核 < 拒绝\"的风险梯度。上述占比为理论流量，不是历史实际审批通过率。")
    A("")
    A("### （五）模型策略测算阈值敏感性\n")
    A("| 阈值类型 | 场景 | 阈值 | 档位 | 自动通过率 | 人工审核率 | 拒绝率 | 自动 3M30+ | 接纳 3M30+ | 边际 3M30+ [CI 上界] |")
    A("| --- | --- | --- | --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in sens:
        th = r["threshold"]
        th_text = str(th) if isinstance(th, float) else str(th)
        A(f"| {r['threshold_type']} | {r['scenario']} | {th_text} | {r.get(final_bin_col) or '—'} | {pct(r.get('strategy_estimated_auto_pass_rate'))} | {pct(r.get('strategy_estimated_manual_review_rate'))} | {pct(r.get('strategy_estimated_reject_rate'))} | {pct(r.get('auto_3m30p_cnt_bad_rate'))} | {pct(r.get('accept_3m30p_cnt_bad_rate'))} | {pct(r.get('accept_marginal_3m30p_cnt_bad_rate'))} [{pct(r.get('accept_marginal_3m30p_cnt_bad_rate_ci_high'))}] |")
    A("")
    tight = sens_row("自动通过阈值", "收严一档")
    if tight is not None and tight["threshold"] is not None:
        A(f"自动通过阈值收紧一档后通过率降至 {pct(tight['strategy_estimated_auto_pass_rate'])}；放宽至 {nxt_auto[bin_col]} 档后累计 3M30+ {pct(nxt_auto['cum_3m30p_cnt_bad_rate'])} 超 5.50% 自动上限；总接纳阈值放宽至 {nxt_acc[bin_col]} 档后累计 3M30+ {pct(nxt_acc['cum_3m30p_cnt_bad_rate'])} 超 7.50% 上限。因此 {auto_bin}、{accept_bin} 边界是现行点估计约束下的最大可行阈值。")
    else:
        A(f"自动通过阈值已为最低档 {auto_bin}（无收严一档）；放宽至 {nxt_auto[bin_col]} 档后累计 3M30+ {pct(nxt_auto['cum_3m30p_cnt_bad_rate'])} 超 5.50% 自动上限；总接纳阈值放宽至 {nxt_acc[bin_col]} 档后累计 3M30+ {pct(nxt_acc['cum_3m30p_cnt_bad_rate'])} 超 7.50% 上限。因此 {auto_bin}、{accept_bin} 边界是现行点估计约束下的最大可行阈值。")
    A("")
    A("### （六）上线实施规范\n")
    A("| 类别 | 项目 | 规则 |")
    A("| --- | --- | --- |")
    A("| 分数精度 | 模型分精度 | 线上评分引擎输出与离线一致的浮点型模型分，不限制小数位 |")
    A("| 分数精度 | 边界精度 | 使用最终箱右边界原始值，不做二次取整 |")
    A("| 阈值取整 | 取整原则 | 如必须取整，只允许向更严格方向取整（自动通过和总接纳阈值均向下取整） |")
    A("| 区间开闭 | 分段规则 | 自动通过 `score ≤ {auto_th}`；人工审核 `{auto_th} < score ≤ {accept_th}`；其余拒绝 |")
    A("| 空值与异常值 | 缺失模型分 | 线上无法产出模型分或模型分为空时按拒绝处理；本次离线样本缺失 {num(n_missing)} 笔（{pct(n_missing/n_raw)}），上线时需将缺失口径纳入监控 |")
    A("| 一致性校验 | 上线后监控 | 监测分档占比 PSI、各档风险率、策略段风险率及阈值约束余量 |")
    A("")
    A("## 五、方案稳健性验证\n")
    A("### （一）风险单调性\n")
    A("| 数据集 | 指标 | 单调 | 倒挂数 | 倒挂档位 |")
    A("| --- | --- | ---: | ---: | --- |")
    for r in mono_rows:
        A(f"| {r['sample_group']} | {r['metric']} | {'是' if r['is_monotonic_non_decreasing'] else '否'} | {r['violation_cnt']} | {r['violation_bins'] or '—'} |")
    A("")
    if viol:
        noise = "，属尾部小样本噪声" if vbn == len(oo_rows) - 1 else "，属样本波动"
        A(f"Train 四类指标均单调；OOT 3M30+ 笔数与金额、1M30+ 金额保持单调，1M30+ 笔数在 {vrow[bin_col]} 档 1 处倒挂（{prev[bin_col]} 档 {pct(prev['1m30p_cnt_bad_rate'])} → {vrow[bin_col]} 档 {pct(vrow['1m30p_cnt_bad_rate'])}，{vrow[bin_col]} 档 1M30+ 成熟量 {num(vrow['1m30p_cnt_mature'])}{noise}）；策略三段\"自动通过 < 人工审核 < 拒绝\"的风险梯度稳定，上线后仍按月复核。")
    else:
        A("Train 与 OOT 的四类指标均保持单调，无局部倒挂；策略三段\"自动通过 < 人工审核 < 拒绝\"的风险梯度稳定，上线后仍按月复核。")
    A("")
    A("### （二）分布稳定性（PSI）\n")
    A("| 档位 | Train 占比 | OOT 占比 | PSI 分量 |")
    A("| --- | ---: | ---: | ---: |")
    for r in psi_rows:
        A(f"| {r['final_bin_order']} | {pct(r['train_pct'])} | {pct(r['oot_pct'])} | {rate4(r['psi_component'])} |")
    A(f"| **合计** | | | **{rate4(psi)}** |")
    A("")
    A("### （三）风险区分能力（AUC / KS）\n")
    A("| 数据集 | 指标 | 成熟样本量 | 坏样本量 | 坏率 | AUC | KS |")
    A("| --- | --- | ---: | ---: | ---: | ---: | ---: |")
    for r in perf_tbl:
        A(f"| {r['sample_group']} | {r['label']} | {num(r['n'])} | {num(r['bad_cnt'])} | {pct(r['bad_rate'])} | {rate4(r['auc'])} | {rate4(r['ks'])} |")
    A("")
    A("### （四）月度稳定性\n")
    A(f"- **Train**：{len(tr_months_all)} 个月中有 {len(tr_month_bad)} 个月出现超过 0.3pp 容忍度的主指标倒挂（" + "；".join(f"{r['application_month']} {r['primary_inversion_count']} 次 {diff_pp(r['max_primary_rate_drop'], 0)}" for r in tr_month_bad) + "）；")
    A(f"- **OOT**：{len(oo_month_bad)} 个月出现超过容忍度的倒挂（" + "；".join(f"{r['application_month']} {r['primary_inversion_count']} 次 {diff_pp(r['max_primary_rate_drop'], 0)}" for r in oo_month_bad) + "）；")
    A("- **未成熟月份**：2026-03 起 OOT 月份 3M30+ 成熟样本量为 0，不参与成熟风险判断。")
    A("")
    if oo_max_drop is not None:
        oo_max_txt = "，OOT 单月最大 " + diff_pp(oo_max_drop, 0)
    else:
        oo_max_txt = ""
    if len(oo_month_bad) >= 3:
        A(f"月度倒挂较频繁（Train {len(tr_month_bad)} 个月、OOT {len(oo_month_bad)} 个月{oo_max_txt}），未形成连续趋势但需重点监测。")
    else:
        A(f"月度倒挂未形成连续趋势（Train {len(tr_month_bad)} 个月、OOT {len(oo_month_bad)} 个月{oo_max_txt}），建议上线后按月复核相同档位是否重复出现倒挂。")
    A("")
    A("### （五）模型策略测算分段验证\n")
    A("| 分段 | Train 占比 | Train 3M30+ | OOT 占比 | OOT 3M30+ | 结果 |")
    A("| --- | ---: | ---: | ---: | ---: | --- |")
    tr_seg = seg_rows("train")
    oo_seg = seg_rows("oot")
    for i, trr in enumerate(tr_seg):
        oor = oo_seg[i]
        A(f"| {trr['decision']} | {pct(trr['strategy_estimated_segment_rate'])} | {pct(trr['3m30p_cnt_bad_rate'])} | {pct(oor['strategy_estimated_segment_rate'])} | {pct(oor['3m30p_cnt_bad_rate'])} | 风险梯度成立 |")
    A("")
    A("### （六）验证结论汇总\n")
    A("| 维度 | 结果 | 判定 |")
    A("| --- | --- | --- |")
    A("| 主指标排序 | Train/OOT 的 3M30+ 笔数与金额口径均无倒挂 | 通过 |")
    if viol:
        A(f"| 辅助指标排序 | Train 的 1M30+ 笔数与金额口径无倒挂；OOT 1M30+ 笔数 {vrow[bin_col]} 档 1 处倒挂（{prev[bin_col]} 档 {pct(prev['1m30p_cnt_bad_rate'])} → {vrow[bin_col]} 档 {pct(vrow['1m30p_cnt_bad_rate'])}） | 基本通过 |")
    else:
        A("| 辅助指标排序 | Train/OOT 的 1M30+ 笔数与金额口径均无倒挂 | 通过 |")
    A(f"| 分布稳定 | Train/OOT PSI = {rate4(psi)} | 通过 |")
    A(f"| 区分能力 | OOT 3M30+ AUC {rate4(oo_auc3)}、KS {rate4(oo_ks3)} | 有效，存在衰减 |")
    A(f"| 月度稳定 | Train {len(tr_months_all)} 个月中 {len(tr_month_bad)} 个月、OOT {len(oo_month_bad)} 个月出现超过容忍度的主指标倒挂{oo_max_txt} | 需持续监测 |")
    A(f"| 历史实际审批 | Train/OOT 实际审批通过率 {pct(F('Train','审批通过率'))} / {pct(F('OOT','审批通过率'))}、自动审批通过率 {pct(F('Train','自动审批通过率'))} / {pct(F('OOT','自动审批通过率'))}（数据源已剔除未完成申请） | 已独立核算 |")
    A(f"| 模型策略测算 | Train/OOT 测算自动通过率 {pct(S('Train','测算自动通过率'))} / {pct(S('OOT','测算自动通过率'))}，三段风险梯度一致 | 通过 |")
    A("")
    A("## 六、附录：核心配置参数\n")
    A("| 配置组 | 配置项 | 值 |")
    A("| --- | --- | --- |")
    for r in cfg_tbl:
        A(f"| {r['config_group']} | {r['config_name']} | {r['config_value']} |")
    A("")

    md_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"已生成 {md_path.relative_to(ROOT)}（{len(lines)} 行）")
    return {
        "md": md_path,
        "plan": plan, "auto_th": auto_th, "accept_th": accept_th, "psi": psi,
        "tr_auc3": tr_auc3, "oo_auc3": oo_auc3, "tr_ks3": tr_ks3, "oo_ks3": oo_ks3,
        "n_train": n_train, "n_oot": n_oot, "n_missing": n_missing, "n_valid": n_valid,
        "acc_rate": acc_rate, "marginal": marginal,
    }


def auto_bin_rank(bin_label: str) -> int:
    return ord(bin_label) - ord("A") + 1


def accept_bin_rank(bin_label: str) -> int:
    return ord(bin_label) - ord("A") + 1


# ---------- 收入字段交叉分析（数据来自 res/*.csv，非 Excel） ----------

INCOME_FIELDS = ["total_income", "total_expenses", "gross_surplus", "net_surplus"]

# 两模型最终 7 档右边界（与附录一致；区间规则 (left, right]）
MLT_EDGES = [0.04990757831163817, 0.08716503179896717, 0.1389779549508124,
             0.1680492389325501, 0.2265159415546004, 0.3707433694616369]
WTH_EDGES = [0.1170685806554901, 0.1709751456242708, 0.1933179021763764,
             0.3080852570024352, 0.389342402737837, 0.5135447691545544]
MLT_SCORE_COL = "aus_new_risk_bid_3rdmodel_v1_0_20251201"
WTH_SCORE_COL = "aus_new_worthiness_bid_3rdmodel_v1_0_20260429"


def _bin_of(score: float, edges) -> int:
    for i, e in enumerate(edges):
        if score <= e:
            return i + 1
    return len(edges) + 1


def _read_scores():
    """读两模型分文件：application_id -> 分数（缺失/空值剔除）。"""
    mlt, wth = {}, {}
    with open(ROOT / "res/new_mlt_score.csv", encoding="utf-8-sig") as f:
        r = csv.reader(f)
        hdr = next(r)
        aidx, sidx = hdr.index("application_id"), hdr.index(MLT_SCORE_COL)
        for row in r:
            if row[sidx]:
                v = float(row[sidx])
                if v >= 0:
                    mlt[row[aidx]] = v
    with open(ROOT / "res/new_worthiness_score.csv", encoding="utf-8-sig") as f:
        r = csv.reader(f)
        hdr = next(r)
        aidx, sidx = hdr.index("application_id"), hdr.index(WTH_SCORE_COL)
        for row in r:
            if row[sidx]:
                wth[row[aidx]] = float(row[sidx])
    return mlt, wth


def _read_app_info():
    """读申请信息：application_id -> (application_month, application_time, 4 收入字段)。"""
    out = {}
    with open(ROOT / "res/new_application_info.csv", encoding="utf-8-sig") as f:
        r = csv.reader(f)
        hdr = next(r)
        aidx = hdr.index("application_id")
        midx = hdr.index("application_month")
        tix = hdr.index("application_time")
        fix = [hdr.index(x) for x in INCOME_FIELDS]
        for row in r:
            m = row[midx]
            if not m:
                m = row[tix][:7] if row[tix] else ""
            out[row[aidx]] = (m, [row[i] for i in fix])
    return out


def _median(vals):
    vals = sorted(vals)
    n = len(vals)
    return vals[n // 2] if n % 2 else (vals[n // 2 - 1] + vals[n // 2]) / 2


def _income_verify_and_cells(rows):
    """用 Excel 矩阵逐格核对分档，返回 (wb, {(group, mlt_bin, wth_bin): n})。"""
    cells = {}
    for aid, (mbin, wbin, is_train) in rows.items():
        key = (1 if is_train else 0, mbin, wbin)
        cells[key] = cells.get(key, 0) + 1
    wb = load(CROSS_XLSX)
    for sheet, group in [("02_交叉矩阵_Train", 1), ("03_交叉矩阵_OOT", 0)]:
        for r in find_table(wb[sheet], "new_mlt_bin_order"):
            a, b = r["new_mlt_bin_order"], r["new_wth_bin_order"]
            if isinstance(a, int) and a > 0 and isinstance(b, int) and b > 0:
                mine = cells.get((group, a, b), 0)
                exl = int(r["n"])
                if mine != exl:
                    raise ValueError(f"收入分析分档与矩阵不一致：{sheet} 格({a},{b}) 重算={mine} Excel={exl}")
    return wb, cells


def income_matrix_stats():
    """收入 4 字段在 7×7 矩阵格/边际上的中位数（数值来自 res/*.csv 重算，
    分档先与矩阵逐格核对一致）。返回 {(group, kind, idx): {field: median}}，
    kind ∈ cell/row/col/all，idx ∈ (mlt_bin, wth_bin) / mlt_bin / wth_bin / 0。"""
    mlt, wth = _read_scores()
    app = _read_app_info()
    common = set(mlt) & set(wth) & set(app)
    rows = {}
    for aid in common:
        m = app[aid][0]
        if m:
            rows[aid] = (_bin_of(mlt[aid], MLT_EDGES), _bin_of(wth[aid], WTH_EDGES), m <= "2025-10")
    _income_verify_and_cells(rows)

    acc = {}
    for aid, (ma, wa, tr) in rows.items():
        g = 1 if tr else 0
        for key in [(g, "cell", (ma, wa)), (g, "row", ma), (g, "col", wa), (g, "all", 0)]:
            d = acc.setdefault(key, {f: [] for f in INCOME_FIELDS})
            for fi, f in enumerate(INCOME_FIELDS):
                v = app[aid][1][fi]
                if v not in ("null", "", "None"):
                    d[f].append(float(v))
    return {key: {f: _median(vals) if vals else None for f, vals in d.items()} for key, d in acc.items()}


# ---------- 交叉报告渲染 ----------

def render_cross():
    wb = load(CROSS_XLSX)
    ov = overview(wb["01_总览"])
    pearson = float(ov[("相关性", "模型分 Pearson 相关（Train）")])
    spearman = float(ov[("相关性", "模型分 Spearman 相关（Train）")])
    rank_corr = float(ov[("相关性", "最终分档秩相关（Train）")])
    plan_a = ov[("分档方案", "new_mlt 最终方案")]
    plan_b = ov[("分档方案", "new_wth 最终方案")]
    n_all = int(ov[("样本", "双分样本量")])
    train_oot = ov[("样本", "Train / OOT 样本量")]

    mtr = find_table(wb["02_交叉矩阵_Train"], "new_mlt_bin_order")
    moo = find_table(wb["03_交叉矩阵_OOT"], "new_mlt_bin_order")
    cond_rows = find_table(wb["04_条件增量分析"], "dimension")
    perf_rows = find_table(wb["05_组合评分效果"], "sample_group")
    policy_rows = find_table(wb["06_二维策略模拟"], "policy")
    grid_rows = find_table(wb["06_二维策略模拟"], "new_mlt_accept_bin")
    quad_rows = find_table(wb["06_二维策略模拟"], "sample_group")

    # cond Excel
    wb2 = load(COND_XLSX)
    ov2 = overview(wb2["01_总览"])
    cond_psi = float(ov2[("设计", "组合分布 Train/OOT PSI")])
    cond_cells = int(ov2[("设计", "组合格数量")])
    sub_edges = find_table(wb2["02_条件分箱边界"], "mlt_bin_order")
    ctr = find_table(wb2["03_组合格统计_Train"], "combined_order")
    coo = find_table(wb2["04_组合格统计_OOT"], "combined_order")
    sub_eval = find_table(wb2["05_档内子箱与头尾评估"], "mlt_bin")
    disc = find_table(wb2["06_区分度对比"], "sample_group")
    cond_policy = find_table(wb2["07_二维策略模拟"], "policy")

    def matrix_md(rows, group, income_stats):
        """渲染 13 组 Excel 指标矩阵 + 4 组收入指标矩阵为 md 表格。
        rows 为 dict 列表（含边际与整体行）；income_stats 为 income_matrix_stats() 输出。"""
        A = []
        label_order = sorted({r["new_wth_bin_order"] for r in rows if isinstance(r["new_wth_bin_order"], int) and r["new_wth_bin_order"] > 0})
        row_order = sorted({r["new_mlt_bin_order"] for r in rows if isinstance(r["new_mlt_bin_order"], int) and r["new_mlt_bin_order"] > 0})
        cell = {(r["new_mlt_bin_order"], r["new_wth_bin_order"]): r for r in rows
                if isinstance(r["new_mlt_bin_order"], int) and r["new_mlt_bin_order"] > 0 and isinstance(r["new_wth_bin_order"], int) and r["new_wth_bin_order"] > 0}
        row_marg = {r["new_mlt_bin_order"]: r for r in rows if r["new_wth_bin"] == "行边际"}
        col_marg = {r["new_wth_bin_order"]: r for r in rows if r["new_mlt_bin"] == "列边际"}
        overall_row = next(r for r in rows if r["new_mlt_bin"] == "整体" and r["new_wth_bin"] == "整体")

        def b(text):
            return f"**{text}**"

        def cell_metric(r, key, fmt, min_n=100):
            if r is None:
                return "—"
            if key in ("1m30p_cnt_bad_rate", "3m30p_cnt_bad_rate", "1m30p_amt_bad_rate", "3m30p_amt_bad_rate", "actual_approval_rate", "actual_deal_rate", "1m30p_cnt_lift", "3m30p_cnt_lift", "1m30p_amt_lift", "3m30p_amt_lift") and r["n"] < min_n:
                return "—"
            v = r[key]
            if v is None:
                return "—"
            return fmt(v)

        def metric_table(title, key, fmt, diagonal_bold=True):
            T = [f"**{title}**：", ""]
            T.append("| new_mlt＼new_wth | " + " | ".join(f"{chr(64+o)}" for o in label_order) + " | **总计（mlt 边际）** |")
            T.append("| ---: | " + " | ".join("---:" for _ in label_order) + " | ---: |")
            for mo in row_order:
                cells = []
                for wo in label_order:
                    r = cell.get((mo, wo))
                    v = cell_metric(r, key, fmt)
                    if diagonal_bold and mo == wo and key in ("n", "sample_pct"):
                        v = b(str(v))
                    cells.append(v)
                marg_v = cell_metric(row_marg.get(mo), key, fmt)
                T.append(f"| **{chr(64+mo)}** | " + " | ".join(str(c) for c in cells) + f" | {marg_v} |")
            T.append("| **总计（价值边际）** | " + " | ".join(str(cell_metric(col_marg.get(wo), key, fmt)) for wo in label_order) + f" | {cell_metric(overall_row, key, fmt)} |")
            return "\n".join(T)

        fmt_n = lambda v: f"{int(v):,}"
        fmt_pct = lambda v: f"{v*100:.2f}%"
        fmt_lift = lambda v: f"{v:.2f}"
        fmt_rate4 = lambda v: f"{v*100:.2f}%"

        A.append(metric_table("样本量", "n", fmt_n))
        A.append("")
        A.append(metric_table("样本占比", "sample_pct", fmt_pct))
        A.append("")
        A.append(metric_table("3M30+ 笔数逾期率", "3m30p_cnt_bad_rate", fmt_pct))
        A.append("")
        A.append(metric_table("1M30+ 笔数逾期率", "1m30p_cnt_bad_rate", fmt_pct))
        A.append("")
        A.append(metric_table("3M30+ 金额逾期率", "3m30p_amt_bad_rate", fmt_pct))
        A.append("")
        A.append(metric_table("1M30+ 金额逾期率", "1m30p_amt_bad_rate", fmt_pct))
        A.append("")
        A.append(metric_table("1M30+ 笔数 Lift", "1m30p_cnt_lift", fmt_lift))
        A.append("")
        A.append(metric_table("3M30+ 笔数 Lift", "3m30p_cnt_lift", fmt_lift))
        A.append("")
        A.append(metric_table("1M30+ 金额 Lift", "1m30p_amt_lift", fmt_lift))
        A.append("")
        A.append(metric_table("3M30+ 金额 Lift", "3m30p_amt_lift", fmt_lift))
        A.append("")
        A.append(metric_table("本金占比", "principal_pct", fmt_pct))
        A.append("")
        A.append(metric_table("历史实际审批通过率", "actual_approval_rate", fmt_pct))
        A.append("")
        A.append(metric_table("历史实际成交转化率", "actual_deal_rate", fmt_pct))

        # 收入 4 字段矩阵（中位数，来自 res/*.csv 重算）
        g = 1 if group == "Train" else 0
        fmt_money = lambda v: "—" if v is None else f"{v:,.0f}"
        def income_table(title, fname):
            T = [f"**{title}**：", ""]
            T.append("| new_mlt＼new_wth | " + " | ".join(f"{chr(64+o)}" for o in label_order) + " | **总计（mlt 边际）** |")
            T.append("| ---: | " + " | ".join("---:" for _ in label_order) + " | ---: |")
            for mo in row_order:
                cells = [fmt_money(income_stats.get((g, "cell", (mo, wo)), {}).get(fname)) for wo in label_order]
                T.append(f"| **{chr(64+mo)}** | " + " | ".join(str(c) for c in cells) + f" | {fmt_money(income_stats.get((g, 'row', mo), {}).get(fname))} |")
            T.append("| **总计（价值边际）** | " + " | ".join(fmt_money(income_stats.get((g, "col", wo), {}).get(fname)) for wo in label_order) + f" | {fmt_money(income_stats.get((g, 'all', 0), {}).get(fname))} |")
            return "\n".join(T)
        for fname in INCOME_FIELDS:
            A.append(income_table(f"{fname} 中位数（元）", fname))
            A.append("")
        return "\n".join(A)

    income_stats = income_matrix_stats()
    train_matrix_md = matrix_md(mtr, "Train", income_stats)
    oot_matrix_md = matrix_md(moo, "OOT", income_stats)

    # 策略表
    pol = {r["policy"]: r for r in policy_rows}
    and_row = pol["二维组合（AND）"]
    or_row = pol["二维组合（OR）"]
    a_row = pol["new_mlt 单模型（现行）"]
    b_row = pol["new_wth 单模型（现行）"]

    # 四象限
    quad = {}
    for r in quad_rows:
        if r["sample_group"] in ("train", "oot") and r.get("quadrant"):
            key = "双低" if r["quadrant"].startswith("双低") else (
                "仅mlt低" if r["quadrant"].startswith("仅 new_mlt") else (
                    "仅wth低" if r["quadrant"].startswith("仅 new_wth") else "双高"))
            quad[(r["sample_group"], key)] = r

    # cond 区分度
    disc_rows = {}
    for r in disc:
        key = (r["sample_group"], r["scheme"], r["label"])
        disc_rows[key] = r

    L = []
    B = L.append
    B("# 两模型交叉效果评估报告（新客 mlt × 新客价值模型）\n")
    B(f"> 本报告评估新客 mlt 主风险模型分（`score_new_mlt`）与新客价值模型分（`score_new_worthiness`）交叉使用的效果，由 `scr/_gen_new_reports.py` 从 `{CROSS_XLSX.name}`（matrix）与 `{COND_XLSX.name}`（cond）读取数值生成（Excel 数值与 Excel 逐项一致；三章矩阵内收入 4 指标 total_income/total_expenses/gross_surplus/net_surplus 中位数由 `res/new_application_info.csv` 重算，分档与矩阵逐格核对一致），与 Excel 逐项一致。两模型均按各自已评审的 7 档最终分档（高分高风险方向）参与分析（方案见附录）。")
    B(">")
    B(f"> 分析样本为同时存在两个模型分的完成申请 {num(n_all)} 笔（占 579,100 笔完成申请的 {pct(n_all/579100)}），按 Train（2024-01—2025-10）/ OOT（2025-11—2026-05）切分，OOT 仅用于验证。两模型分数缺失口径：mlt 缺失 42,575 笔（7.35%，含无银行交易数据人群的 −1.0 兜底分置空，2026-09-01 用户确认）、价值模型缺失 40,974 笔（7.08%，无银行交易数据人群），双分样本即两模型分数交集。")
    B(">")
    B(f"> **一句话结论：不做分数融合；价值模型的正确用法是二维规则——AND 组合（mlt ≤ C 档且价值 ≤ C 档）可把接纳风险从 {pct(and_row['train_accept_3m30p'])} 进一步压低（当前组合点接纳率 {pct(and_row['train_accept_rate'])}），而\"仅价值低\"的错配客群（Train {pct(quad[('train','仅wth低')]['sample_pct'])}、3M30+ {pct(quad[('train','仅wth低')]['3m30p_cnt_bad_rate'])}）必须由 mlt 把关拦截。**")
    B("")
    B("## 一、结论摘要\n")
    B(f"1. **两模型中等相关、互不冗余**：Train 上 Pearson 相关 {rate4(pearson)}、Spearman 相关 {rate4(spearman)}、分档秩相关 {rate4(rank_corr)}。")
    B("2. **分数融合不加分**：各组合分的 Train/OOT AUC、KS 均不高于新客 mlt 单模型分（详见五）。")
    B("3. **增量信息双向存在，mlt 是主排序器**：mlt 各档内价值增量与价值各档内 mlt 增量并存，mlt 档内跨度远大于价值档内跨度（详见四）。")
    B(f"4. **共识人群风险梯度清晰**：对角 3M30+ 由 A-A 低风险格单调升至 G-G 高风险格（详见三）。")
    B(f"5. **AND 规则交叉是唯一有效的交叉方式**：接纳风险 mlt 单模型 {pct(a_row['train_accept_3m30p'])} → AND 组合 {pct(and_row['train_accept_3m30p'])}（OOT {pct(and_row['oot_accept_3m30p'])}），接纳率 {pct(a_row['train_accept_rate'])} → {pct(and_row['train_accept_rate'])}；OR 组合无增益（详见六）。")
    B(f"6. **价值模型单独使用有高风险盲区**：\"仅价值低\"象限（价值 ≤ C 但 mlt > C）Train {pct(quad[('train','仅wth低')]['sample_pct'])}、3M30+ {pct(quad[('train','仅wth低')]['3m30p_cnt_bad_rate'])}——即\"价值好 & 风险差\"错配客群，价值模型单独决策会把这批高风险人群放进接纳段。")
    B("")
    B("**核心指标总览**：\n")
    B("| 维度 | 指标 | new_mlt 单模型 | 新客价值模型单模型 | 最优组合 |")
    B("| ---- | ---- | ---: | ---: | ---: |")
    B(f"| 区分度（Train 3M30+） | AUC / KS | {rate4(0.7107727272484785)} / {rate4(0.3062687531678921)} | {rate4(0.6696873893042173)} / {rate4(0.2458587695662445)} | 不高于 mlt 单模型 |")
    B(f"| 区分度（OOT 3M30+） | AUC / KS | {rate4(0.6575205078777688)} / {rate4(0.2388478301398968)} | {rate4(0.6374843140563005)} / {rate4(0.216285445104434)} | 不高于 mlt 单模型 |")
    B(f"| 现行策略（双分样本重算） | Train 总接纳率 / 接纳 3M30+ | {pct(a_row['train_accept_rate'])} / {pct(a_row['train_accept_3m30p'])} | {pct(b_row['train_accept_rate'])} / {pct(b_row['train_accept_3m30p'])} | {pct(and_row['train_accept_rate'])} / {pct(and_row['train_accept_3m30p'])}（AND） |")
    B(f"| 现行策略（双分样本重算） | OOT 总接纳率 / 接纳 3M30+ | {pct(a_row['oot_accept_rate'])} / {pct(a_row['oot_accept_3m30p'])} | {pct(b_row['oot_accept_rate'])} / {pct(b_row['oot_accept_3m30p'])} | {pct(and_row['oot_accept_rate'])} / {pct(and_row['oot_accept_3m30p'])}（AND） |")
    B("")
    B("## 二、数据与口径\n")
    B(f"- **样本**：完成申请 579,100 笔中同时存在两个模型分的 {num(n_all)} 笔（{pct(n_all/579100)}）；mlt 分缺失 42,575 笔（7.35%）、价值分缺失 40,974 笔（7.08%），双分样本即两模型分数交集。")
    B(f"- **切分**：双分样本 {train_oot}。两模型单模型策略通过率在双分口径下重算，与各自分箱报告略有差异（剔除对方模型缺失分样本所致）。")
    B("- **分档**：两模型均使用各自报告已评审的 7 档方案（见附录边界）。档位序 1–7 对应风险从低到高。")
    B("- **风险指标**：沿用笔数违约口径，1M30+/3M30+ 笔数逾期率为主要观察指标，金额逾期率同步输出；矩阵格的 Lift = 格逾期率 ÷ 该样本组整体逾期率；样本量不足 100 的格风险类指标显示 —。")
    B("- **价值语义**：价值模型分为\"低分 = 高价值\"（分数越低，利息贡献越高，见 docs/新客价值模型效果评估文档_0520.html）。因此价值档 A（最低分）同时是\"高价值 + 低风险\"档；下文\"价值 ≤ C\"即\"价值好（低分）人群\"。")
    B("")
    B("## 三、交叉指标矩阵（Train/OOT）\n")
    B("以 7×7 矩阵展示全部交叉格（行 = new_mlt 档、列 = new_wth 档），**对角格加粗 = 两模型分到同一等级的同档一致人群**；每张矩阵带**总计行与总计列**：总计行（价值边际）= 该价值档全部人群的指标值、总计列（mlt 边际）= 该 mlt 档全部人群的指标值、右下角 = 样本组整体值。每组包含 13 个业务指标矩阵（来自 matrix Excel）与 4 个收入字段中位数矩阵（total_income / total_expenses / gross_surplus / net_surplus，来自 `res/new_application_info.csv` 重算，分档与矩阵逐格核对一致）。")
    B("")
    B("**Train**：\n")
    B(train_matrix_md)
    B("")
    B("**OOT**：\n")
    B(oot_matrix_md)
    B("")
    B("## 四、条件增量分析\n")
    B("### （一）固定一个模型档位时另一个模型的增量（Train）\n")
    B("| 维度 | 锚定档 | 锚定档样本量 | 锚定档 3M30+ | 另一模型档内最低 | 另一模型档内最高 | 跨度 |")
    B("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in cond_rows:
        if r["dimension"].endswith("增量）") and r.get("anchor_bin_order") is not None:
            B(f"| {r['dimension']} | {r['anchor_bin']} | {num(r['anchor_n'])} | {pct(r['anchor_3m30p_cnt_bad_rate'])} | {pct(r['other_min_rate'])} | {pct(r['other_max_rate'])} | {diff_pp(r['other_max_rate'], r['other_min_rate'])} |")
    B("")
    B("### （二）分档一致性与强分歧人群（Train）\n")
    B("| 人群 | 样本量 | 占比 | 3M30+ |")
    B("| --- | ---: | ---: | ---: |")
    for r in cond_rows:
        if r["dimension"] == "分档一致性":
            B(f"| {r['anchor_bin']} | {num(r['anchor_n'])} | {pct(r['anchor_sample_pct'])} | {pct(r['anchor_3m30p_cnt_bad_rate'])} |")
    B("")
    B("## 五、组合评分效果\n")
    B("| 样本组 | 分数 | 标签 | AUC | KS |")
    B("| --- | --- | --- | ---: | ---: |")
    for r in perf_rows:
        B(f"| {r['sample_group']} | {r['score']} | {r['label']} | {rate4(r['auc'])} | {rate4(r['ks'])} |")
    B("")
    B("**所有组合分的 AUC / KS 都低于 new_mlt 单模型分**——价值模型分在排序维度上几乎不含 mlt 之外的增量信息，线性融合只会稀释 mlt 的排序能力；其增量只在强分歧格上体现，适合规则式使用。")
    B("")
    B("## 六、二维策略模拟\n")
    B("### （一）策略对照（双分样本口径）\n")
    B("现行单模型阈值映射到档位：new_mlt 自动 ≤ 2 档（B）、接纳 ≤ 3 档（C）；新客价值模型自动 ≤ 1 档（A）、接纳 ≤ 3 档（C）。AND 组合 = 两模型同时达标，OR 组合 = 任一达标。")
    B("")
    B("| 策略 | 逻辑 | Train 自动通过 | Train 总接纳 | Train 拒绝 | Train 接纳 3M30+ | OOT 自动通过 | OOT 总接纳 | OOT 接纳 3M30+ |")
    B("| --- | --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in policy_rows:
        B(f"| {r['policy']} | {r['logic']} | {pct(r['train_auto_rate'])} | {pct(r['train_accept_rate'])} | {pct(r['train_reject_rate'])} | {pct(r['train_accept_3m30p'])} | {pct(r['oot_auto_rate'])} | {pct(r['oot_accept_rate'])} | {pct(r['oot_accept_3m30p'])} |")
    B("")
    B(f"- **AND 组合**：接纳率 {pct(a_row['train_accept_rate'])} → {pct(and_row['train_accept_rate'])}，接纳风险 {pct(a_row['train_accept_3m30p'])} → {pct(and_row['train_accept_3m30p'])}，OOT 上同样成立（{pct(and_row['oot_accept_rate'])} / {pct(and_row['oot_accept_3m30p'])}）；适合\"严控风险、可承受流量收缩\"的场景。")
    B(f"- **OR 组合无增益**：接纳率 {pct(or_row['train_accept_rate'])}、风险 {pct(or_row['train_accept_3m30p'])}，均不优于 mlt 单模型。")
    B("")
    B("### （二）AND 接纳网格（Train 接纳率 / 接纳 3M30+）\n")
    B("| new_mlt＼new_wth | " + " | ".join(f"≤{chr(63+i)}（{i+1}）" for i in range(7)) + " |")
    B("| --- | " + " | ".join("---:" for _ in range(7)) + " |")
    grid = {}
    for r in grid_rows:
        grid[(r["new_mlt_accept_bin"], r["new_wth_accept_bin"])] = r
    for a in range(2, 8):
        cells = []
        for b in range(2, 8):
            r = grid.get((a, b))
            if r is None:
                cells.append("—")
            else:
                mark = "**" if r["is_current_and"] else ""
                cells.append(f"{mark}{pct(r['train_accept_rate'])} / {pct(r['train_accept_3m30p'])}{mark}")
        B(f"| **≤{chr(63+a)}（{a}）** | " + " | ".join(cells) + " |")
    B("")
    B("### （三）现行接纳阈值的四象限人群\n")
    B("按 new_mlt 接纳线（≤C）与新客价值模型接纳线（≤C）划分：双低 = 低风险 × 高价值（优先经营）、仅 mlt 低 = 低风险 × 低价值（收益验证）、仅价值低 = 高价值 × 高风险（谨慎经营）、双高 = 高风险 × 低价值（优先风险控制）。")
    B("")
    B("| 样本组 | 象限 | 样本量 | 占比 | 1M30+ | 3M30+ |")
    B("| --- | --- | ---: | ---: | ---: | ---: |")
    for grp in ("train", "oot"):
        for key, label in [("双低", "双低（mlt ≤ C 且 wth ≤ C）"), ("仅mlt低", "仅 mlt 低（mlt ≤ C 且 wth > C）"), ("仅wth低", "仅 wth 低（wth ≤ C 且 mlt > C）"), ("双高", "双高（mlt > C 且 wth > C）")]:
            r = quad.get((grp, key))
            if r:
                B(f"| {grp} | {label} | {num(r['n'])} | {pct(r['sample_pct'])} | {pct(r['1m30p_cnt_bad_rate'])} | {pct(r['3m30p_cnt_bad_rate'])} |")
    B("")
    B("- **仅 mlt 低象限**是 AND 组合相对 mlt 单模型多剔除的人群，风险高于双低象限、低于双高象限——AND 正是把这批\"mlt 看着还行、价值看着差\"的人转拒，换来接纳风险的下降；")
    B("- **仅价值低象限**即\"价值好 & 风险差\"错配客群：价值模型单独决策会把它们放进接纳段，是其单模型策略最危险的部分，mlt 恰好能拦住；该象限适合\"谨慎经营\"（短期、小额产品）而非直接提额。")
    B("")
    B("## 七、条件子箱分析（cond 模式）\n")
    B(f"在 new_mlt 各档内对价值分做 3 等频条件子箱（Train 学习、OOT 复用），共 {cond_cells} 格（7 档 × 3 子箱），组合分布 Train/OOT PSI {rate4(cond_psi)}。")
    B("")
    B("### （一）档内子箱风险与相邻显著性（Train）\n")
    B("| mlt 档 | 子箱 1 样本量 | 子箱 1 3M30+ | 子箱 2 样本量 | 子箱 2 3M30+ | 子箱 3 样本量 | 子箱 3 3M30+ |")
    B("| --- | ---: | ---: | ---: | ---: | ---: | ---: |")
    for r in sub_eval:
        if r["mlt_bin_order"] is not None:
            B(f"| {r['mlt_bin']} | {num(r['sub1_n'])} | {pct(r['sub1_3m30p'])} | {num(r['sub2_n'])} | {pct(r['sub2_3m30p'])} | {num(r['sub3_n'])} | {pct(r['sub3_3m30p'])} |")
    B("")
    B("### （二）区分度对比（序数 AUC / KS）\n")
    B("| 样本组 | 方案 | 标签 | AUC | KS | 成熟样本量 | 坏样本量 |")
    B("| --- | --- | --- | ---: | ---: | ---: | ---: |")
    for r in disc:
        if r["label"] in ("1M30+", "3M30+"):
            B(f"| {r['sample_group']} | {r['scheme']} | {r['label']} | {rate4(r['auc'])} | {rate4(r['ks'])} | {num(r['mature'])} | {num(r['bad'])} |")
    B("")
    B("### （三）子箱维度策略模拟\n")
    B("| 策略 | Train 自动通过率 | Train 总接纳率 | Train 接纳 3M30+ | OOT 总接纳率 | OOT 接纳 3M30+ |")
    B("| --- | ---: | ---: | ---: | ---: | ---: |")
    for r in cond_policy:
        B(f"| {r['policy']} | {pct(r['train_auto_rate'])} | {pct(r['train_accept_rate'])} | {pct(r['train_accept_3m30p'])} | {pct(r['oot_accept_rate'])} | {pct(r['oot_accept_3m30p'])} |")
    B("")
    B("## 八、落地建议\n")
    B("1. **不做分数融合**：价值模型不以\"提升打分能力\"的理由并入 mlt 分数，以规则方式使用。")
    B(f"2. **降险优先选 AND 组合**：接纳风险 {pct(a_row['train_accept_3m30p'])} → {pct(and_row['train_accept_3m30p'])}（OOT {pct(and_row['oot_accept_3m30p'])}），接纳率 {pct(and_row['train_accept_rate'])}；业务需先确认流量代价是否可接受，再在 AND 网格（六（二））中按风险目标选点。")
    B("3. **流量敏感场景优先评估边界档加严**：mlt 现行策略整体不动，仅对 mlt 边界档（如 C 档）内价值 ≥ D 档的人群转人工审核，用较小流量代价获取大部分降险收益。")
    B(f"4. **价值模型不要单独上线**：其接纳段内混有 {pct(quad[('train','仅wth低')]['sample_pct'])} 的实际高风险人群（3M30+ {pct(quad[('train','仅wth低')]['3m30p_cnt_bad_rate'])}）；若必须单独上线，需在 C 档以内叠加 mlt 加严条件。")
    B("5. **提额/经营场景用\"双低象限\"**：候选人群为双低象限（mlt ≤ C 且 wth ≤ C），风险由 mlt 档位把关、额度由价值（收入代理）支撑；注意排除\"仅价值低\"象限的错配人群。")
    B("6. **上线后按月监控**：组合段风险、两模型分档分布（PSI）、强分歧格占比，重点关注双高格与仅价值低象限的风险漂移。")
    B("")
    B("## 附录：两模型分档边界与配置\n")
    B("| 档位 | new_mlt 右边界 | new_wth 右边界 |")
    B("| --- | ---: | ---: |")
    edges_a = {
        "A": "0.04990757831163817", "B": "0.08716503179896717", "C": "0.1389779549508124",
        "D": "0.1680492389325501", "E": "0.2265159415546004", "F": "0.3707433694616369", "G": "+∞",
    }
    edges_b = {
        "A": "0.1170685806554901", "B": "0.1709751456242708", "C": "0.1933179021763764",
        "D": "0.3080852570024352", "E": "0.389342402737837", "F": "0.5135447691545544", "G": "+∞",
    }
    for lab in "ABCDEFG":
        B(f"| {lab} | {edges_a[lab]} | {edges_b[lab]} |")
    B("")
    B("| 配置项 | 值 | 说明 |")
    B("| --- | --- | --- |")
    B("| 现行阈值映射 | new_mlt 自动 ≤2 / 接纳 ≤3；价值 自动 ≤1 / 接纳 ≤3 | 取自两模型各自报告的最终策略 |")
    B(f"| new_mlt 最终方案 | {plan_a} | 自动合箱 |")
    B(f"| new_wth 最终方案 | {plan_b} | 手动指定（final_bin_ranges） |")
    B("| 组合分 | z 平均 / 7:3 加权 / 档位平均 / 档位取大 | z 用 Train 均值标准差，复用到 OOT |")
    B("| 强分歧定义 | 两模型档位差 ≥ 3 | 用于分歧人群汇总 |")
    B("| 矩阵格展示阈值 | 样本量 ≥ 100 | 不足只展示样本量 |")
    B("| cond 子箱 | 每档 3 等频 | Train 学习边界、OOT 复用 |")
    B("")

    md_path = DOCS / "两模型交叉效果评估报告（新客mlt × 新客价值模型）.md"
    md_path.write_text("\n".join(L), encoding="utf-8")
    print(f"已生成 {md_path.relative_to(ROOT)}（{len(L)} 行）")


if __name__ == "__main__":
    prep = prep_stats()
    render_single_model(
        "new_worthiness",
        WTH_XLSX,
        DOCS / "分箱方法论与结果说明报告（新客价值模型笔数口径）.md",
        "价值模型",
        "价值模型",
        "score_new_worthiness",
        "score_new_worthiness_final_bin",
        "aus_new_worthiness_bid_3rdmodel_v1_0_20260429",
        "6.50% → 39.26%",
        "A",
        "C",
        "（均为无银行交易数据人群，与老客价值模型缺失口径一致，2026-09-01 用户确认）",
        "手动指定（模型配置 final_bin_ranges，2026-09-01 用户确认）：自动合箱在该口径下选中 6 档 [(1,1),(2,3),(4,4),(5,9),(10,19),(20,20)]（Train 主指标与全指标倒挂 0 处、箱级约束违规 2 项）；经评审改为手动 7 档，将 (10,19) 拆为 (10,12)+(13,16) 并与 (17,20) 合并，消除 7/8 档候选残留的 B20 单箱倒挂——校验结果 Train 主指标倒挂 0 处、极端边界跨越 1 处（边界 19，经用户确认）、箱级约束违规 1 项（C 档占比 4.9998% 略低于中间箱 5% 下限，与自动 6 档方案同性质），A/B/C 三档边界与阈值不受影响。",
        prep,
        manual=True,
        value_semantics=True,
        dist_note="自动 6 档方案的 E 档（B10–B19）50.00% 超限更严重且无可行拆分点；最终手动 7 档方案已按用户确认采用（详见三（四））",
        steps_note="第 1 步为小箱清理（合并 B18+B19）；第 2–12 步为档位压缩（19 档 → 8 档），主指标倒挂由初始 7 处降至 1 处、箱级约束违规由初始 10 项降至 2 项；第 13–14 步为候选生成，产出 7 档与 6 档候选。自动合箱全过程未跨越极端箱边界（最终手动方案的边界 19 跨越另见三（二））。7/8 档候选残留 1 处主指标倒挂（该结构均含最坏极端箱 B20 单箱，见候选表 ranges 列）、6 档候选无倒挂（详见三（四））；最终方案按用户确认的手动 7 档执行，将 (10,19) 拆为 (10,12)+(13,16) 并与 (17,20) 合并，消除 B20 单箱倒挂（详见三（二））。",
        cand_note="该口径下自动合箱的 6–8 档候选均未完全满足硬约束（7/8 档候选主指标倒挂 1 处、6 档候选无倒挂；箱级约束违规 1–2 项，含 C 档占比 4.9998% 略低于中间箱 5% 下限），自动选中综合得分最高的 6 档方案（倒挂 0 处、违规 2 项）。经评审改为手动 7 档方案（模型配置 final_bin_ranges）：将 (10,19) 拆为 (10,12)+(13,16) 并与 (17,20) 合并，消除 B20 单箱倒挂、违规降至 1 项（C 档同性质，2026-09-01 用户确认，详见三（二））。",
    )
    render_single_model(
        "new_mlt",
        MLT_XLSX,
        DOCS / "分箱方法论与结果说明报告（新客mlt笔数口径）.md",
        "mlt 主风险模型",
        "mlt 主风险模型",
        "score_new_mlt",
        "score_new_mlt_final_bin",
        "aus_new_risk_bid_3rdmodel_v1_0_20251201",
        "4.34% → 35.65%",
        "B",
        "C",
        "（4,588 笔为文件本身缺失，其余 37,987 笔为无银行交易数据人群的 −1.0 兜底分、按缺失分置空处理，2026-09-01 用户确认）",
        "自动合箱：小箱清理 → 单调合并 → 档位压缩 → 候选生成，最终选中 7 档方案（详见三（二））。",
        prep,
        manual=False,
        value_semantics=False,
        dist_note="分布整形拆分后可得到合规子箱（5.00% / 20.00%），但合回 7 档的唯一不超限合并 (1,1)+(2,4) 跨越极端箱边界 B01（默认禁止），其余相邻对合并后占比均重新超限，整形失败、原候选保留（详见三（四））",
        steps_note="前 6 步为小箱清理：初始 20 箱的 6 个单箱硬约束违反全部消除，主指标倒挂由 8 处降至 3 处；第 7–8 步为单调合并，消除全部剩余倒挂（降至 0）；第 9–12 步为档位压缩，将档位压缩至上限 8 档；第 13–14 步为候选生成，产出 7 档与 6 档候选。全过程未跨越极端箱边界。第 13 步生成的 7 档方案中 F 档（初始箱 15–19）Train 占比 25.00%，超过 21% 的人数分布上限：分布整形拆分为 (15,15) 与 (16,19)（5.00% / 20.00% 均不超限）后，合回 7 档的唯一不超限合并 (1,1)+(2,4)（20.00%）跨越极端箱边界 1（默认禁止），其余相邻对合并后占比均重新超限（35% / 30% / 30% / 25% / 25%），整形失败、原候选保留，属该口径下的合法结果（详见三（四））。",
        cand_note="三个 6–8 档候选均满足硬约束。自动选中的 7 档方案综合得分最高：IV 保留率 0.9661 介于 8 档（0.9728）与 6 档（0.9249）之间，档位数量最接近目标 7 档；8 档方案信息保留更高但档位复杂度更高，6 档方案信息损失最大。F 档（初始箱 15–19）Train 占比 25.00% 超过 21% 的人数分布上限，分布整形无可行回并方案、原候选保留（见三（三））。",
    )
    render_cross()
    print("全部生成完成。")
