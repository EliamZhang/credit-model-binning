"""核对 分箱方法论与结果说明报告（mlt 笔数口径）.md 与最新分箱报告 Excel 的数值一致性。

用法：重跑 binning_mlt_cnt.py 生成最新 Excel 后运行
    python scr/_verify_report_sync_mlt_cnt.py
逐格比对 md 报告各数值表格与 Excel（按 md 显示格式归一化：百分数、千分位、CI、pp、
全精度阈值、布尔约束列），输出所有不一致项；全部一致时打印核对通过。

核对范围：md 中全部数值表格 + 月度稳定性文字断言 + 附录配置关键项。
正文叙述性数字（如 "3.57%→3.53%"）不在此范围。
"""
import glob
import re
import sys

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

MD_PATH = "分箱方法论与结果说明报告（mlt 笔数口径）.md"
SRC_PATTERN = "out/binning_strategy_report_*.xlsx"

# Excel 各 sheet 的表头起始字段，用于按表分段
TABLE_HEADERS = {
    "initial_bin_order", "selected", "step_no", "sample_group", "metric_scope",
    "selected_role", "strategy_name", "threshold_type", "section",
    "final_bin_order", "config_group", "category",
}

issues: list[str] = []
checked = 0


# ---------- md 表格解析 ----------

def strip_md(s: str) -> str:
    return re.sub(r"\*\*", "", s).strip()


def split_row(line: str) -> list[str]:
    return [strip_md(c) for c in line.strip().strip("|").split("|")]


def parse_md_tables(text: str) -> list[dict]:
    tables = []
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if not line.lstrip().startswith("|"):
            i += 1
            continue
        if i + 1 < len(lines) and re.match(r"^\s*\|[\s\-:|]+\|?\s*$", lines[i + 1]):
            header = split_row(line)
            rows = []
            j = i + 2
            while j < len(lines) and lines[j].lstrip().startswith("|"):
                rows.append(split_row(lines[j]))
                j += 1
            tables.append({"header": header, "rows": rows})
            i = j
        else:
            i += 1
    return tables


# ---------- Excel 读取 ----------

def read_sheet_tables(wb, name: str) -> list[dict]:
    ws = wb[name]
    tables = []
    cur = None
    for r in ws.iter_rows(values_only=True):
        if not r or r[0] is None:
            continue
        if isinstance(r[0], str) and r[0] in TABLE_HEADERS:
            cur = {"header": [str(x) if x is not None else "" for x in r], "rows": []}
            tables.append(cur)
        elif cur is not None:
            cur["rows"].append(list(r))
    return tables


def col_of(ex_table, name: str) -> int:
    return ex_table["header"].index(name)


def row_of(ex_table, **criteria):
    idx = {k: col_of(ex_table, k) for k in criteria}
    for r in ex_table["rows"]:
        if all(str(r[idx[k]]).strip() == str(v) for k, v in criteria.items()):
            return r
    return None


# ---------- 单元格比较 ----------

NUM_RE = re.compile(r"^([\-−]?\d[\d,]*(?:\.\d+)?)([%％])?$")


def parse_md_number(cell: str):
    m = NUM_RE.match(cell)
    if not m:
        return None
    num = m.group(1).replace(",", "")
    is_pct = bool(m.group(2))
    decimals = len(num.split(".")[1]) if "." in num else 0
    value = float(num)
    return (value / 100.0 if is_pct else value), decimals, is_pct


def fmt_excel(value, decimals: int, is_pct: bool) -> str:
    if is_pct:
        return f"{value * 100:.{decimals}f}%"
    if decimals == 0:
        return f"{value:,.0f}"
    return f"{value:.{decimals}f}"


def same(a: str, b: str) -> bool:
    return a.replace(",", "") == b.replace(",", "")


def check_cell(md_cell: str, excel_value, where: str):
    global checked
    parsed = parse_md_number(md_cell)
    if parsed is None:
        return
    checked += 1
    value, decimals, is_pct = parsed
    if decimals >= 14:
        expected = repr(excel_value)
    else:
        expected = fmt_excel(excel_value, decimals, is_pct)
    if not same(expected, md_cell):
        issues.append(f"{where}: md={md_cell!r} excel={expected!r} (原始 {excel_value})")


def check_pp_cell(md_cell: str, excel_value, where: str):
    global checked
    m = re.match(r"^(-?\d[\d,]*(?:\.\d+)?)pp$", md_cell)
    if not m:
        return
    checked += 1
    decimals = len(m.group(1).split(".")[1]) if "." in m.group(1) else 0
    expected = f"{excel_value * 100:.{decimals}f}pp"
    if not same(expected, md_cell):
        issues.append(f"{where}: md={md_cell!r} excel={expected!r} (原始 {excel_value})")


def check_ci_cell(md_cell: str, excel_value, ci_low, ci_high, where: str):
    m = re.match(r"^(.*?)\s*\[([^\]]+)\]$", md_cell)
    if not m:
        check_cell(md_cell, excel_value, where)
        return
    check_cell(m.group(1), excel_value, where)
    parts = [p.strip() for p in m.group(2).split(",")]
    if len(parts) == 2:
        check_cell(parts[0], ci_low, f"{where} CI下界")
        check_cell(parts[1], ci_high, f"{where} CI上界")
    else:
        check_cell(parts[0], ci_high, f"{where} CI上界")


def check_bool_cell(md_cell: str, excel_value, where: str):
    global checked
    if "通过" not in md_cell:
        return
    checked += 1
    want = "不通过" not in md_cell
    if bool(excel_value) is not want:
        issues.append(f"{where}: md={md_cell!r} excel={excel_value}")


def check_p_cell(md_cell: str, excel_value, where: str):
    global checked
    if md_cell == "<0.001":
        checked += 1
        if not (isinstance(excel_value, (int, float)) and 0 <= excel_value < 0.001):
            issues.append(f"{where}: md={md_cell!r} excel={excel_value}")
    else:
        check_cell(md_cell, excel_value, where)


def check_text(md_cell: str, excel_value, where: str):
    global checked
    if md_cell in ("", "—", "−"):
        return
    checked += 1
    if str(excel_value).strip() != md_cell:
        issues.append(f"{where}: md={md_cell!r} excel={excel_value}")


def norm_scheme(s: str) -> str:
    return re.sub(r"[\s`]", "", s)


# ---------- 各表核对 ----------

# 分箱大表：md 列名 → 03_最终分箱统计 列名
BIG_COL_MAP = {
    "分数下界": "score_left", "分数上界": "score_right", "样本量": "n",
    "档位占比": "sample_pct", "累计流量": "cum_pass_rate", "策略分段": "strategy_estimated_decision",
    "策略箱流量": "strategy_estimated_bin_flow_rate",
    "1M30+笔数逾期率": "1m30p_cnt_bad_rate", "1M30+金额逾期率": "1m30p_amt_bad_rate",
    "3M30+笔数逾期率": "3m30p_cnt_bad_rate", "3M30+金额逾期率": "3m30p_amt_bad_rate",
    "1M30+笔数Lift": "1m30p_cnt_lift", "1M30+金额Lift": "1m30p_amt_lift",
    "3M30+笔数Lift": "3m30p_cnt_lift", "3M30+金额Lift": "3m30p_amt_lift",
    "累计1M30+笔数逾期率": "cum_1m30p_cnt_bad_rate", "累计1M30+金额逾期率": "cum_1m30p_amt_bad_rate",
    "累计3M30+笔数逾期率": "cum_3m30p_cnt_bad_rate", "累计3M30+金额逾期率": "cum_3m30p_amt_bad_rate",
    "1M30+ IV分项": "1m30p_iv_component", "3M30+ IV分项": "3m30p_iv_component",
    "1M30+ KS曲线值": "1m30p_ks_curve", "3M30+ KS曲线值": "3m30p_ks_curve",
    "PSI分项": "train_oot_psi_component",
    "实际完成率": "actual_completion_rate", "实际审批通过率": "actual_approval_rate",
    "实际自动审批通过率": "actual_auto_approval_rate", "实际人工审批通过率": "actual_manual_approval_rate",
    "实际自动审批占比": "actual_auto_approval_share", "实际人工审批占比": "actual_manual_approval_share",
    "实际成交转化率": "actual_deal_rate",
    "测算自动通过率": "strategy_estimated_overall_auto_pass_rate",
    "测算人工审核率": "strategy_estimated_overall_manual_review_rate",
    "测算总接纳率": "strategy_estimated_overall_total_accept_rate",
    "测算拒绝率": "strategy_estimated_overall_reject_rate",
    "1M30+整体AUC": "overall_1m30p_auc", "3M30+整体AUC": "overall_3m30p_auc",
    "1M30+整体KS": "overall_1m30p_ks", "3M30+整体KS": "overall_3m30p_ks",
    "整体PSI": "train_oot_psi_total",
}


def build_overall(ex, sheet03, sheet05) -> dict:
    """整体行的取值来源：01_总览 + 05 Train/OOT 对照加总。"""
    o1 = {str(r[1]): r[2] for r in ex["01_总览"][0]["rows"]}
    cmp05 = [t for t in ex["05_模型验证"] if "n_train" in t["header"]][0]
    out = {}
    for grp in ("train", "oot"):
        rows = [r for r in cmp05["rows"] if r]
        key = grp
        out[key] = {
            "n": int(o1["Train 样本量" if grp == "train" else "OOT 样本量"]),
            "1m_bad_rate": o1[f"{'train' if grp=='train' else 'oot'}_duedate_1m_30_bad_rate"],
            "3m_bad_rate": o1[f"{'train' if grp=='train' else 'oot'}_duedate_3m_30_bad_rate"],
            "auc_1m": o1[f"{'train' if grp=='train' else 'oot'}_duedate_1m_30_auc"],
            "ks_1m": o1[f"{'train' if grp=='train' else 'oot'}_duedate_1m_30_ks"],
            "auc_3m": o1[f"{'train' if grp=='train' else 'oot'}_duedate_3m_30_auc"],
            "ks_3m": o1[f"{'train' if grp=='train' else 'oot'}_duedate_3m_30_ks"],
            "approval": o1["Train_审批通过率" if grp == "train" else "OOT_审批通过率"],
            "auto_approval": o1["Train_自动审批通过率" if grp == "train" else "OOT_自动审批通过率"],
            "manual_approval": o1["Train_人工审批通过率" if grp == "train" else "OOT_人工审批通过率"],
            "auto_share": o1["Train_自动审批占比" if grp == "train" else "OOT_自动审批占比"],
            "manual_share": o1["Train_人工审批占比" if grp == "train" else "OOT_人工审批占比"],
            "deal": o1["Train_成交转化率" if grp == "train" else "OOT_成交转化率"],
            "est_auto": o1["Train_测算自动通过率" if grp == "train" else "OOT_测算自动通过率"],
            "est_manual": o1["Train_测算人工审核率" if grp == "train" else "OOT_测算人工审核率"],
            "est_accept": o1["Train_测算总接纳率" if grp == "train" else "OOT_测算总接纳率"],
            "est_reject": o1["Train_测算拒绝率" if grp == "train" else "OOT_测算拒绝率"],
            "psi": o1["最终箱 Train/OOT PSI"],
        }
        suf = "_train" if grp == "train" else "_oot"
        out[key]["1m_amt_rate"] = (
            sum(r[col_of(cmp05, f"1m30p_amt_bad{suf}")] or 0 for r in rows)
            / sum(r[col_of(cmp05, f"1m30p_amt_exposure{suf}")] or 0 for r in rows)
        )
        out[key]["3m_amt_rate"] = (
            sum(r[col_of(cmp05, f"3m30p_amt_bad{suf}")] or 0 for r in rows)
            / sum(r[col_of(cmp05, f"3m30p_amt_exposure{suf}")] or 0 for r in rows)
        )
        iv1 = [r[col_of(sheet03, "1m30p_iv_component")] or 0
               for r in sheet03["rows"] if str(r[0]) == ("Train" if grp == "train" else "OOT")]
        iv3 = [r[col_of(sheet03, "3m30p_iv_component")] or 0
               for r in sheet03["rows"] if str(r[0]) == ("Train" if grp == "train" else "OOT")]
        out[key]["iv_1m"] = sum(iv1)
        out[key]["iv_3m"] = sum(iv3)
    return out


OVERALL_SRC = {
    "样本量": "n", "档位占比": None, "累计流量": None, "策略箱流量": None,
    "1M30+笔数逾期率": "1m_bad_rate", "3M30+笔数逾期率": "3m_bad_rate",
    "1M30+金额逾期率": "1m_amt_rate", "3M30+金额逾期率": "3m_amt_rate",
    "1M30+笔数Lift": None, "1M30+金额Lift": None, "3M30+笔数Lift": None, "3M30+金额Lift": None,
    "1M30+ IV分项": "iv_1m", "3M30+ IV分项": "iv_3m",
    "实际完成率": None,
    "实际审批通过率": "approval", "实际自动审批通过率": "auto_approval",
    "实际人工审批通过率": "manual_approval", "实际自动审批占比": "auto_share",
    "实际人工审批占比": "manual_share", "实际成交转化率": "deal",
    "测算自动通过率": "est_auto", "测算人工审核率": "est_manual",
    "测算总接纳率": "est_accept", "测算拒绝率": "est_reject",
    "1M30+整体AUC": "auc_1m", "3M30+整体AUC": "auc_3m",
    "1M30+整体KS": "ks_1m", "3M30+整体KS": "ks_3m", "整体PSI": "psi",
}


def check_big_table(md_tbl, sheet03, overall, where_prefix: str):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        bin_name = row[0]
        where = f"{where_prefix} {bin_name}"
        if bin_name == "整体":
            ov = overall[where_prefix.lower()]
            for md_col, key in OVERALL_SRC.items():
                if md_col not in h:
                    continue
                md_cell = row[h.index(md_col)]
                if md_cell in ("—", ""):
                    continue
                check_cell(md_cell, 1.0 if key is None else ov[key], f"{where}.{md_col}")
            continue
        er = row_of(sheet03, sample_group=where_prefix, score_mlt_final_bin=bin_name)
        if er is None:
            issues.append(f"{where}: Excel 03 中找不到对应行")
            continue
        for md_col, ex_col in BIG_COL_MAP.items():
            if md_col not in h:
                continue
            md_cell = row[h.index(md_col)]
            if md_cell in ("—", ""):
                continue
            ex_val = er[col_of(sheet03, ex_col)]
            if md_col == "策略分段":
                check_text(md_cell, ex_val, f"{where}.{md_col}")
            else:
                check_cell(md_cell, ex_val, f"{where}.{md_col}")


def check_candidates(md_tbl, ex02, issues):
    cand = [t for t in ex02 if t["header"][0] == "selected"][0]
    rows = [r for r in cand["rows"] if str(r[col_of(cand, "hard_constraints_ok")]).strip() == "True"]
    rows = [r for r in rows if int(r[col_of(cand, "final_bin_count")]) in (6, 7, 8)]
    for row in md_tbl["rows"]:
        scheme = norm_scheme(row[1])
        er = next((r for r in rows if norm_scheme(str(r[col_of(cand, "ranges")])) == scheme), None)
        where = f"候选方案 {row[1][:60]}"
        if er is None:
            issues.append(f"{where}: Excel 候选表中找不到")
            continue
        bins = re.search(r"\d+", row[0])
        if bins and int(bins.group()) != int(er[col_of(cand, "final_bin_count")]):
            issues.append(f"{where}: 档位数 md={row[0]} excel={er[col_of(cand, 'final_bin_count')]}")
        md_src = row[2]
        ex_src = "整形" if str(er[col_of(cand, "stage")]) == "share_balancing" else "候选生成"
        if md_src != ex_src:
            issues.append(f"{where}: 来源 md={md_src!r} excel={ex_src!r}")
        check_cell(row[3], er[col_of(cand, "train_primary_inversion_cnt")], f"{where}.主指标倒挂")
        check_cell(row[4], er[col_of(cand, "train_all_inversion_cnt")], f"{where}.全指标倒挂")
        check_cell(row[5], er[col_of(cand, "primary_iv_retention")], f"{where}.IV保留率")
        check_pp_cell(row[6], er[col_of(cand, "min_adjacent_primary_rate_diff")], f"{where}.最小相邻差距")
        check_cell(row[7], er[col_of(cand, "candidate_score")], f"{where}.综合得分")


def check_merge_steps(md_tbl, ex02):
    steps = [t for t in ex02 if t["header"][0] == "step_no"][0]
    stage_map = {"small_bin_cleanup": "约束修正", "granularity_reduction": "档位压缩",
                 "candidate_reduction": "候选生成"}
    for row in md_tbl["rows"]:
        step = int(row[0])
        er = row_of(steps, step_no=step)
        where = f"合箱执行 步{step}"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        ex_stage = stage_map.get(str(er[col_of(steps, "stage")]), str(er[col_of(steps, "stage")]))
        if row[1] != ex_stage:
            issues.append(f"{where}: 阶段 md={row[1]!r} excel={ex_stage!r}")
        merged = norm_scheme(str(er[col_of(steps, "merged_range")]))
        if merged != norm_scheme(row[2]):
            issues.append(f"{where}: 合并箱 md={row[2]!r} excel={merged!r}")
        m = re.match(r"^(.*?)→(.*)$", row[3])
        if m:
            check_cell(m.group(1), er[col_of(steps, "left_primary_rate")], f"{where}.左率")
            check_cell(m.group(2), er[col_of(steps, "right_primary_rate")], f"{where}.右率")
        check_p_cell(row[4], er[col_of(steps, "two_proportion_p_value")], f"{where}.p值")
        check_cell(row[5], er[col_of(steps, "primary_iv_loss")], f"{where}.IV损失")
        if int(row[6]) != 20 - step:
            issues.append(f"{where}: 合并后档位数 md={row[6]} excel={20 - step}")


def check_constraint_table(md_tbl, sheet03):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        bin_name = row[0]
        er = row_of(sheet03, sample_group="Train", score_mlt_final_bin=bin_name)
        where = f"单箱约束 {bin_name}"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        check_cell(row[h.index("样本量")], er[col_of(sheet03, "n")], f"{where}.样本量")
        check_cell(row[h.index("占比")], er[col_of(sheet03, "sample_pct")], f"{where}.占比")
        check_cell(row[h.index("3M30+ 成熟量")], er[col_of(sheet03, "3m30p_cnt_mature")], f"{where}.成熟量")
        check_cell(row[h.index("坏样本量")], er[col_of(sheet03, "3m30p_cnt_bad")], f"{where}.坏样本量")
        check_cell(row[h.index("好样本量")], er[col_of(sheet03, "3m30p_cnt_good")], f"{where}.好样本量")


def check_ci_table(md_tbl, sheet03, sample_group: str):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        bin_name = row[0]
        er = row_of(sheet03, sample_group=sample_group, score_mlt_final_bin=bin_name)
        where = f"最终分箱统计[{sample_group}] {bin_name}"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        check_cell(row[h.index("样本量")], er[col_of(sheet03, "n")], f"{where}.样本量")
        check_cell(row[h.index("占比")], er[col_of(sheet03, "sample_pct")], f"{where}.占比")
        check_ci_cell(row[h.index("1M30+ 笔数逾期率 [95% CI]")],
                      er[col_of(sheet03, "1m30p_cnt_bad_rate")],
                      er[col_of(sheet03, "1m30p_cnt_bad_rate_ci_low")],
                      er[col_of(sheet03, "1m30p_cnt_bad_rate_ci_high")], f"{where}.1M30+率")
        check_ci_cell(row[h.index("3M30+ 笔数逾期率 [95% CI]")],
                      er[col_of(sheet03, "3m30p_cnt_bad_rate")],
                      er[col_of(sheet03, "3m30p_cnt_bad_rate_ci_low")],
                      er[col_of(sheet03, "3m30p_cnt_bad_rate_ci_high")], f"{where}.3M30+率")
        check_cell(row[h.index("3M30+ 金额逾期率")], er[col_of(sheet03, "3m30p_amt_bad_rate")],
                   f"{where}.3M30+金额率")


FUNNEL_CNT_COLS = {
    "申请数": "actual_apply_cnt", "完成进件数": "actual_completed_application_cnt",
    "审批通过数": "actual_approved_application_cnt", "自动审批通过数": "actual_auto_approved_application_cnt",
    "人工审批通过数": "actual_manual_approved_application_cnt", "成交数": "actual_deal_sample_cnt",
}
FUNNEL_RATIO_COLS = {
    "进件完成率": "actual_completion_rate", "审批通过率": "actual_approval_rate",
    "自动审批通过率": "actual_auto_approval_rate", "人工审批通过率": "actual_manual_approval_rate",
    "自动审批占比": "actual_auto_approval_share", "人工审批占比": "actual_manual_approval_share",
    "成交转化率": "actual_deal_rate",
}


def check_funnel(md_cnt, md_ratio, ex04_funnel):
    for md_tbl, colmap in ((md_cnt, FUNNEL_CNT_COLS), (md_ratio, FUNNEL_RATIO_COLS)):
        h = md_tbl["header"]
        for row in md_tbl["rows"]:
            sg = row[0]
            er = row_of(ex04_funnel, sample_group=sg)
            where = f"历史漏斗[{sg}]"
            if er is None:
                issues.append(f"{where}: Excel 中找不到")
                continue
            for md_col, ex_col in colmap.items():
                check_cell(row[h.index(md_col)], er[col_of(ex04_funnel, ex_col)], f"{where}.{md_col}")


def norm_sg(s: str) -> str:
    """md 数据集列（Train/OOT）→ Excel sample_group 值（train/oot）。"""
    return {"Train": "train", "OOT": "oot"}.get(s, s)


def check_threshold_table(md_tbl, sheet03):
    """md 候选档表（7 行 A-G）对应 03_最终分箱统计：阈值=右边界（末档用分数上限），
    累计/边际列直接取 cum_* / 单箱 3m30p 率；约束两列是判断文本，无 Excel 字段。"""
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        bin_name = row[0]
        er = row_of(sheet03, sample_group="Train", score_mlt_final_bin=bin_name)
        where = f"阈值选择 {bin_name}"
        if er is None:
            issues.append(f"{where}: Excel 03 中找不到")
            continue
        sr = er[col_of(sheet03, "score_right")]
        if isinstance(sr, str) and sr.lower() == "inf":
            thr_val = er[col_of(sheet03, "score_max")]
        else:
            thr_val = sr
        check_text(row[h.index("候选档")], er[col_of(sheet03, "score_mlt_final_bin")], f"{where}.档位")
        check_cell(row[h.index("阈值")], thr_val, f"{where}.阈值")
        check_cell(row[h.index("累计通过率")], er[col_of(sheet03, "cum_pass_rate")], f"{where}.累计通过率")
        check_cell(row[h.index("累计 1M30+")], er[col_of(sheet03, "cum_1m30p_cnt_bad_rate")], f"{where}.累计1M30+")
        check_ci_cell(row[h.index("累计 3M30+ [CI 上界]")],
                      er[col_of(sheet03, "cum_3m30p_cnt_bad_rate")],
                      None, er[col_of(sheet03, "cum_3m30p_cnt_bad_rate_ci_high")], f"{where}.累计3M30+")
        check_ci_cell(row[h.index("边际 3M30+ [CI 上界]")],
                      er[col_of(sheet03, "3m30p_cnt_bad_rate")],
                      None, er[col_of(sheet03, "3m30p_cnt_bad_rate_ci_high")], f"{where}.边际3M30+")


SEGMENT_COLS = {
    "样本量": "n", "占比": "strategy_estimated_segment_rate",
    "1M30+ 笔数逾期率": "1m30p_cnt_bad_rate", "3M30+ 笔数逾期率": "3m30p_cnt_bad_rate",
    "3M30+ 金额逾期率": "3m30p_amt_bad_rate",
}


def check_segments(md_tbl, ex04_seg, sample_group: str):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        decision = row[0]
        er = row_of(ex04_seg, sample_group=sample_group, decision=decision)
        where = f"分段流量[{sample_group}] {decision}"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        for md_col, ex_col in SEGMENT_COLS.items():
            check_cell(row[h.index(md_col)], er[col_of(ex04_seg, ex_col)], f"{where}.{md_col}")


SCENARIO_MAP = {"当前": "当前", "收紧一档": "收严一档", "放宽一档": "放松一档"}


def check_sensitivity(md_tbl, ex04_sens):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        th_type = {"自动通过": "自动通过阈值", "总接纳": "总接纳阈值"}.get(row[0], row[0])
        scenario = SCENARIO_MAP.get(row[1], row[1])
        er = row_of(ex04_sens, threshold_type=th_type, scenario=scenario)
        where = f"敏感性[{row[0]} {row[1]}]"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        check_cell(row[h.index("阈值")], er[col_of(ex04_sens, "threshold")], f"{where}.阈值")
        check_text(row[h.index("档位")], er[col_of(ex04_sens, "score_mlt_final_bin")], f"{where}.档位")
        check_cell(row[h.index("自动通过率")], er[col_of(ex04_sens, "strategy_estimated_auto_pass_rate")], f"{where}.自动通过率")
        check_cell(row[h.index("人工审核率")], er[col_of(ex04_sens, "strategy_estimated_manual_review_rate")], f"{where}.人工审核率")
        check_cell(row[h.index("拒绝率")], er[col_of(ex04_sens, "strategy_estimated_reject_rate")], f"{where}.拒绝率")
        check_cell(row[h.index("自动 3M30+")], er[col_of(ex04_sens, "auto_3m30p_cnt_bad_rate")], f"{where}.自动3M30+")
        check_cell(row[h.index("接纳 3M30+")], er[col_of(ex04_sens, "accept_3m30p_cnt_bad_rate")], f"{where}.接纳3M30+")
        check_ci_cell(row[h.index("边际 3M30+ [CI 上界]")],
                      er[col_of(ex04_sens, "accept_marginal_3m30p_cnt_bad_rate")],
                      None, er[col_of(ex04_sens, "accept_marginal_3m30p_cnt_bad_rate_ci_high")],
                      f"{where}.边际3M30+")


MONO_METRICS = {
    "1M30+ 笔数倒挂": "1m30p_cnt_bad_rate", "3M30+ 笔数倒挂": "3m30p_cnt_bad_rate",
    "1M30+ 金额倒挂": "1m30p_amt_bad_rate", "3M30+ 金额倒挂": "3m30p_amt_bad_rate",
}


def check_monotonic(md_tbl, ex05_mono):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        sg = norm_sg(row[0])
        for md_col, metric in MONO_METRICS.items():
            er = row_of(ex05_mono, sample_group=sg, metric=metric)
            where = f"单调性[{sg}] {md_col}"
            if er is None:
                issues.append(f"{where}: Excel 中找不到")
                continue
            check_cell(row[h.index(md_col)], er[col_of(ex05_mono, "violation_cnt")], where)


def check_psi(md_tbl, ex05_psi):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        where = "PSI 表 " + (row[0] or "合计")
        if row[0] == "合计":
            check_cell(row[h.index("PSI 分量")],
                       [r[col_of(ex05_psi, "psi_total")] for r in ex05_psi["rows"]][0], where)
            continue
        er = row_of(ex05_psi, score_mlt_final_bin=row[0])
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        check_cell(row[h.index("Train 占比")], er[col_of(ex05_psi, "train_pct")], f"{where}.Train占比")
        check_cell(row[h.index("OOT 占比")], er[col_of(ex05_psi, "oot_pct")], f"{where}.OOT占比")
        check_cell(row[h.index("PSI 分量")], er[col_of(ex05_psi, "psi_component")], f"{where}.PSI分量")


LABEL_MAP = {"1M30+": "duedate_1m_30", "3M30+": "duedate_3m_30"}


def check_auc_ks(md_tbl, ex05_auc):
    h = md_tbl["header"]
    for row in md_tbl["rows"]:
        sg = norm_sg(row[0])
        label = LABEL_MAP.get(row[1], row[1])
        er = row_of(ex05_auc, sample_group=sg, label=label)
        where = f"AUC/KS[{sg} {row[1]}]"
        if er is None:
            issues.append(f"{where}: Excel 中找不到")
            continue
        check_cell(row[h.index("成熟样本量")], er[col_of(ex05_auc, "n")], f"{where}.成熟样本量")
        check_cell(row[h.index("坏样本量")], er[col_of(ex05_auc, "bad_cnt")], f"{where}.坏样本量")
        check_cell(row[h.index("坏率")], er[col_of(ex05_auc, "bad_rate")], f"{where}.坏率")
        check_cell(row[h.index("AUC")], er[col_of(ex05_auc, "auc")], f"{where}.AUC")
        check_cell(row[h.index("KS")], er[col_of(ex05_auc, "ks")], f"{where}.KS")


def check_monthly(ex05) -> list[str]:
    monthly = [t for t in ex05 if "primary_inversion_count" in t["header"]][0]
    train_rows = [r for r in monthly["rows"] if str(r[0]) == "train"]
    oot_rows = [r for r in monthly["rows"] if str(r[0]) == "oot"]
    t_ok = all(int(r[col_of(monthly, "primary_inversion_count")]) == 0 for r in train_rows)
    oot_inv = [(str(r[col_of(monthly, "application_month")]),
                int(r[col_of(monthly, "primary_inversion_count")]),
                r[col_of(monthly, "max_primary_rate_drop")])
               for r in oot_rows if int(r[col_of(monthly, "primary_inversion_count")]) > 0]
    may = [r for r in oot_rows if str(r[col_of(monthly, "application_month")]) == "2026-05"]
    msg = [f"月度稳定性断言: Train {len(train_rows)} 个月全部无倒挂 = {t_ok}",
           f"  OOT 倒挂月份（应仅 2026-02 且 1 次）: {oot_inv}",
           f"  2026-05 成熟样本量（应为 0）: {may[0][col_of(monthly, 'mature_count')] if may else '未找到'}"]
    return msg


# md 附录未展示、跳过核对的 Excel 配置项
APPENDIX_SKIP = {
    "DATA_DIR", "ACTUAL_FUNNEL_SOURCE", "ACTUAL_FUNNEL_COUNT_KEY",
    "ACTUAL_COMPLETED_EXCLUSIONS", "ACTUAL_APPROVED_PREFIXES", "ACTUAL_DEAL_STATUSES",
    "PSI_EPS", "IV_SMOOTHING_EPS", "ALLOW_EXTREME_BIN_MERGE_FOR_HARD_CONSTRAINTS",
    "EXTREME_BOUNDARY_VIOLATION_PENALTY",
}


def check_appendix(ex06, md_text) -> list[str]:
    appendix = re.sub(r"\s", "", md_text.split("## 六、附录")[-1])
    appendix = appendix.replace("−", "-")  # md 用 Unicode 减号，Excel 值用 ASCII 连字符
    cfg = [t for t in ex06 if t["header"][0] == "config_group"][0]
    missing = []
    for r in cfg["rows"]:
        name, val = r[1], r[2]
        if val is None or name in APPENDIX_SKIP:
            continue
        if not isinstance(val, (int, float, bool)) and not re.search(r"[\d\-\[\]()]", str(val)):
            continue
        sval = re.sub(r"\s", "", str(val))
        if sval not in appendix:
            missing.append(f"{name}={val}")
    if missing:
        return [f"附录配置核对: md 附录未找到 {missing}"]
    return ["附录配置核对: 全部关键配置在 md 附录中一致"]


def main():
    xlsx = sorted(glob.glob(SRC_PATTERN))[-1]
    wb = load_workbook(xlsx, data_only=True)
    ex = {name: read_sheet_tables(wb, name) for name in wb.sheetnames}
    md_text = open(MD_PATH, encoding="utf-8").read()
    md_tables = parse_md_tables(md_text)

    def md_tables_with(*cols, any_of=()):
        return [t for t in md_tables
                if all(c in t["header"] for c in cols)
                and all(any(sub in c for c in t["header"]) for sub in any_of)]

    sheet03 = ex["03_最终分箱统计"][0]
    ex02 = ex["02_分箱详情"]
    ex04 = ex["04_策略方案"]
    ex05 = ex["05_模型验证"]

    overall = build_overall(ex, sheet03, ex05)

    big = md_tables_with("档位", "分数下界")
    check_big_table(big[0], sheet03, overall, "Train")
    check_big_table(big[1], sheet03, overall, "OOT")

    check_candidates(md_tables_with("档位数", "方案")[0], ex02, issues)
    check_merge_steps(md_tables_with("合并初始箱")[0], ex02)
    check_constraint_table(md_tables_with("3M30+ 成熟量")[0], sheet03)
    ci_tbls = md_tables_with(any_of=["[95% CI]"])
    check_ci_table(ci_tbls[0], sheet03, "Train")
    check_ci_table(ci_tbls[1], sheet03, "OOT")

    funnel = [t for t in ex04 if "actual_apply_cnt" in t["header"]][0]
    check_funnel(md_tables_with("申请数")[0], md_tables_with("进件完成率")[0], funnel)

    check_threshold_table(md_tables_with("候选档")[0], sheet03)

    seg = [t for t in ex04 if t["header"][0] == "sample_group"][0]
    seg_tbls = md_tables_with("分段", "占比")
    check_segments(seg_tbls[0], seg, "train")
    check_segments(seg_tbls[1], seg, "oot")

    sens = [t for t in ex04 if t["header"][0] == "threshold_type"][0]
    check_sensitivity(md_tables_with("阈值类型")[0], sens)

    mono = [t for t in ex05 if "is_monotonic_non_decreasing" in t["header"]][0]
    check_monotonic(md_tables_with("1M30+ 笔数倒挂")[0], mono)

    psi = [t for t in ex05 if "psi_component" in t["header"]][0]
    check_psi(md_tables_with("PSI 分量")[0], psi)

    auc = [t for t in ex05 if "auc" in t["header"] and t["header"][0] == "sample_group"][0]
    check_auc_ks(md_tables_with("成熟样本量")[0], auc)

    print(f"=== 核对 {xlsx} vs {MD_PATH} ===")
    for msg in check_monthly(ex05):
        print(msg)
    for msg in check_appendix(ex["06_附录"], md_text):
        print(msg)
    if issues:
        print(f"\n不一致 {len(issues)} 处：")
        for it in issues:
            print(" -", it)
        sys.exit(1)
    print(f"\n核对通过：共 {checked} 个数值单元全部与 Excel 一致。")


if __name__ == "__main__":
    main()
