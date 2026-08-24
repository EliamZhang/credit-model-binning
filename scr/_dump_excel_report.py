"""核对工具：把 out/ 下最新的分箱报告 Excel 全部 sheet 按行 dump 为文本，用于核对 md 报告数值。

用法：重跑 binning.py 后直接 `python scr/_dump_excel_report.py`，输出到 out/_excel_dump.txt。
"""
import glob
import sys

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook

SRC_PATTERN = "out/binning_strategy_report_*.xlsx"
DST = "out/_excel_dump.txt"

matches = sorted(glob.glob(SRC_PATTERN))
if not matches:
    raise SystemExit(f"未找到匹配 {SRC_PATTERN} 的 Excel 文件")
src = matches[-1]

wb = load_workbook(src, data_only=True)
with open(DST, "w", encoding="utf-8") as f:
    for name in wb.sheetnames:
        ws = wb[name]
        f.write(f"===== SHEET: {name} ({ws.max_row} rows) =====\n")
        for row in ws.iter_rows(values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                f.write(" | ".join(vals) + "\n")
        f.write("\n")
print(f"written: {DST} <- {src}")
