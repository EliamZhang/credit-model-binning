# -*- coding: utf-8 -*-
"""报告输出：6-sheet Excel 组装、格式化、总览/附录/指标字典/上线规则。

本模块由旧脚本机械迁移生成：函数体与原实现逐行一致，常量经 settings 同步注入。
新增/修改逻辑时请同步更新对应报告与测试。
"""
import ast
import math
import time
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import pipeline.settings as settings


def _sync_settings() -> None:
    """把 pipeline.settings 的常量刷入本模块全局命名空间（入口脚本在换模型/样本后调用）。"""
    settings.sync(globals())


_sync_settings()


from pipeline.binning_cnt import format_merge_ranges, identify_extreme_boundaries
from pipeline.common import flatten_dict, safe_div



def build_metric_dictionary() -> pd.DataFrame:
    """输出 Excel 核心字段和计算口径说明。"""
    rows = [
        ("通用", "n", "箱内或区间内的申请样本量", "COUNT(application_id)"),
        ("通用", "sample_pct", "箱内样本占全部样本的比例", "n / total_n"),
        ("通用", "principal_amt", "箱内样本本金合计", "SUM(principal)"),
        ("分箱", "score_left / score_right", "分箱的模型分左右边界", "(score_left, score_right]"),
        ("分箱", "merged_from", "最终风险档位由哪些初始箱合并而来", "例如 B06-B08"),
        ("分箱", "extreme_bin_role", "最好/最坏极端箱标识", "由 PROTECT_EXTREME_INITIAL_BINS 相关配置生成"),
        ("分箱", "extreme_boundary_violation_count", "候选方案跨越极端保护边界的数量", "0 表示最好/最坏极端边界被保留"),
        ("笔数风险", "1m30p_cnt_mature", "1M30+ 已成熟样本量", "duedate_1m_30 IN (0, 1)"),
        ("笔数风险", "1m30p_cnt_bad", "1M30+ 逾期样本量", "duedate_1m_30 = 1"),
        ("笔数风险", "1m30p_cnt_bad_rate", "1M30+ 笔数逾期率", "1m30p_cnt_bad / 1m30p_cnt_mature"),
        ("笔数风险", "1m30p_cnt_lift", "1M30+ 笔数 Lift", "1m30p_cnt_bad_rate ÷ 整体 1M30+ 笔数逾期率"),
        ("笔数风险", "3m30p_cnt_mature", "3M30+ 已成熟样本量", "duedate_3m_30 IN (0, 1)"),
        ("笔数风险", "3m30p_cnt_bad", "3M30+ 逾期样本量", "duedate_3m_30 = 1"),
        ("笔数风险", "3m30p_cnt_bad_rate", "3M30+ 笔数逾期率", "3m30p_cnt_bad / 3m30p_cnt_mature"),
        ("笔数风险", "3m30p_cnt_lift", "3M30+ 笔数 Lift", "3m30p_cnt_bad_rate ÷ 整体 3M30+ 笔数逾期率"),
        ("笔数风险", "1m30p_cnt_bad_rate_ci_low / ci_high", "1M30+ 笔数逾期率 95% Wilson 置信区间下/上界", "Wilson 区间（z=1.96）；成熟量为 0 时为空"),
        ("笔数风险", "3m30p_cnt_bad_rate_ci_low / ci_high", "3M30+ 笔数逾期率 95% Wilson 置信区间下/上界", "Wilson 区间（z=1.96）；成熟量为 0 时为空"),
        ("笔数风险", "cum_1m30p_cnt_mature / cum_1m30p_cnt_bad", "累计 1M30+ 已成熟样本量 / 逾期样本量", "按 bin_order 从低风险向高风险逐箱累加"),
        ("笔数风险", "cum_1m30p_cnt_bad_rate", "累计 1M30+ 笔数逾期率", "cum_1m30p_cnt_bad / cum_1m30p_cnt_mature"),
        ("笔数风险", "cum_3m30p_cnt_mature / cum_3m30p_cnt_bad", "累计 3M30+ 已成熟样本量 / 逾期样本量", "按 bin_order 从低风险向高风险逐箱累加"),
        ("笔数风险", "cum_3m30p_cnt_bad_rate", "累计 3M30+ 笔数逾期率", "cum_3m30p_cnt_bad / cum_3m30p_cnt_mature"),
        ("笔数风险", "cum_*_cnt_bad_rate_ci_low / ci_high", "累计笔数逾期率的 95% Wilson 置信区间下/上界", "按累计成熟量与逾期量计算"),
        ("金额风险", "1m30p_amt_exposure", "1M30+ 已成熟样本的本金敞口", "SUM(principal) WHERE MOB1 已成熟"),
        ("金额风险", "1m30p_amt_bad", "MOB1 30+ 样本的剩余本金", "SUM(estimate_principal_remaining_mob1)"),
        ("金额风险", "1m30p_amt_bad_rate", "1M30+ 金额逾期率", "1m30p_amt_bad / 1m30p_amt_exposure"),
        ("金额风险", "1m30p_amt_lift", "1M30+ 金额 Lift", "1m30p_amt_bad_rate ÷ 整体 1M30+ 金额逾期率"),
        ("金额风险", "3m30p_amt_exposure", "3M30+ 已成熟样本的本金敞口", "SUM(principal) WHERE MOB3 已成熟"),
        ("金额风险", "3m30p_amt_bad", "MOB3 30+ 样本的剩余本金", "SUM(estimate_principal_remaining_mob3)"),
        ("金额风险", "3m30p_amt_bad_rate", "3M30+ 金额逾期率", "3m30p_amt_bad / 3m30p_amt_exposure"),
        ("金额风险", "3m30p_amt_lift", "3M30+ 金额 Lift", "3m30p_amt_bad_rate ÷ 整体 3M30+ 金额逾期率"),
        ("金额风险", "cum_1m30p_amt_exposure / cum_1m30p_amt_bad", "累计 1M30+ 本金敞口 / 逾期剩余本金", "按 bin_order 从低风险向高风险逐箱累加"),
        ("金额风险", "cum_1m30p_amt_bad_rate", "累计 1M30+ 金额逾期率", "cum_1m30p_amt_bad / cum_1m30p_amt_exposure"),
        ("金额风险", "cum_3m30p_amt_exposure / cum_3m30p_amt_bad", "累计 3M30+ 本金敞口 / 逾期剩余本金", "按 bin_order 从低风险向高风险逐箱累加"),
        ("金额风险", "cum_3m30p_amt_bad_rate", "累计 3M30+ 金额逾期率", "cum_3m30p_amt_bad / cum_3m30p_amt_exposure"),
        ("模型策略测算", "cum_pass_rate", "从低风险端累计到当前阈值的模型策略测算通过率", "cum_n / total_n"),
        ("阈值", "marginal_sample_pct", "当前档位新增样本占比", "marginal_n / total_n"),
        ("阈值", "marginal_3m30p_cnt_bad_rate", "当前新增档位自身的 3M30+ 风险", "marginal_bad / marginal_mature"),
        ("阈值", "marginal_*_cnt_bad_rate_ci_low / ci_high", "边际档位笔数逾期率 95% Wilson 置信区间下/上界", "按边际档位成熟量与逾期量计算"),
        ("阈值", "auto_all_constraints_ok", "该候选阈值是否满足自动通过全部约束", "全部自动通过检查项均为 True"),
        ("阈值", "accept_all_constraints_ok", "该候选阈值是否满足整体接纳全部约束", "全部整体接纳检查项均为 True"),
        ("历史实际审批漏斗", "actual_completion_rate", "历史实际进件完成率", "completed_application_cnt / apply_cnt"),
        ("历史实际审批漏斗", "actual_approval_rate", "历史实际审批通过率", "approved_application_cnt / completed_application_cnt"),
        ("历史实际审批漏斗", "actual_auto_approval_rate", "历史实际自动审批通过率", "auto_approved_application_cnt / completed_application_cnt"),
        ("历史实际审批漏斗", "actual_manual_approval_rate", "历史实际人工审批通过率", "manual_approved_application_cnt / completed_application_cnt"),
        ("历史实际审批漏斗", "actual_auto_approval_share", "历史实际通过件中的自动审批占比", "auto_approved_application_cnt / approved_application_cnt"),
        ("历史实际审批漏斗", "actual_manual_approval_share", "历史实际通过件中的人工审批占比", "manual_approved_application_cnt / approved_application_cnt"),
        ("历史实际审批漏斗", "actual_deal_rate", "历史实际成交转化率", "deal_sample_cnt / approved_application_cnt"),
        ("箱级历史实际审批漏斗", "actual_*（03_最终分箱统计）", "按 Train/OOT 与最终风险档下钻的历史实际数量和比率", "在每个 score_mlt_final_bin 内按唯一 application_id 复算"),
        ("模型策略测算", "strategy_estimated_auto_pass_rate", "模型阈值测算的自动通过样本占比", "score_mlt 满足自动通过阈值的申请数 / 有效模型分申请数"),
        ("模型策略测算", "strategy_estimated_manual_review_rate", "模型阈值测算的人工审核样本占比", "人工审核分数区间申请数 / 有效模型分申请数"),
        ("模型策略测算", "strategy_estimated_total_accept_rate", "模型阈值测算的总接纳样本占比", "自动通过数与人工审核数之和 / 有效模型分申请数"),
        ("模型策略测算", "strategy_estimated_reject_rate", "模型阈值测算的拒绝样本占比", "超过总接纳阈值的申请数 / 有效模型分申请数"),
        ("箱级模型策略测算", "strategy_estimated_decision", "最终风险档对应的策略归属", "按自动通过阈值和总接纳阈值映射为自动通过/人工审核/拒绝"),
        ("箱级模型策略测算", "strategy_estimated_bin_flow_rate", "该风险档对策略流量的贡献", "箱内申请数 / 当前样本组有效模型分申请数"),
        ("箱级模型策略测算", "strategy_estimated_cumulative_flow_rate", "从低风险端累计至当前档的策略流量", "累计申请数 / 当前样本组有效模型分申请数"),
        ("分箱表整体指标", "strategy_estimated_overall_*_rate", "当前 Train/OOT 样本组的整体策略测算转化率", "整体指标在分箱表中独立成列并重复展示；不作为单箱指标解释"),
        ("分箱表整体指标", "overall_1m30p_auc / overall_3m30p_auc", "当前 Train/OOT 样本组的整体 AUC", "整体指标在分箱表中独立成列并重复展示；AUC 不定义为单箱指标"),
        ("分箱表整体指标", "overall_1m30p_ks / overall_3m30p_ks", "当前 Train/OOT 样本组的整体 KS", "整体指标在分箱表中独立成列并重复展示；与箱级 *_ks_curve 区分"),
        ("箱级模型诊断", "*_iv_component", "1M30+/3M30+ 的箱级 IV 分项", "(bad_dist-good_dist) * LN(bad_dist/good_dist)"),
        ("箱级模型诊断", "*_ks_curve", "由高风险端累计至当前档的 KS 曲线值", "ABS(cum_bad_dist_from_high-cum_good_dist_from_high)"),
        ("箱级模型诊断", "train_oot_psi_component", "当前风险档对 Train/OOT PSI 的贡献", "(OOT%-Train%) * LN(OOT%/Train%)"),
        ("验证", "PSI", "Train 与 OOT 的分箱分布稳定性", "SUM((OOT%-Train%) * LN(OOT%/Train%))"),
        ("验证", "AUC / KS", "模型风险区分能力指标", "分别衡量排序能力和好坏样本累计差异"),
    ]
    return pd.DataFrame(rows, columns=["category", "field", "definition", "calculation"])
def build_online_execution_rules() -> pd.DataFrame:
    """输出上线执行规则的静态清单，供引擎团队上线时逐项核对。"""
    rows = [
        ("分数精度", "模型分精度", "线上评分引擎输出与离线一致的 float 模型分（score_mlt），不限制小数位"),
        ("分数精度", "边界精度", "阈值 = 最终箱右边界原始值（末档为 Train 最大分数），不做二次取整"),
        ("阈值取整", "取整原则", "默认按原始精度部署；若工程必须取整（如存储小数位限制），只允许向更严方向取整：自动通过阈值、总接纳阈值均向下取整（floor），不放大接纳人群"),
        ("阈值取整", "取整后复核", "取整后须重新计算三段占比与风险，与未取整版本对比，确认无风险放大"),
        ("区间开闭", "分档规则", "采用 (left, right]：自动通过 = score ≤ 自动通过阈值；人工审核 = 自动通过阈值 < score ≤ 总接纳阈值；拒绝 = score > 总接纳阈值"),
        ("区间开闭", "边界相等", "分数精确等于阈值时归入右闭档（score == 阈值 → 通过/接纳）"),
        ("区间开闭", "比较方式", "线上用数值比较（float），禁止格式化字符串比较"),
        ("空值与异常值", "缺失模型分", "线上无法产出模型分或为空 → 按拒绝处理；离线报告中缺失样本已从分箱统计剔除，并在 01_总览 展示缺失量与缺失率，两个口径需知悉"),
        ("空值与异常值", "异常分数", "NaN/Inf 不入自动通过，按拒绝处理；分数超出训练范围 → 归入对应极端箱（±∞ 边界兜底）"),
        ("一致性校验", "阈值清单核对", "上线前逐项核对：阈值原始值、开闭符号、末档全量通过点（score ≤ Train 最大分数）"),
        ("一致性校验", "离线/线上对照", "上线前取最近批次样本，离线分档 vs 线上引擎分档，一致率必须 100%，重点覆盖等于边界及边界 ± 1 ulp 的分数"),
        ("一致性校验", "上线后监控", "监控线上分档占比与离线报告各档占比（PSI），边界漂移及时告警复核"),
    ]
    return pd.DataFrame(rows, columns=["category", "item", "rule"])
def build_overview(
    data: pd.DataFrame,
    train: pd.DataFrame,
    oot: pd.DataFrame,
    initial_bin_count: int,
    final_bin_count: int,
    selected_merge_ranges: Sequence[Tuple[int, int]],
    selected_candidate: Optional[pd.Series],
    protected_boundaries: Set[int],
    psi: pd.DataFrame,
    performance: pd.DataFrame,
    monotonicity: pd.DataFrame,
    strategy_plan: pd.DataFrame,
    actual_funnel_report: pd.DataFrame,
    strategy_estimated_flow: pd.DataFrame,
) -> pd.DataFrame:
    """整理报告首页的核心结论，并按模块分组展示。"""
    source_row_count = int(data.attrs.get("source_row_count", len(data)))
    removed_incomplete_count = int(data.attrs.get("removed_incomplete_count", 0))
    score_missing_count = int(data.attrs.get("score_missing_count", 0))

    rows = [
        ("样本", "原始样本量", source_row_count),
        ("样本", "剔除未完成申请量", removed_incomplete_count),
        ("样本", "剔除未完成申请率", safe_div(removed_incomplete_count, source_row_count)),
        ("样本", "有效模型分样本量", len(data)),
        ("样本", "模型分缺失量", score_missing_count),
        ("样本", "模型分缺失率", safe_div(score_missing_count, source_row_count)),
        ("样本", "Train 样本量", len(train)),
        ("样本", "OOT 样本量", len(oot)),
        ("时间切分", "Train 截止月份", TRAIN_END_MONTH),
        ("时间切分", "OOT 起始月份", OOT_START_MONTH),
        ("分箱", "初始箱数量", initial_bin_count),
        ("分箱", "最终箱数量", final_bin_count),
        ("分箱", "合箱主指标", " / ".join(PRIMARY_RATE_COLS)),
        ("分箱", "最终采用合箱方案", format_merge_ranges(selected_merge_ranges)),
        ("分箱", "受保护初始边界", ",".join(map(str, sorted(protected_boundaries)))),
        ("稳定性", "最终箱 Train/OOT PSI", psi["psi_total"].iloc[0]),
    ]

    if selected_candidate is not None:
        rows.extend(
            [
                ("候选评分", "Train 主指标倒挂数", selected_candidate.get("train_primary_inversion_cnt")),
                ("候选评分", "Train 全指标倒挂数", selected_candidate.get("train_all_inversion_cnt")),
                ("候选评分", "主指标 IV 保留率", selected_candidate.get("primary_iv_retention")),
                ("候选评分", "候选综合得分", selected_candidate.get("candidate_score")),
            ]
        )

    for _, perf in performance.iterrows():
        prefix = f"{perf['sample_group']}_{perf['label']}"
        rows.extend(
            [
                ("模型效果", f"{prefix}_bad_rate", perf["bad_rate"]),
                ("模型效果", f"{prefix}_auc", perf["auc"]),
                ("模型效果", f"{prefix}_ks", perf["ks"]),
            ]
        )

    for sample_group in ["train", "oot"]:
        sample_check = monotonicity.loc[monotonicity["sample_group"].eq(sample_group)]
        if sample_check.empty:
            continue
        rows.append(
            (
                "单调性",
                f"{sample_group}_最终箱全部单调",
                bool(sample_check["is_monotonic_non_decreasing"].all()),
            )
        )

    for sample_group in ["Train", "OOT"]:
        actual_rows = actual_funnel_report.loc[
            actual_funnel_report["sample_group"].eq(sample_group)
        ]
        if not actual_rows.empty:
            actual = actual_rows.iloc[0]
            rows.extend(
                [
                    ("历史实际审批漏斗", f"{sample_group}_进件完成率", actual["actual_completion_rate"]),
                    ("历史实际审批漏斗", f"{sample_group}_审批通过率", actual["actual_approval_rate"]),
                    ("历史实际审批漏斗", f"{sample_group}_自动审批通过率", actual["actual_auto_approval_rate"]),
                    ("历史实际审批漏斗", f"{sample_group}_人工审批通过率", actual["actual_manual_approval_rate"]),
                    ("历史实际审批漏斗", f"{sample_group}_自动审批占比", actual["actual_auto_approval_share"]),
                    ("历史实际审批漏斗", f"{sample_group}_人工审批占比", actual["actual_manual_approval_share"]),
                    ("历史实际审批漏斗", f"{sample_group}_成交转化率", actual["actual_deal_rate"]),
                ]
            )

        estimated_rows = strategy_estimated_flow.loc[
            strategy_estimated_flow["sample_group"].eq(sample_group)
        ]
        if not estimated_rows.empty:
            estimated = estimated_rows.iloc[0]
            rows.extend(
                [
                    ("模型策略测算流量", f"{sample_group}_测算自动通过率", estimated["strategy_estimated_auto_pass_rate"]),
                    ("模型策略测算流量", f"{sample_group}_测算人工审核率", estimated["strategy_estimated_manual_review_rate"]),
                    ("模型策略测算流量", f"{sample_group}_测算总接纳率", estimated["strategy_estimated_total_accept_rate"]),
                    ("模型策略测算流量", f"{sample_group}_测算拒绝率", estimated["strategy_estimated_reject_rate"]),
                ]
            )

    valid_strategy = strategy_plan.loc[strategy_plan["status"].eq("OK")]
    if not valid_strategy.empty:
        row = valid_strategy.iloc[0]
        rows.extend(
            [
                ("模型策略阈值", "策略名称", row["strategy_name"]),
                ("模型策略阈值", "自动通过阈值", row["auto_pass_threshold"]),
                ("模型策略阈值", "自动通过截止风险档", row["auto_pass_bin"]),
                ("模型策略阈值", "人工审核上限/拒绝阈值", row["reject_threshold"]),
                ("模型策略阈值", "人工审核截止风险档", row["manual_review_upper_bin"]),
                ("策略风险", "接纳人群1M30+笔数逾期率", row["accepted_1m30p_cnt_bad_rate"]),
                ("策略风险", "接纳人群3M30+笔数逾期率", row["accepted_3m30p_cnt_bad_rate"]),
                ("策略风险", "接纳人群1M30+金额逾期率", row["accepted_1m30p_amt_bad_rate"]),
                ("策略风险", "接纳人群3M30+金额逾期率", row["accepted_3m30p_amt_bad_rate"]),
                ("策略风险", "最后接纳档边际3M30+", row["last_accepted_marginal_3m30p_cnt_bad_rate"]),
            ]
        )
    else:
        status = strategy_plan.iloc[0]["status"] if not strategy_plan.empty else "未生成"
        rows.append(("策略", "策略状态", status))

    return pd.DataFrame(rows, columns=["section", "metric", "value"])
def build_config_table(
    selected_merge_ranges: Sequence[Tuple[int, int]],
    protected_boundaries: Set[int],
) -> pd.DataFrame:
    """输出便于后续修改和版本管理的参数表。"""
    actual_initial_bin_count = max((end for _, end in selected_merge_ranges), default=0)
    extreme_boundaries_text = ",".join(
        map(str, sorted(identify_extreme_boundaries(actual_initial_bin_count)))
    )
    rows = [
        {"config_group": "基础配置", "config_name": "DATA_DIR", "config_value": str(DATA_DIR)},
        {"config_group": "基础配置", "config_name": "TRAIN_END_MONTH", "config_value": TRAIN_END_MONTH},
        {"config_group": "基础配置", "config_name": "OOT_START_MONTH", "config_value": OOT_START_MONTH},
        {"config_group": "基础配置", "config_name": "INITIAL_BIN_COUNT", "config_value": INITIAL_BIN_COUNT},
        {"config_group": "基础配置", "config_name": "HIGH_SCORE_HIGH_RISK", "config_value": HIGH_SCORE_HIGH_RISK},
        {"config_group": "历史实际审批漏斗", "config_name": "ACTUAL_FUNNEL_SOURCE", "config_value": "application_info.csv"},
        {"config_group": "历史实际审批漏斗", "config_name": "ACTUAL_FUNNEL_COUNT_KEY", "config_value": "COUNT DISTINCT application_id"},
        {"config_group": "历史实际审批漏斗", "config_name": "ACTUAL_COMPLETED_EXCLUSIONS", "config_value": "0.Incomplete,1.In Progress"},
        {"config_group": "历史实际审批漏斗", "config_name": "ACTUAL_APPROVED_PREFIXES", "config_value": "3,4"},
        {"config_group": "历史实际审批漏斗", "config_name": "ACTUAL_DEAL_STATUSES", "config_value": "Active_Account,Closed,Blocked"},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_COUNT", "config_value": MIN_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "MAX_FINAL_BIN_COUNT", "config_value": MAX_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "TARGET_FINAL_BIN_COUNT", "config_value": TARGET_FINAL_BIN_COUNT},
        {"config_group": "合箱配置", "config_name": "PRIMARY_RATE_COLS", "config_value": " / ".join(PRIMARY_RATE_COLS)},
        {"config_group": "合箱配置", "config_name": "PRIMARY_RATE_COL", "config_value": PRIMARY_RATE_COL},
        {"config_group": "合箱配置", "config_name": "MIN_MIDDLE_BIN_SAMPLE_PCT", "config_value": MIN_MIDDLE_BIN_SAMPLE_PCT},
        {"config_group": "合箱配置", "config_name": "MIN_TAIL_BIN_SAMPLE_PCT", "config_value": MIN_TAIL_BIN_SAMPLE_PCT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_MATURE_COUNT", "config_value": MIN_FINAL_BIN_MATURE_COUNT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_BAD_COUNT", "config_value": MIN_FINAL_BIN_BAD_COUNT},
        {"config_group": "合箱配置", "config_name": "MIN_FINAL_BIN_GOOD_COUNT", "config_value": MIN_FINAL_BIN_GOOD_COUNT},
        {"config_group": "极端箱配置", "config_name": "MIN_EXTREME_BIN_MATURE_COUNT", "config_value": MIN_EXTREME_BIN_MATURE_COUNT},
        {"config_group": "极端箱配置", "config_name": "MIN_BEST_EXTREME_BIN_BAD_COUNT", "config_value": MIN_BEST_EXTREME_BIN_BAD_COUNT},
        {"config_group": "极端箱配置", "config_name": "MIN_BEST_EXTREME_BIN_GOOD_COUNT", "config_value": MIN_BEST_EXTREME_BIN_GOOD_COUNT},
        {"config_group": "极端箱配置", "config_name": "MIN_WORST_EXTREME_BIN_BAD_COUNT", "config_value": MIN_WORST_EXTREME_BIN_BAD_COUNT},
        {"config_group": "极端箱配置", "config_name": "MIN_WORST_EXTREME_BIN_GOOD_COUNT", "config_value": MIN_WORST_EXTREME_BIN_GOOD_COUNT},
        {"config_group": "合箱配置", "config_name": "TRAIN_INVERSION_TOLERANCE", "config_value": TRAIN_INVERSION_TOLERANCE},
        {"config_group": "合箱配置", "config_name": "MONTHLY_INVERSION_TOLERANCE", "config_value": MONTHLY_INVERSION_TOLERANCE},
        {"config_group": "合箱配置", "config_name": "ADJACENT_PVALUE_TO_MERGE", "config_value": ADJACENT_PVALUE_TO_MERGE},
        {"config_group": "合箱配置", "config_name": "MIN_ADJACENT_ABS_RATE_DIFF", "config_value": MIN_ADJACENT_ABS_RATE_DIFF},
        {"config_group": "合箱配置", "config_name": "MAX_FINAL_BIN_SHARE", "config_value": MAX_FINAL_BIN_SHARE},
        {"config_group": "合箱配置", "config_name": "PROTECTED_BOUNDARIES", "config_value": ",".join(map(str, sorted(protected_boundaries)))},
        {"config_group": "极端箱配置", "config_name": "PROTECT_EXTREME_INITIAL_BINS", "config_value": PROTECT_EXTREME_INITIAL_BINS},
        {"config_group": "极端箱配置", "config_name": "BEST_EXTREME_INITIAL_BIN_COUNT", "config_value": BEST_EXTREME_INITIAL_BIN_COUNT},
        {"config_group": "极端箱配置", "config_name": "WORST_EXTREME_INITIAL_BIN_COUNT", "config_value": WORST_EXTREME_INITIAL_BIN_COUNT},
        {"config_group": "极端箱配置", "config_name": "ALLOW_EXTREME_BIN_MERGE_FOR_HARD_CONSTRAINTS", "config_value": ALLOW_EXTREME_BIN_MERGE_FOR_HARD_CONSTRAINTS},
        {"config_group": "极端箱配置", "config_name": "EXTREME_BOUNDARIES", "config_value": extreme_boundaries_text},
        {"config_group": "极端箱配置", "config_name": "EXTREME_BOUNDARY_PENALTY", "config_value": EXTREME_BOUNDARY_PENALTY},
        {"config_group": "极端箱配置", "config_name": "EXTREME_BOUNDARY_VIOLATION_PENALTY", "config_value": EXTREME_BOUNDARY_VIOLATION_PENALTY},
        {"config_group": "评分配置", "config_name": "PROTECTED_BOUNDARY_PENALTY", "config_value": PROTECTED_BOUNDARY_PENALTY},
        {"config_group": "评分配置", "config_name": "MERGE_COST_RATE_GAP_WEIGHT", "config_value": MERGE_COST_RATE_GAP_WEIGHT},
        {"config_group": "评分配置", "config_name": "MERGE_COST_IV_LOSS_WEIGHT", "config_value": MERGE_COST_IV_LOSS_WEIGHT},
        {"config_group": "评分配置", "config_name": "IV_RETENTION_SCORE_CAP", "config_value": IV_RETENTION_SCORE_CAP},
        {"config_group": "评分配置", "config_name": "PSI_EPS", "config_value": PSI_EPS},
        {"config_group": "评分配置", "config_name": "IV_SMOOTHING_EPS", "config_value": IV_SMOOTHING_EPS},
        {"config_group": "合箱配置", "config_name": "SELECTED_FINAL_BIN_RANGES", "config_value": format_merge_ranges(selected_merge_ranges)},
    ]

    for name, value in CANDIDATE_SCORE_WEIGHTS.items():
        rows.append(
            {
                "config_group": "候选评分权重",
                "config_name": name,
                "config_value": value,
            }
        )

    flattened = {
        **flatten_dict("auto", STRATEGY_CONFIG["auto_constraints"]),
        **flatten_dict("accept", STRATEGY_CONFIG["accept_constraints"]),
    }
    for name, value in flattened.items():
        rows.append(
            {
                "config_group": STRATEGY_CONFIG["strategy_name"],
                "config_name": name,
                "config_value": value,
            }
        )

    return pd.DataFrame(rows)
def _detect_sections(ws) -> List[Tuple[int, int, int]]:
    """通过空行分隔符检测 sheet 中的 section 边界。"""
    sections: List[Tuple[int, int, int]] = []
    if ws.max_row < 1:
        return sections

    header_row = 1
    data_start = 2
    data_end = data_start
    while data_end <= ws.max_row:
        cells = [ws.cell(data_end, col).value for col in range(1, ws.max_column + 1)]
        if all(c is None for c in cells):
            sections.append((header_row, data_start, data_end - 1))
            header_row = data_end + 1
            data_start = header_row + 1
            data_end = data_start
        else:
            data_end += 1
    sections.append((header_row, data_start, data_end - 1))
    return sections
def _write_sections(writer, sheet_name: str, sections: List[Tuple[str, pd.DataFrame]]) -> None:
    """在一个 sheet 中写入多个 DataFrame，section 之间空一行分隔。"""
    startrow = 0
    for label, df in sections:
        if df is None or df.empty:
            continue
        df.to_excel(writer, sheet_name=sheet_name, startrow=startrow, index=False)
        startrow += len(df) + 2
def format_excel_report(path: Path) -> None:
    """按 section 设置格式：表头样式、数字格式、条件高亮。"""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    body_font = Font(name="微软雅黑", size=10)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="微软雅黑", color="FFFFFF", bold=True, size=10)
    selected_fill = PatternFill("solid", fgColor="C6E0B4")
    selected_reject_fill = PatternFill("solid", fgColor="FCE4D6")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    fail_fill = PatternFill("solid", fgColor="F4CCCC")

    for sheet in workbook.worksheets:
        sections = _detect_sections(sheet)
        if not sections:
            continue

        sheet.freeze_panes = "D2" if sheet.title.startswith("03_") else "A2"
        sheet.sheet_view.showGridLines = False

        first_data_end = sections[0][2]
        if first_data_end >= sections[0][1]:
            sheet.auto_filter.ref = (
                f"A{sections[0][0]}:"
                f"{get_column_letter(sheet.max_column)}{first_data_end}"
            )

        # 列宽。
        for col_idx in range(1, sheet.max_column + 1):
            max_length = 0
            for row_idx in range(1, sheet.max_row + 1):
                value = sheet.cell(row_idx, col_idx).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(
                max(max_length + 2, 10), 42
            )

        # 按 section 格式化。
        for header_row, data_start, data_end in sections:
            headers = {cell.column: str(cell.value) for cell in sheet[header_row]}
            header_to_col = {str(cell.value): cell.column for cell in sheet[header_row]}

            for cell in sheet[header_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for row_idx in range(data_start, data_end + 1):
                for cell in sheet[row_idx]:
                    cell.font = body_font
                    header = headers.get(cell.column, "").lower()
                    if cell.value is None:
                        continue
                    header_tokens = set(header.split("_"))
                    if header_tokens.intersection(
                        {"rate", "pct", "retention", "share", "distribution"}
                    ):
                        cell.number_format = "0.00%"
                    elif (
                        header_tokens.intersection({"auc", "ks", "psi", "corr", "iv", "lift"})
                        or "p_value" in header
                    ):
                        cell.number_format = "0.0000"
                    elif any(key in header for key in [
                        "threshold", "score_left", "score_right",
                        "score_min", "score_max", "score_mean",
                    ]):
                        cell.number_format = "0.0000"
                    elif isinstance(cell.value, (int, np.integer)) and not isinstance(cell.value, bool):
                        cell.number_format = "#,##0"
                    elif isinstance(cell.value, (int, float)):
                        cell.number_format = "#,##0.00"

            # 条件高亮：倒挂标记。
            inversion_cols = [
                header_to_col.get(k)
                for k in ["1m30p_inversion_flag", "3m30p_inversion_flag",
                           "primary_inversion_flag"]
            ]
            inversion_cols = [c for c in inversion_cols if c is not None]
            if inversion_cols:
                for row_idx in range(data_start, data_end + 1):
                    if any(sheet.cell(row_idx, col).value is True for col in inversion_cols):
                        for cell in sheet[row_idx]:
                            cell.fill = warning_fill

            # 条件高亮：阈值选中行。
            selected_role_col = header_to_col.get("selected_role")
            auto_ok_col = header_to_col.get("auto_all_constraints_ok")
            accept_ok_col = header_to_col.get("accept_all_constraints_ok")
            if selected_role_col:
                for row_idx in range(data_start, data_end + 1):
                    role = sheet.cell(row_idx, selected_role_col).value or ""
                    if "自动通过" in str(role):
                        for cell in sheet[row_idx]:
                            cell.fill = selected_fill
                    if "拒绝阈值" in str(role):
                        for cell in sheet[row_idx]:
                            cell.fill = selected_reject_fill
                    for check_col in [auto_ok_col, accept_ok_col]:
                        if check_col and sheet.cell(row_idx, check_col).value is False:
                            sheet.cell(row_idx, check_col).fill = fail_fill

            # 条件高亮：阈值敏感性表中的当前方案行。
            scenario_col = header_to_col.get("scenario")
            if scenario_col:
                for row_idx in range(data_start, data_end + 1):
                    if sheet.cell(row_idx, scenario_col).value == "当前":
                        for cell in sheet[row_idx]:
                            cell.fill = selected_fill

            # 条件高亮：合箱候选选中行。
            selected_flag_col = header_to_col.get("selected")
            hard_ok_col = header_to_col.get("hard_constraints_ok")
            if selected_flag_col:
                for row_idx in range(data_start, data_end + 1):
                    if sheet.cell(row_idx, selected_flag_col).value is True:
                        for cell in sheet[row_idx]:
                            cell.fill = selected_fill
                    elif hard_ok_col and sheet.cell(row_idx, hard_ok_col).value is False:
                        sheet.cell(row_idx, hard_ok_col).fill = fail_fill

    workbook.save(path)
def write_report(
    overview: pd.DataFrame,
    binning_process: pd.DataFrame,
    final_train_stats: pd.DataFrame,
    final_oot_stats: pd.DataFrame,
    train_oot_compare: pd.DataFrame,
    actual_funnel_report: pd.DataFrame,
    strategy_estimated_flow: pd.DataFrame,
    threshold_selection: pd.DataFrame,
    strategy_plan: pd.DataFrame,
    threshold_sensitivity: pd.DataFrame,
    strategy_segments: pd.DataFrame,
    performance: pd.DataFrame,
    psi: pd.DataFrame,
    monotonicity: pd.DataFrame,
    monthly_stability: pd.DataFrame,
    monthly_stability_summary: pd.DataFrame,
    merge_candidates: pd.DataFrame,
    merge_steps: pd.DataFrame,
    config_table: pd.DataFrame,
    online_execution_rules: pd.DataFrame,
    metric_dictionary: pd.DataFrame,
) -> None:
    """输出精简版策略报告（6 个 sheet）。"""
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    # 合并 Train/OOT 最终分箱统计：同结构 + sample_group 列。
    final_stats_parts = []
    for label, stats in [
        ("Train", final_train_stats),
        ("OOT", final_oot_stats),
    ]:
        s = stats.copy()
        s.insert(0, "sample_group", label)
        final_stats_parts.append(s)
    final_stats_all = pd.concat(final_stats_parts, ignore_index=True)

    # 将样本组整体策略转化率与整体 AUC/KS 作为独立字段并入分箱表。
    # 字段名明确标注 overall，避免被误解为单箱指标。
    overall_strategy = strategy_estimated_flow.loc[
        strategy_estimated_flow["sample_group"].isin(["Train", "OOT"]),
        [
            "sample_group",
            "strategy_estimated_auto_pass_rate",
            "strategy_estimated_manual_review_rate",
            "strategy_estimated_total_accept_rate",
            "strategy_estimated_reject_rate",
        ],
    ].rename(
        columns={
            "strategy_estimated_auto_pass_rate": "strategy_estimated_overall_auto_pass_rate",
            "strategy_estimated_manual_review_rate": "strategy_estimated_overall_manual_review_rate",
            "strategy_estimated_total_accept_rate": "strategy_estimated_overall_total_accept_rate",
            "strategy_estimated_reject_rate": "strategy_estimated_overall_reject_rate",
        }
    )
    final_stats_all = final_stats_all.merge(
        overall_strategy,
        on="sample_group",
        how="left",
        validate="many_to_one",
    )

    overall_performance = performance.copy()
    overall_performance["sample_group"] = (
        overall_performance["sample_group"].astype(str).str.lower().map(
            {"train": "Train", "oot": "OOT"}
        )
    )
    performance_wide = overall_performance.pivot(
        index="sample_group",
        columns="label",
        values=["auc", "ks"],
    )
    performance_wide.columns = [
        f"overall_{'1m30p' if label == 'duedate_1m_30' else '3m30p'}_{metric}"
        for metric, label in performance_wide.columns
    ]
    final_stats_all = final_stats_all.merge(
        performance_wide.reset_index(),
        on="sample_group",
        how="left",
        validate="many_to_one",
    )

    headline_columns = [
        "sample_group",
        "bin_order",
        FINAL_BIN_COL,
        "score_left",
        "score_right",
        "n",
        "sample_pct",
        "cum_pass_rate",
        "strategy_estimated_decision",
        "strategy_estimated_bin_flow_rate",
        "actual_completion_rate",
        "actual_approval_rate",
        "actual_auto_approval_rate",
        "actual_manual_approval_rate",
        "actual_auto_approval_share",
        "actual_manual_approval_share",
        "actual_deal_rate",
        "1m30p_cnt_bad_rate",
        "1m30p_amt_bad_rate",
        "3m30p_cnt_bad_rate",
        "3m30p_amt_bad_rate",
        "1m30p_cnt_lift",
        "1m30p_amt_lift",
        "3m30p_cnt_lift",
        "3m30p_amt_lift",
        "cum_1m30p_cnt_mature",
        "cum_1m30p_cnt_bad",
        "cum_1m30p_cnt_bad_rate",
        "cum_1m30p_amt_exposure",
        "cum_1m30p_amt_bad",
        "cum_1m30p_amt_bad_rate",
        "cum_3m30p_cnt_mature",
        "cum_3m30p_cnt_bad",
        "cum_3m30p_cnt_bad_rate",
        "cum_3m30p_amt_exposure",
        "cum_3m30p_amt_bad",
        "cum_3m30p_amt_bad_rate",
        "1m30p_iv_component",
        "3m30p_iv_component",
        "1m30p_ks_curve",
        "3m30p_ks_curve",
        "train_oot_psi_component",
        "strategy_estimated_overall_auto_pass_rate",
        "strategy_estimated_overall_manual_review_rate",
        "strategy_estimated_overall_total_accept_rate",
        "strategy_estimated_overall_reject_rate",
        "overall_1m30p_auc",
        "overall_3m30p_auc",
        "overall_1m30p_ks",
        "overall_3m30p_ks",
        "train_oot_psi_total",
    ]
    headline_columns = [col for col in headline_columns if col in final_stats_all.columns]
    remaining_columns = [
        col for col in final_stats_all.columns if col not in headline_columns
    ]
    final_stats_all = final_stats_all[headline_columns + remaining_columns]

    with pd.ExcelWriter(REPORT_PATH, engine="openpyxl") as writer:
        overview.to_excel(writer, sheet_name="01_总览", index=False)

        _write_sections(writer, "02_分箱详情", [
            ("分箱过程", binning_process),
            ("合箱候选评分", merge_candidates),
            ("合箱步骤", merge_steps),
        ])

        final_stats_all.to_excel(writer, sheet_name="03_最终分箱统计", index=False)

        _write_sections(writer, "04_策略方案", [
            ("历史实际审批漏斗", actual_funnel_report),
            ("模型策略测算流量", strategy_estimated_flow),
            ("阈值选择过程", threshold_selection),
            ("模型策略测算结果", strategy_plan),
            ("模型策略测算阈值敏感性", threshold_sensitivity),
            ("模型策略测算分段风险验证", strategy_segments),
        ])

        _write_sections(writer, "05_模型验证", [
            ("Train_OOT对比", train_oot_compare),
            ("AUC_KS", performance),
            ("PSI", psi),
            ("单调性", monotonicity),
            ("月度稳定性汇总", monthly_stability_summary),
            ("月度箱表现", monthly_stability),
        ])

        _write_sections(writer, "06_附录", [
            ("配置参数", config_table),
            ("上线执行规则", online_execution_rules),
            ("指标说明", metric_dictionary),
        ])

    format_excel_report(REPORT_PATH)
