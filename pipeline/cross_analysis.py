# -*- coding: utf-8 -*-
"""
两模型交叉分析（任意两模型组合，配置驱动）。

matrix 模式：A 模型 7 档 × B 模型 7 档的交叉矩阵、条件增量、组合评分与二维策略模拟
（对应旧脚本 binning_cross_mlt_wth.py 的逻辑，数值口径不变）。
cond 模式：在 A 模型各档内对 B 模型分做 3 等频条件子箱（A1–G3 共 21 格），
评估档内显著性、头尾拉开、整体区分度与二维策略（对应旧脚本
binning_worthiness_cond_cnt.py 的逻辑，数值口径不变）。

模型组合通过 configs/models.py 注册的 key 指定，入口为 scripts/cross_models.py。
"""
import contextlib
import time
from itertools import product
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import pipeline.settings as settings
from configs.datasets import DATASETS
from configs.models import MODELS
from pipeline import bin_amt, binning_cnt, data_loading, monthly, reporting, risk_metrics, strategy
from pipeline.common import wilson_ci
from pipeline.data_loading import _actual_funnel_row
from pipeline.risk_metrics import calc_auc_ks, two_proportion_pvalue

# ============================================================
# 1. 模型上下文：在指定数据集/模型配置下运行管线
# ============================================================

_SYNC_MODULES = (data_loading, risk_metrics, binning_cnt, strategy, monthly, reporting, bin_amt)


@contextlib.contextmanager
def model_context(dataset_cfg: dict, model_cfg: dict, metric: str = "cnt"):
    """临时应用数据集+模型配置并同步各模块；退出时恢复默认（老客 mlt cnt）。"""
    settings.apply_dataset(dataset_cfg)
    settings.apply_model(model_cfg)
    settings.apply_metric(metric)
    for module in _SYNC_MODULES:
        module._sync_settings()
    try:
        yield
    finally:
        settings.apply_dataset(DATASETS["laoke"])
        settings.apply_model(MODELS["mlt"])
        settings.apply_metric("cnt")
        for module in _SYNC_MODULES:
            module._sync_settings()


def build_model_final_bins(
    dataset_cfg: dict,
    model_cfg: dict,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[Tuple[int, int]]]:
    """跑单模型管线的分箱部分，返回 (最终分档宽表, 最终边界表, 选中方案)。"""
    with model_context(dataset_cfg, model_cfg):
        data = data_loading.load_analysis_data()
        all_data, train, oot = data_loading.split_train_oot(data)
        edges = binning_cnt.learn_equal_freq_edges(train, settings.SCORE_COL, settings.INITIAL_BIN_COUNT)
        initial_edges = binning_cnt.build_initial_edge_table(edges)
        initial_bin_count = len(edges) - 1
        all_binned = binning_cnt.apply_edges(all_data, settings.SCORE_COL, edges, settings.INITIAL_BIN_COL)
        train_binned = all_binned.loc[all_binned["sample_group"].eq("train")].copy()
        train_initial_stats = binning_cnt.calc_complete_initial_stats(train_binned, initial_edges)
        candidates, _steps, _protected = binning_cnt.build_merge_candidate_score_table(
            train_initial_stats,
            initial_bin_count,
            settings.STRATEGY_CONFIG,
        )
        selected_ranges = binning_cnt.selected_ranges_from_candidate_table(candidates)
        merge_map = binning_cnt.build_merge_map(selected_ranges, initial_bin_count)
        final_edges = binning_cnt.build_final_edge_table(initial_edges, merge_map, initial_bin_count)
        final = binning_cnt.apply_merge_map(all_binned, merge_map)
        print(
            f"[{model_cfg['name']}] 最终方案：{len(final_edges)} 档 "
            f"{binning_cnt.format_merge_ranges(selected_ranges)}"
        )
        return final, final_edges, selected_ranges


# ============================================================
# 2. 共享辅助
# ============================================================


def rate_of(frame: pd.DataFrame, due_col: str) -> Tuple[int, int, float]:
    """笔数口径：返回 (成熟量, 坏样本量, 逾期率)。"""
    values = pd.to_numeric(frame[due_col], errors="coerce")
    mature_mask = values.isin([0, 1])
    mature = int(mature_mask.sum())
    bad = int(values.eq(1).sum())
    return mature, bad, (bad / mature if mature else np.nan)


def amt_rate_of(frame: pd.DataFrame, dpd_col: str, remaining_col: str) -> float:
    """金额口径逾期率：dpd >= 30 样本的剩余本金 / 成熟样本本金敞口。"""
    dpd = pd.to_numeric(frame[dpd_col], errors="coerce")
    principal = pd.to_numeric(frame["principal"], errors="coerce").fillna(0)
    remaining = pd.to_numeric(frame[remaining_col], errors="coerce")
    exposure = float(principal.loc[dpd.notna()].sum())
    bad_amt = float(remaining.loc[dpd.ge(30)].fillna(0).sum())
    return bad_amt / exposure if exposure else np.nan


def group_overall_rates(frame: pd.DataFrame) -> Dict[str, float]:
    """样本组整体四项风险率，用于 Lift 计算。"""
    out = {}
    for prefix, due_col, dpd_col, remaining_col in [
        ("1m30p_cnt", "duedate_1m_30", "dpd_days_ever_mob1", "estimate_principal_remaining_mob1"),
        ("3m30p_cnt", "duedate_3m_30", "dpd_days_ever_mob3", "estimate_principal_remaining_mob3"),
    ]:
        _, _, rate = rate_of(frame, due_col)
        out[f"{prefix}_bad_rate"] = rate
        out[f"{prefix[:-4]}_amt_bad_rate"] = amt_rate_of(frame, dpd_col, remaining_col)
    return out


def write_sheet(
    wb: Workbook,
    sheet_name: str,
    frames: Sequence[Tuple[Optional[str], Optional[pd.DataFrame]]],
    freeze: str = "A2",
) -> None:
    """把若干 DataFrame 依次写入一个 sheet（标题行可选）。"""
    ws = wb.create_sheet(sheet_name)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    row_idx = 1
    for title, frame in frames:
        if title:
            ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=12)
            row_idx += 1
        if frame is None:
            continue
        for col_idx, col_name in enumerate(frame.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        for _, record in frame.iterrows():
            for col_idx, col_name in enumerate(frame.columns, start=1):
                value = record[col_name]
                if pd.isna(value):
                    value = None
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, float):
                    col_key = str(col_name)
                    if "rate" in col_key or "pct" in col_key or "1m30p" in col_key or "3m30p" in col_key:
                        cell.number_format = "0.00%"
                    elif "auc" in col_key or "ks" in col_key or "lift" in col_key:
                        cell.number_format = "0.0000"
                    elif "corr" in col_key:
                        cell.number_format = "0.0000"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"
            row_idx += 1
        row_idx += 1
    ws.freeze_panes = freeze
    for col_idx in range(1, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


# ============================================================
# 3. matrix 模式：A 模型 × B 模型全局交叉
# ============================================================

# 现行单模型策略阈值（档位口径，来源：两模型分箱报告的最终方案；老客口径）。
# 新组合上线时需在 configs/models.py 的 comment 中记录并同步此处。
MLT_AUTO_BIN = 3
MLT_ACCEPT_BIN = 5
WTH_AUTO_BIN = 2
WTH_ACCEPT_BIN = 3
COMBO_MLT_WEIGHT = 0.7
DISAGREE_RANK_GAP = 3
MIN_CELL_N_FOR_RATE = 50


def build_cross_sample(
    final_a: pd.DataFrame,
    final_b: pd.DataFrame,
    model_a_cfg: dict,
    model_b_cfg: dict,
) -> pd.DataFrame:
    """按 application_id 内连接两模型最终分档，得到双分样本。"""
    tag_a = model_a_cfg["cross_tag"]
    tag_b = model_b_cfg["cross_tag"]
    cols_a = [
        "application_id", "sample_group", "application_month",
        "duedate_1m_30", "duedate_3m_30", "principal",
        "estimate_principal_remaining_mob1", "estimate_principal_remaining_mob3",
        "dpd_days_ever_mob1", "dpd_days_ever_mob3",
        "application_status", "assessment_status", "status",
        model_a_cfg["final_bin_col"], "bin_order", model_a_cfg["score_col"],
    ]
    cols_b = ["application_id", model_b_cfg["final_bin_col"], "bin_order", model_b_cfg["score_col"]]
    side_a = final_a[cols_a].rename(columns={
        "bin_order": f"{tag_a}_bin_order",
        model_a_cfg["final_bin_col"]: f"{tag_a}_bin",
        model_a_cfg["score_col"]: f"{tag_a}_score",
    })
    side_b = final_b[cols_b].rename(columns={
        "bin_order": f"{tag_b}_bin_order",
        model_b_cfg["final_bin_col"]: f"{tag_b}_bin",
        model_b_cfg["score_col"]: f"{tag_b}_score",
    })
    cross = side_a.merge(side_b, on="application_id", how="inner")
    for col in [f"{tag_a}_score", f"{tag_b}_score"]:
        cross[col] = pd.to_numeric(cross[col], errors="coerce")
    return cross


def add_bin_labels(edges: pd.DataFrame, bin_col: str) -> pd.DataFrame:
    result = edges.copy()
    result["final_bin_col_label"] = result[bin_col]
    return result


def _label(edges: pd.DataFrame, order: int) -> str:
    hit = edges.loc[edges["final_bin_order"].eq(order)]
    return str(hit.iloc[0]["final_bin_col_label"]) if not hit.empty else str(order)


def build_cross_matrix(
    cross: pd.DataFrame,
    sample_group: str,
    edges_a: pd.DataFrame,
    edges_b: pd.DataFrame,
    tag_a: str,
    tag_b: str,
) -> pd.DataFrame:
    """某样本组的 7×7 交叉矩阵（含行列边际与 13 组指标）。"""
    group = cross.loc[cross["sample_group"].eq(sample_group)].copy()
    overall = group_overall_rates(group)
    group_total = len(group)
    group_principal = float(pd.to_numeric(group["principal"], errors="coerce").fillna(0).sum())

    def cell_metrics(frame: pd.DataFrame) -> Dict[str, object]:
        n = len(frame)
        _, _, r1 = rate_of(frame, "duedate_1m_30")
        _, _, r3 = rate_of(frame, "duedate_3m_30")
        a1 = amt_rate_of(frame, "dpd_days_ever_mob1", "estimate_principal_remaining_mob1")
        a3 = amt_rate_of(frame, "dpd_days_ever_mob3", "estimate_principal_remaining_mob3")
        principal = float(pd.to_numeric(frame["principal"], errors="coerce").fillna(0).sum())
        funnel = _actual_funnel_row(frame, "cell")
        return {
            "n": n,
            "sample_pct": n / group_total,
            "1m30p_cnt_bad_rate": r1,
            "3m30p_cnt_bad_rate": r3,
            "1m30p_amt_bad_rate": a1,
            "3m30p_amt_bad_rate": a3,
            "1m30p_cnt_lift": r1 / overall["1m30p_cnt_bad_rate"] if overall["1m30p_cnt_bad_rate"] else np.nan,
            "3m30p_cnt_lift": r3 / overall["3m30p_cnt_bad_rate"] if overall["3m30p_cnt_bad_rate"] else np.nan,
            "1m30p_amt_lift": a1 / overall["1m30p_amt_bad_rate"] if overall["1m30p_amt_bad_rate"] else np.nan,
            "3m30p_amt_lift": a3 / overall["3m30p_amt_bad_rate"] if overall["3m30p_amt_bad_rate"] else np.nan,
            "principal_pct": principal / group_principal if group_principal else np.nan,
            "actual_approval_rate": funnel["actual_approval_rate"],
            "actual_deal_rate": funnel["actual_deal_rate"],
        }

    orders = sorted(
        set(group[f"{tag_a}_bin_order"].dropna().astype(int))
        & set(group[f"{tag_b}_bin_order"].dropna().astype(int))
    )
    rows: List[Dict] = []
    for mo, wo in product(orders, orders):
        cell = group.loc[group[f"{tag_a}_bin_order"].eq(mo) & group[f"{tag_b}_bin_order"].eq(wo)]
        rows.append({
            f"{tag_a}_bin_order": mo,
            f"{tag_a}_bin": _label(edges_a, mo),
            f"{tag_b}_bin_order": wo,
            f"{tag_b}_bin": _label(edges_b, wo),
            **cell_metrics(cell),
        })
    for mo in orders:
        row_cell = group.loc[group[f"{tag_a}_bin_order"].eq(mo)]
        rows.append({
            f"{tag_a}_bin_order": mo,
            f"{tag_a}_bin": _label(edges_a, mo),
            f"{tag_b}_bin_order": 0,
            f"{tag_b}_bin": "行边际",
            **cell_metrics(row_cell),
        })
    for wo in orders:
        col_cell = group.loc[group[f"{tag_b}_bin_order"].eq(wo)]
        rows.append({
            f"{tag_a}_bin_order": 0,
            f"{tag_a}_bin": "列边际",
            f"{tag_b}_bin_order": wo,
            f"{tag_b}_bin": _label(edges_b, wo),
            **cell_metrics(col_cell),
        })
    overall_metrics = cell_metrics(group)
    overall_metrics.update({
        "sample_pct": 1.0,
        "1m30p_cnt_lift": 1.0, "3m30p_cnt_lift": 1.0,
        "1m30p_amt_lift": 1.0, "3m30p_amt_lift": 1.0,
        "principal_pct": 1.0,
    })
    rows.append({
        f"{tag_a}_bin_order": 0,
        f"{tag_a}_bin": "整体",
        f"{tag_b}_bin_order": 0,
        f"{tag_b}_bin": "整体",
        **overall_metrics,
    })
    return pd.DataFrame(rows)


def build_conditional_analysis(cross: pd.DataFrame, tag_a: str, tag_b: str) -> pd.DataFrame:
    """每个 A 模型档位内 B 模型档位的风险跨度，与强分歧格汇总。"""
    group = cross.loc[cross["sample_group"].eq("train")].copy()
    group_total = len(group)
    rows: List[Dict] = []

    for mo in sorted(group[f"{tag_a}_bin_order"].dropna().unique().astype(int)):
        row_cell = group.loc[group[f"{tag_a}_bin_order"].eq(mo)]
        _, _, row_rate = rate_of(row_cell, "duedate_3m_30")
        col_rates = {}
        for wo in sorted(group[f"{tag_b}_bin_order"].dropna().unique().astype(int)):
            cell = row_cell.loc[row_cell[f"{tag_b}_bin_order"].eq(wo)]
            if len(cell) >= MIN_CELL_N_FOR_RATE:
                _, _, r3 = rate_of(cell, "duedate_3m_30")
                col_rates[wo] = r3
        vals = sorted(col_rates.values())
        rows.append({
            "dimension": f"{tag_a} 行内（{tag_b} 增量）",
            "anchor_bin_order": mo,
            "anchor_bin": chr(ord("A") + mo - 1),
            "anchor_n": len(row_cell),
            "anchor_3m30p_cnt_bad_rate": row_rate,
            "other_min_rate": vals[0] if vals else np.nan,
            "other_max_rate": vals[-1] if vals else np.nan,
            "other_spread": (vals[-1] - vals[0]) if vals else np.nan,
        })

    for wo in sorted(group[f"{tag_b}_bin_order"].dropna().unique().astype(int)):
        col_cell = group.loc[group[f"{tag_b}_bin_order"].eq(wo)]
        _, _, col_rate = rate_of(col_cell, "duedate_3m_30")
        row_rates = {}
        for mo in sorted(group[f"{tag_a}_bin_order"].dropna().unique().astype(int)):
            cell = col_cell.loc[col_cell[f"{tag_a}_bin_order"].eq(mo)]
            if len(cell) >= MIN_CELL_N_FOR_RATE:
                _, _, r3 = rate_of(cell, "duedate_3m_30")
                row_rates[mo] = r3
        vals = sorted(row_rates.values())
        rows.append({
            "dimension": f"{tag_b} 列内（{tag_a} 增量）",
            "anchor_bin_order": wo,
            "anchor_bin": chr(ord("A") + wo - 1),
            "anchor_n": len(col_cell),
            "anchor_3m30p_cnt_bad_rate": col_rate,
            "other_min_rate": vals[0] if vals else np.nan,
            "other_max_rate": vals[-1] if vals else np.nan,
            "other_spread": (vals[-1] - vals[0]) if vals else np.nan,
        })

    rank_gap = (group[f"{tag_a}_bin_order"].astype(int) - group[f"{tag_b}_bin_order"].astype(int)).abs()
    disagree = group.loc[rank_gap.ge(DISAGREE_RANK_GAP)].copy()
    agree = group.loc[rank_gap.le(1)].copy()
    for label, frame in [("强分歧（档位差 ≥ 3）", disagree), ("近似一致（档位差 ≤ 1）", agree)]:
        _, _, r3 = rate_of(frame, "duedate_3m_30")
        _, _, r1 = rate_of(frame, "duedate_1m_30")
        rows.append({
            "dimension": "分档一致性",
            "anchor_bin_order": np.nan,
            "anchor_bin": label,
            "anchor_n": len(frame),
            "anchor_sample_pct": len(frame) / group_total,
            "anchor_3m30p_cnt_bad_rate": r3,
            "other_1m30p_cnt_bad_rate": r1,
            "other_min_rate": np.nan,
            "other_max_rate": np.nan,
            "other_spread": np.nan,
        })
    for label, mask in [
        (f"{tag_a} 低风险 & {tag_b} 高风险（{tag_a} ≤ 3 且 {tag_b} ≥ 5）",
         group[f"{tag_a}_bin_order"].le(3) & group[f"{tag_b}_bin_order"].ge(5)),
        (f"{tag_b} 低风险 & {tag_a} 高风险（{tag_b} ≤ 3 且 {tag_a} ≥ 5）",
         group[f"{tag_b}_bin_order"].le(3) & group[f"{tag_a}_bin_order"].ge(5)),
    ]:
        frame = group.loc[mask]
        _, _, r3 = rate_of(frame, "duedate_3m_30")
        _, _, r1 = rate_of(frame, "duedate_1m_30")
        rows.append({
            "dimension": "分歧方向",
            "anchor_bin_order": np.nan,
            "anchor_bin": label,
            "anchor_n": len(frame),
            "anchor_sample_pct": len(frame) / group_total,
            "anchor_3m30p_cnt_bad_rate": r3,
            "other_1m30p_cnt_bad_rate": r1,
            "other_min_rate": np.nan,
            "other_max_rate": np.nan,
            "other_spread": np.nan,
        })
    return pd.DataFrame(rows)


def build_combined_scores(cross: pd.DataFrame, tag_a: str, tag_b: str) -> pd.DataFrame:
    """z-score 标准化两模型分并构造组合分。"""
    result = cross.copy()
    train_mask = result["sample_group"].eq("train")
    for col in [f"{tag_a}_score", f"{tag_b}_score"]:
        mean = result.loc[train_mask, col].mean()
        std = result.loc[train_mask, col].std()
        result[f"z_{col}"] = (result[col] - mean) / std
    result["combo_avg"] = (result[f"z_{tag_a}_score"] + result[f"z_{tag_b}_score"]) / 2
    result["combo_w7030"] = (
        COMBO_MLT_WEIGHT * result[f"z_{tag_a}_score"]
        + (1 - COMBO_MLT_WEIGHT) * result[f"z_{tag_b}_score"]
    )
    result["combo_rank_avg"] = (
        result[f"{tag_a}_bin_order"].astype(float) + result[f"{tag_b}_bin_order"].astype(float)
    ) / 2
    result["combo_rank_max"] = result[[f"{tag_a}_bin_order", f"{tag_b}_bin_order"]].max(axis=1).astype(float)
    return result


def build_score_performance(cross: pd.DataFrame, tag_a: str, tag_b: str, model_a_cfg: dict, model_b_cfg: dict) -> pd.DataFrame:
    """单模型分与组合分的 Train/OOT AUC / KS 对比。"""
    score_cols = [
        (f"{model_a_cfg['display_short']} 单模型分", f"{tag_a}_score"),
        (f"{model_b_cfg['display_short']} 单模型分", f"{tag_b}_score"),
        ("组合分：z 平均", "combo_avg"),
        ("组合分：7:3 加权", "combo_w7030"),
        ("组合分：档位平均", "combo_rank_avg"),
        ("组合分：档位取大", "combo_rank_max"),
    ]
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        for score_name, score_col in score_cols:
            for label_name, label_col in [("1M30+", "duedate_1m_30"), ("3M30+", "duedate_3m_30")]:
                mask = frame[label_col].isin([0, 1]) & frame[score_col].notna()
                n = int(mask.sum())
                bad = int(frame.loc[mask, label_col].eq(1).sum())
                perf = calc_auc_ks(frame, score_col, label_col)
                auc, ks = perf["auc"], perf["ks"]
                rows.append({
                    "sample_group": group_name,
                    "score": score_name,
                    "label": label_name,
                    "n": n,
                    "bad_cnt": bad,
                    "bad_rate": bad / n if n else np.nan,
                    "auc": auc,
                    "ks": ks,
                })
    return pd.DataFrame(rows)


def segment_mask(cross, tag_a, tag_b, a_auto, b_auto, a_accept, b_accept, logic):
    a_o = cross[f"{tag_a}_bin_order"].astype(int)
    b_o = cross[f"{tag_b}_bin_order"].astype(int)
    combine = np.logical_and if logic == "AND" else np.logical_or
    auto = combine(a_o.le(a_auto), b_o.le(b_auto))
    accept = combine(a_o.le(a_accept), b_o.le(b_accept))
    return {"auto": auto, "accept": accept, "reject": ~accept, "manual": accept & ~auto}


def simulate_policy(cross, tag_a, tag_b, name, logic, a_auto, b_auto, a_accept, b_accept):
    masks = segment_mask(cross, tag_a, tag_b, a_auto, b_auto, a_accept, b_accept, logic)
    row = {
        "policy": name,
        "logic": logic,
        f"{tag_a}_auto_bin": a_auto,
        f"{tag_b}_auto_bin": b_auto,
        f"{tag_a}_accept_bin": a_accept,
        f"{tag_b}_accept_bin": b_accept,
    }
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        total = len(frame)
        for seg in ["auto", "manual", "accept", "reject"]:
            seg_frame = frame.loc[masks[seg]]
            _, _, r1 = rate_of(seg_frame, "duedate_1m_30")
            _, _, r3 = rate_of(seg_frame, "duedate_3m_30")
            row[f"{group_name}_{seg}_rate"] = len(seg_frame) / total
            row[f"{group_name}_{seg}_1m30p"] = r1
            row[f"{group_name}_{seg}_3m30p"] = r3
    return row


def build_policy_table(cross, tag_a, tag_b):
    """两模型单模型现行策略 + 二维 AND / OR 组合策略对照。"""
    auto_a, accept_a, auto_b, accept_b = _current_thresholds(tag_a, tag_b)
    rows: List[Dict] = []
    rows.append(simulate_policy(cross, tag_a, tag_b, f"{tag_a} 单模型（现行）", "AND", auto_a, 7, accept_a, 7))
    rows.append(simulate_policy(cross, tag_a, tag_b, f"{tag_b} 单模型（现行）", "AND", 7, auto_b, 7, accept_b))
    rows.append(simulate_policy(cross, tag_a, tag_b, "二维组合（AND）", "AND", auto_a, auto_b, accept_a, accept_b))
    rows.append(simulate_policy(cross, tag_a, tag_b, "二维组合（OR）", "OR", auto_a, auto_b, accept_a, accept_b))
    return pd.DataFrame(rows)


def _current_thresholds(tag_a: str, tag_b: str):
    """现行单模型阈值的档位口径（老客 mlt × worthiness 的评审值；新组合需在此登记）。"""
    if (tag_a, tag_b) == ("mlt", "wth"):
        return MLT_AUTO_BIN, MLT_ACCEPT_BIN, WTH_AUTO_BIN, WTH_ACCEPT_BIN
    if (tag_a, tag_b) == ("wth", "mlt"):
        return WTH_AUTO_BIN, WTH_ACCEPT_BIN, MLT_AUTO_BIN, MLT_ACCEPT_BIN
    raise ValueError(f"未登记 {tag_a} × {tag_b} 的现行阈值档位，请在 pipeline/cross_analysis.py 补充")


def build_accept_grid(cross, tag_a, tag_b):
    """AND 逻辑下（A 档上限 × B 档上限）接纳网格：Train 接纳率与 3M30+ 风险。"""
    _, accept_a, _, accept_b = _current_thresholds(tag_a, tag_b)
    rows: List[Dict] = []
    for a_cut in range(2, 8):
        for b_cut in range(2, 8):
            row = simulate_policy(cross, tag_a, tag_b, "grid", "AND", a_cut, b_cut, a_cut, b_cut)
            rows.append({
                f"{tag_a}_accept_bin": a_cut,
                f"{tag_b}_accept_bin": b_cut,
                "train_accept_rate": row["train_accept_rate"],
                "train_accept_3m30p": row["train_accept_3m30p"],
                "oot_accept_rate": row["oot_accept_rate"],
                "oot_accept_3m30p": row["oot_accept_3m30p"],
                "is_current_and": a_cut == accept_a and b_cut == accept_b,
            })
    return pd.DataFrame(rows)


def build_quadrant_table(cross, tag_a, tag_b):
    """按现行接纳阈值划分四个象限人群。"""
    _, accept_a, _, accept_b = _current_thresholds(tag_a, tag_b)
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        total = len(frame)
        quadrants = [
            (f"双低（{tag_a} ≤ {chr(64+accept_a)} 且 {tag_b} ≤ {chr(64+accept_b)}）",
             frame[f"{tag_a}_bin_order"].le(accept_a) & frame[f"{tag_b}_bin_order"].le(accept_b)),
            (f"仅 {tag_a} 低（{tag_a} ≤ {chr(64+accept_a)} 且 {tag_b} > {chr(64+accept_b)}）",
             frame[f"{tag_a}_bin_order"].le(accept_a) & frame[f"{tag_b}_bin_order"].gt(accept_b)),
            (f"仅 {tag_b} 低（{tag_b} ≤ {chr(64+accept_b)} 且 {tag_a} > {chr(64+accept_a)}）",
             frame[f"{tag_b}_bin_order"].le(accept_b) & frame[f"{tag_a}_bin_order"].gt(accept_a)),
            (f"双高（{tag_a} > {chr(64+accept_a)} 且 {tag_b} > {chr(64+accept_b)}）",
             frame[f"{tag_a}_bin_order"].gt(accept_a) & frame[f"{tag_b}_bin_order"].gt(accept_b)),
        ]
        for label, mask in quadrants:
            seg = frame.loc[mask]
            _, _, r1 = rate_of(seg, "duedate_1m_30")
            _, _, r3 = rate_of(seg, "duedate_3m_30")
            rows.append({
                "sample_group": group_name,
                "quadrant": label,
                "n": len(seg),
                "sample_pct": len(seg) / total,
                "1m30p_cnt_bad_rate": r1,
                "3m30p_cnt_bad_rate": r3,
            })
    return pd.DataFrame(rows)


def run_matrix(
    dataset_cfg: dict,
    model_a_cfg: dict,
    model_b_cfg: dict,
    report_path: Path,
) -> None:
    """matrix 模式主流程：输出 7 个 sheet 的交叉分析 Excel。"""
    t0 = time.time()
    tag_a, tag_b = model_a_cfg["cross_tag"], model_b_cfg["cross_tag"]
    print("=" * 60)
    print(f"两模型交叉分析（matrix）：{model_a_cfg['name']} × {model_b_cfg['name']}")
    print("=" * 60)

    final_a, edges_a, ranges_a = build_model_final_bins(dataset_cfg, model_a_cfg)
    final_b, edges_b, ranges_b = build_model_final_bins(dataset_cfg, model_b_cfg)
    edges_a = add_bin_labels(edges_a, model_a_cfg["final_bin_col"])
    edges_b = add_bin_labels(edges_b, model_b_cfg["final_bin_col"])

    cross = build_cross_sample(final_a, final_b, model_a_cfg, model_b_cfg)
    train_n = int(cross["sample_group"].eq("train").sum())
    oot_n = int(cross["sample_group"].eq("oot").sum())
    print(f"[1/7] 双分样本：{len(cross):,} 笔（Train {train_n:,} / OOT {oot_n:,}）")

    train_cross = cross.loc[cross["sample_group"].eq("train")]
    pearson = float(train_cross[f"{tag_a}_score"].corr(train_cross[f"{tag_b}_score"]))
    spearman = float(train_cross[f"{tag_a}_score"].corr(train_cross[f"{tag_b}_score"], method="spearman"))
    bin_spearman = float(
        train_cross[f"{tag_a}_bin_order"].astype(int).corr(
            train_cross[f"{tag_b}_bin_order"].astype(int), method="spearman"
        )
    )
    print(f"[2/7] 相关性：Pearson {pearson:.4f} / Spearman {spearman:.4f} / 分档秩相关 {bin_spearman:.4f}")

    train_matrix = build_cross_matrix(cross, "train", edges_a, edges_b, tag_a, tag_b)
    oot_matrix = build_cross_matrix(cross, "oot", edges_a, edges_b, tag_a, tag_b)
    print(f"[3/7] 交叉矩阵完成（Train {len(train_matrix)} 行 / OOT {len(oot_matrix)} 行）")

    conditional = build_conditional_analysis(cross, tag_a, tag_b)
    print(f"[4/7] 条件增量分析完成（{len(conditional)} 行）")

    cross = build_combined_scores(cross, tag_a, tag_b)
    score_perf = build_score_performance(cross, tag_a, tag_b, model_a_cfg, model_b_cfg)
    print("[5/7] 组合评分效果完成")

    policy_table = build_policy_table(cross, tag_a, tag_b)
    accept_grid = build_accept_grid(cross, tag_a, tag_b)
    quadrant_table = build_quadrant_table(cross, tag_a, tag_b)
    print("[6/7] 二维策略模拟完成")

    overview = pd.DataFrame([
        {"section": "样本", "metric": "双分样本量", "value": len(cross)},
        {"section": "样本", "metric": "Train / OOT 样本量", "value": f"{train_n:,} / {oot_n:,}"},
        {"section": "相关性", "metric": "模型分 Pearson 相关（Train）", "value": pearson},
        {"section": "相关性", "metric": "模型分 Spearman 相关（Train）", "value": spearman},
        {"section": "相关性", "metric": "最终分档秩相关（Train）", "value": bin_spearman},
        {"section": "分档方案", "metric": f"{tag_a} 最终方案", "value": binning_cnt.format_merge_ranges(ranges_a)},
        {"section": "分档方案", "metric": f"{tag_b} 最终方案", "value": binning_cnt.format_merge_ranges(ranges_b)},
    ])
    appendix = pd.DataFrame([
        {"config_group": "基础配置", "config_name": "COMBO_MLT_WEIGHT", "config_value": COMBO_MLT_WEIGHT},
        {"config_group": "基础配置", "config_name": "DISAGREE_RANK_GAP", "config_value": DISAGREE_RANK_GAP},
        {"config_group": "基础配置", "config_name": "MIN_CELL_N_FOR_RATE", "config_value": MIN_CELL_N_FOR_RATE},
    ])
    print("[7/7] 总览与附录完成")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "01_总览", [("两模型交叉分析总览", None), (None, overview)])
    write_sheet(wb, "02_交叉矩阵_Train", [(f"Train 7×7 交叉矩阵（行 = {tag_a} 档，列 = {tag_b} 档）", None), (None, train_matrix)])
    write_sheet(wb, "03_交叉矩阵_OOT", [(f"OOT 7×7 交叉矩阵（行 = {tag_a} 档，列 = {tag_b} 档）", None), (None, oot_matrix)])
    write_sheet(wb, "04_条件增量分析", [("条件增量与分档一致性", None), (None, conditional)])
    write_sheet(wb, "05_组合评分效果", [("单模型分与组合分 AUC / KS 对比", None), (None, score_perf)])
    write_sheet(wb, "06_二维策略模拟", [
        ("策略对照（单模型现行 vs 二维 AND / OR）", None),
        (None, policy_table),
        ("AND 接纳网格", None),
        (None, accept_grid),
        ("现行阈值四象限人群", None),
        (None, quadrant_table),
    ])
    write_sheet(wb, "07_附录", [("配置参数", None), (None, appendix)])
    wb.save(report_path)
    print(f"完成 => {report_path}（耗时 {time.time() - t0:.1f}s）")


# ============================================================
# 4. cond 模式：A 模型各档内的 B 模型条件子箱
# ============================================================

SUB_BIN_COUNT = 3
ANCHOR_DUE_COL = "duedate_3m_30"
ANCHOR_DPD_COL = "dpd_days_ever_mob3"
ANCHOR_REMAINING_COL = "estimate_principal_remaining_mob3"


def learn_sub_bin_edges(train_cross: pd.DataFrame, tag_a: str, tag_b: str) -> Dict[int, np.ndarray]:
    """在每个 A 档内（Train）对 B 模型分做等频切分，返回 {A 档位: 边界数组}。"""
    tier_edges: Dict[int, np.ndarray] = {}
    for tier in range(1, 8):
        frame = train_cross.loc[train_cross[f"{tag_a}_bin_order"].eq(tier)]
        score = pd.to_numeric(frame[f"{tag_b}_score"], errors="coerce").dropna()
        _, raw_edges = pd.qcut(score, q=SUB_BIN_COUNT, retbins=True, duplicates="drop")
        edges = np.unique(np.asarray(raw_edges, dtype="float64"))
        if len(edges) < 2:
            raise ValueError(f"{tag_a} {chr(64 + tier)} 档内 {tag_b} 分唯一值不足，无法切分子箱")
        edges[0] = -np.inf
        edges[-1] = np.inf
        tier_edges[tier] = edges
    return tier_edges


def apply_sub_bins(cross: pd.DataFrame, tier_edges: Dict[int, np.ndarray], tag_a: str, tag_b: str) -> pd.DataFrame:
    """把各档内子箱边界应用到样本，生成 wth_sub_bin 与组合标签/序。"""
    result = cross.copy()
    sub_parts: List[pd.Series] = []
    for tier in range(1, 8):
        mask = result[f"{tag_a}_bin_order"].eq(tier)
        sub = pd.cut(
            result.loc[mask, f"{tag_b}_score"],
            bins=tier_edges[tier],
            labels=list(range(1, len(tier_edges[tier]))),
            include_lowest=True,
        )
        sub_parts.append(sub)
    result[f"{tag_b}_sub_bin"] = pd.concat(sub_parts).astype("Int64")
    result["combined_label"] = (
        result[f"{tag_a}_bin"] + result[f"{tag_b}_sub_bin"].astype("Int64").astype("string")
    )
    result["combined_order"] = (
        (result[f"{tag_a}_bin_order"].astype(int) - 1) * SUB_BIN_COUNT
        + result[f"{tag_b}_sub_bin"]
    ).astype(int)
    return result


def build_sub_bin_edge_table(tier_edges: Dict[int, np.ndarray]) -> pd.DataFrame:
    rows = []
    for tier in range(1, 8):
        edges = tier_edges[tier]
        for idx in range(len(edges) - 1):
            rows.append({
                "mlt_bin_order": tier,
                "mlt_bin": chr(64 + tier),
                "wth_sub_bin": idx + 1,
                "combined_label": f"{chr(64 + tier)}{idx + 1}",
                "score_left": edges[idx],
                "score_right": edges[idx + 1],
            })
    return pd.DataFrame(rows)


def build_combined_stats(cross: pd.DataFrame, sample_group: str, tag_a: str, tag_b: str) -> pd.DataFrame:
    """某样本组的 21 个组合格统计（含 95% Wilson CI 与 Lift）。"""
    group = cross.loc[cross["sample_group"].eq(sample_group)].copy()
    overall = group_overall_rates(group)
    total = len(group)
    group_principal = float(pd.to_numeric(group["principal"], errors="coerce").fillna(0).sum())
    rows: List[Dict] = []
    for tier in range(1, 8):
        for sub in range(1, SUB_BIN_COUNT + 1):
            cell = group.loc[group[f"{tag_a}_bin_order"].eq(tier) & group[f"{tag_b}_sub_bin"].eq(sub)]
            if cell.empty:
                continue
            n = len(cell)
            m3, b3, r3 = rate_of(cell, ANCHOR_DUE_COL)
            m1, b1, r1 = rate_of(cell, "duedate_1m_30")
            ci3_low, ci3_high = wilson_ci(np.array([b3]), np.array([m3]))
            ci1_low, ci1_high = wilson_ci(np.array([b1]), np.array([m1]))
            a3 = amt_rate_of(cell, ANCHOR_DPD_COL, ANCHOR_REMAINING_COL)
            principal = float(pd.to_numeric(cell["principal"], errors="coerce").fillna(0).sum())
            funnel = _actual_funnel_row(cell, "cell")
            rows.append({
                "combined_order": (tier - 1) * SUB_BIN_COUNT + sub,
                "combined_label": f"{chr(64 + tier)}{sub}",
                "mlt_bin_order": tier,
                "mlt_bin": chr(64 + tier),
                "wth_sub_bin": sub,
                "n": n,
                "sample_pct": n / total,
                "1m30p_cnt_bad_rate": r1,
                "1m30p_cnt_bad_rate_ci_low": float(ci1_low[0]),
                "1m30p_cnt_bad_rate_ci_high": float(ci1_high[0]),
                "3m30p_cnt_bad_rate": r3,
                "3m30p_cnt_bad_rate_ci_low": float(ci3_low[0]),
                "3m30p_cnt_bad_rate_ci_high": float(ci3_high[0]),
                "3m30p_amt_bad_rate": a3,
                "3m30p_cnt_lift": r3 / overall["3m30p_cnt_bad_rate"] if overall["3m30p_cnt_bad_rate"] else np.nan,
                "principal_pct": principal / group_principal if group_principal else np.nan,
                "actual_approval_rate": funnel["actual_approval_rate"],
                "actual_deal_rate": funnel["actual_deal_rate"],
                "3m30p_cnt_mature": m3,
                "3m30p_cnt_bad": b3,
                "3m30p_cnt_good": m3 - b3,
            })
    return pd.DataFrame(rows)


def build_tier_sub_bin_summary(train_stats: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict] = []
    for tier in range(1, 8):
        frame = train_stats.loc[train_stats["mlt_bin_order"].eq(tier)].sort_values("wth_sub_bin")
        tier_cells = frame[["combined_label", "n", "3m30p_cnt_bad", "3m30p_cnt_mature", "3m30p_cnt_bad_rate"]].copy()
        rates = [row["3m30p_cnt_bad_rate"] for _, row in tier_cells.iterrows() if row["n"] >= MIN_CELL_N_FOR_RATE]
        spread = (max(rates) - min(rates)) if len(rates) >= 2 else np.nan
        p_values = []
        for k in range(1, SUB_BIN_COUNT):
            left = tier_cells.loc[tier_cells["combined_label"].eq(f"{chr(64 + tier)}{k}")]
            right = tier_cells.loc[tier_cells["combined_label"].eq(f"{chr(64 + tier)}{k + 1}")]
            if left.empty or right.empty:
                p_values.append(np.nan)
                continue
            p_values.append(two_proportion_pvalue(
                float(left.iloc[0]["3m30p_cnt_bad"]), float(left.iloc[0]["3m30p_cnt_mature"]),
                float(right.iloc[0]["3m30p_cnt_bad"]), float(right.iloc[0]["3m30p_cnt_mature"]),
            ))
        row = {"mlt_bin": chr(64 + tier), "mlt_bin_order": tier}
        for k in range(1, SUB_BIN_COUNT + 1):
            hit = tier_cells.loc[tier_cells["combined_label"].eq(f"{chr(64 + tier)}{k}")]
            if hit.empty:
                row[f"sub{k}_3m30p"] = np.nan
                row[f"sub{k}_n"] = 0
            else:
                row[f"sub{k}_3m30p"] = hit.iloc[0]["3m30p_cnt_bad_rate"]
                row[f"sub{k}_n"] = int(hit.iloc[0]["n"])
        row["within_tier_spread"] = spread
        for idx in range(SUB_BIN_COUNT - 1):
            row[f"p_sub{idx + 1}_vs_sub{idx + 2}"] = p_values[idx]
        rows.append(row)
    return pd.DataFrame(rows)


def build_head_tail_eval(train_stats, oot_stats, cross, tag_a) -> pd.DataFrame:
    rows: List[Dict] = []
    for group_key, display_name, stats in [("train", "Train", train_stats), ("oot", "OOT", oot_stats)]:
        group = cross.loc[cross["sample_group"].eq(group_key)]
        for tier_label, tier in [("头部", 1), ("尾部", 7)]:
            tier_frame = group.loc[group[f"{tag_a}_bin_order"].eq(tier)]
            _, _, tier_rate = rate_of(tier_frame, ANCHOR_DUE_COL)
            best = stats.loc[stats["combined_label"].eq(f"{chr(64 + tier)}1")]
            worst = stats.loc[stats["combined_label"].eq(f"{chr(64 + tier)}{SUB_BIN_COUNT}")]
            rows.append({
                "sample_group": display_name,
                "position": tier_label,
                "mlt_tier": chr(64 + tier),
                "mlt_tier_n": len(tier_frame),
                "mlt_tier_3m30p": tier_rate,
                "best_sub_label": f"{chr(64 + tier)}1",
                "best_sub_3m30p": float(best.iloc[0]["3m30p_cnt_bad_rate"]) if not best.empty else np.nan,
                "best_sub_lift": float(best.iloc[0]["3m30p_cnt_lift"]) if not best.empty else np.nan,
                "worst_sub_label": f"{chr(64 + tier)}{SUB_BIN_COUNT}",
                "worst_sub_3m30p": float(worst.iloc[0]["3m30p_cnt_bad_rate"]) if not worst.empty else np.nan,
                "worst_sub_lift": float(worst.iloc[0]["3m30p_cnt_lift"]) if not worst.empty else np.nan,
                "sub_spread": (
                    float(worst.iloc[0]["3m30p_cnt_bad_rate"]) - float(best.iloc[0]["3m30p_cnt_bad_rate"])
                    if not best.empty and not worst.empty else np.nan
                ),
            })
    return pd.DataFrame(rows)


def build_discrimination_comparison(cross: pd.DataFrame, tag_a: str, tag_b: str, model_a_cfg: dict, model_b_cfg: dict) -> pd.DataFrame:
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        group = cross.loc[cross["sample_group"].eq(group_name)]
        for scheme, key in [
            (f"{model_a_cfg['display_short']} 7 档", f"{tag_a}_bin_order"),
            (f"{model_b_cfg['display_short']} 7 档", f"{tag_b}_bin_order"),
            ("组合 21 格", "combined_order"),
        ]:
            for label_name, label_col in [("1M30+", "duedate_1m_30"), ("3M30+", "duedate_3m_30")]:
                frame = group.loc[group[label_col].isin([0, 1]) & group[key].notna()]
                bad = frame[label_col].eq(1).sum()
                mature = len(frame)
                perf = calc_auc_ks(group, key, label_col)
                auc, ks = perf["auc"], perf["ks"]
                rows.append({
                    "sample_group": group_name,
                    "scheme": scheme,
                    "label": label_name,
                    "auc": auc,
                    "ks": ks,
                    "mature": int(mature),
                    "bad": int(bad),
                })
    return pd.DataFrame(rows)


def calc_scheme_iv(cross, key, group_name="train"):
    group = cross.loc[cross["sample_group"].eq(group_name)]
    parts = []
    for value in sorted(group[key].dropna().unique().astype(int)):
        frame = group.loc[group[key].eq(value)]
        _, bad, _ = rate_of(frame, ANCHOR_DUE_COL)
        mature = int(pd.to_numeric(frame[ANCHOR_DUE_COL], errors="coerce").isin([0, 1]).sum())
        parts.append({"3m30p_cnt_bad": bad, "3m30p_cnt_good": mature - bad})
    stats = pd.DataFrame(parts)
    return risk_metrics.calc_iv_from_stats(stats)


def build_iv_comparison(cross, tag_a, tag_b, model_a_cfg, model_b_cfg):
    rows = []
    for scheme, key in [
        (f"{model_a_cfg['display_short']} 7 档", f"{tag_a}_bin_order"),
        (f"{model_b_cfg['display_short']} 7 档", f"{tag_b}_bin_order"),
        ("组合 21 格", "combined_order"),
    ]:
        rows.append({
            "scheme": scheme,
            "bin_count": int(cross.loc[cross["sample_group"].eq("train"), key].nunique()),
            "train_3m30p_iv": calc_scheme_iv(cross, key, "train"),
            "oot_3m30p_iv": calc_scheme_iv(cross, key, "oot"),
        })
    return pd.DataFrame(rows)


def check_combined_monotonicity(stats: pd.DataFrame, sample_group: str) -> pd.DataFrame:
    ordered = stats.sort_values("combined_order").reset_index(drop=True)
    rows = []
    for rate_col in ["1m30p_cnt_bad_rate", "3m30p_cnt_bad_rate"]:
        rates = ordered[rate_col].astype(float)
        inversions = int((rates.diff() < 0).sum())
        drop = float(rates.diff().min()) if inversions else 0.0
        rows.append({"sample_group": sample_group, "rate_col": rate_col, "inversion_cnt": inversions, "max_drop": drop})
    return pd.DataFrame(rows)


def calc_combined_psi(train_stats: pd.DataFrame, oot_stats: pd.DataFrame) -> float:
    merged = train_stats[["combined_label", "sample_pct"]].merge(
        oot_stats[["combined_label", "sample_pct"]],
        on="combined_label", how="outer", suffixes=("_train", "_oot"),
    ).fillna(0.0)
    eps = 1e-6
    return float(np.sum(
        (merged["sample_pct_oot"] - merged["sample_pct_train"])
        * np.log((merged["sample_pct_oot"] + eps) / (merged["sample_pct_train"] + eps))
    ))


def build_segment_relayer_table(cross, tag_a, tag_b) -> pd.DataFrame:
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        group = cross.loc[cross["sample_group"].eq(group_name)]
        for seg_label, mask in [
            (f"自动通过（{tag_a} ≤ 3）", group[f"{tag_a}_bin_order"].le(3)),
            (f"人工审核（3 < {tag_a} ≤ 5）", group[f"{tag_a}_bin_order"].gt(3) & group[f"{tag_a}_bin_order"].le(5)),
            (f"拒绝（{tag_a} > 5）", group[f"{tag_a}_bin_order"].gt(5)),
        ]:
            seg = group.loc[mask]
            best = seg.loc[seg[f"{tag_b}_sub_bin"].eq(1)]
            worst = seg.loc[seg[f"{tag_b}_sub_bin"].eq(SUB_BIN_COUNT)]
            _, _, seg_rate = rate_of(seg, ANCHOR_DUE_COL)
            _, _, best_rate = rate_of(best, ANCHOR_DUE_COL)
            _, _, worst_rate = rate_of(worst, ANCHOR_DUE_COL)
            rows.append({
                "sample_group": group_name,
                "segment": seg_label,
                "segment_n": len(seg),
                "segment_3m30p": seg_rate,
                "sub1_3m30p": best_rate,
                "sub3_3m30p": worst_rate,
                "sub_spread": (worst_rate - best_rate)
                if worst_rate == worst_rate and best_rate == best_rate else np.nan,
            })
    return pd.DataFrame(rows)


def build_sub_bin_strategy_table(cross, tag_a, tag_b) -> pd.DataFrame:
    rows: List[Dict] = []
    policies = [
        (f"{tag_a} 单模型（现行）", 3, SUB_BIN_COUNT, 5, SUB_BIN_COUNT),
        ("接纳收子箱 ≤2", 3, SUB_BIN_COUNT, 5, 2),
        ("自动通过收子箱 ≤2", 3, 2, 5, SUB_BIN_COUNT),
        ("两端均收子箱 ≤2", 3, 2, 5, 2),
        ("两端均收子箱 ≤1", 3, 1, 5, 1),
    ]
    for name, auto_mlt, auto_sub, acc_mlt, acc_sub in policies:
        row = {"policy": name, "mlt_auto_bin": auto_mlt, "sub_auto_bin": auto_sub,
               "mlt_accept_bin": acc_mlt, "sub_accept_bin": acc_sub}
        for group_key, prefix in [("train", "train"), ("oot", "oot")]:
            g = cross.loc[cross["sample_group"].eq(group_key)]
            total = len(g)
            auto = g[f"{tag_a}_bin_order"].le(auto_mlt) & g[f"{tag_b}_sub_bin"].le(auto_sub)
            accept = g[f"{tag_a}_bin_order"].le(acc_mlt) & g[f"{tag_b}_sub_bin"].le(acc_sub)
            for seg_label, mask in [("auto", auto), ("manual", accept & ~auto), ("accept", accept), ("reject", ~accept)]:
                seg = g.loc[mask]
                _, _, r3 = rate_of(seg, ANCHOR_DUE_COL)
                row[f"{prefix}_{seg_label}_rate"] = len(seg) / total
                row[f"{prefix}_{seg_label}_3m30p"] = r3
        rows.append(row)
    return pd.DataFrame(rows)


def build_sub_bin_strategy_grid(cross, tag_a, tag_b) -> pd.DataFrame:
    rows: List[Dict] = []
    for mlt_cut in range(1, 8):
        for sub_cut in range(1, SUB_BIN_COUNT + 1):
            row = {"mlt_accept_bin": mlt_cut, "sub_accept_bin": sub_cut}
            for group_key, prefix in [("train", "train"), ("oot", "oot")]:
                g = cross.loc[cross["sample_group"].eq(group_key)]
                total = len(g)
                accept = g[f"{tag_a}_bin_order"].le(mlt_cut) & g[f"{tag_b}_sub_bin"].le(sub_cut)
                seg = g.loc[accept]
                _, _, r3 = rate_of(seg, ANCHOR_DUE_COL)
                row[f"{prefix}_accept_rate"] = len(seg) / total
                row[f"{prefix}_accept_3m30p"] = r3
            row["is_current_mlt"] = mlt_cut == 5 and sub_cut == SUB_BIN_COUNT
            rows.append(row)
    return pd.DataFrame(rows)


def run_cond(
    dataset_cfg: dict,
    model_a_cfg: dict,
    model_b_cfg: dict,
    report_path: Path,
) -> None:
    """cond 模式主流程：A 模型各档内 B 模型条件子箱分析 Excel。"""
    t0 = time.time()
    tag_a, tag_b = model_a_cfg["cross_tag"], model_b_cfg["cross_tag"]
    print("=" * 60)
    print(f"两模型交叉分析（cond）：{model_a_cfg['name']} 档内 {model_b_cfg['name']} 条件子箱")
    print("=" * 60)

    final_a, edges_a, ranges_a = build_model_final_bins(dataset_cfg, model_a_cfg)
    final_b, edges_b, ranges_b = build_model_final_bins(dataset_cfg, model_b_cfg)
    cross = build_cross_sample(final_a, final_b, model_a_cfg, model_b_cfg)
    train_cross = cross.loc[cross["sample_group"].eq("train")].copy()
    print(f"[1/6] 双分样本：{len(cross):,} 笔（Train {len(train_cross):,}）")

    tier_edges = learn_sub_bin_edges(train_cross, tag_a, tag_b)
    cross = apply_sub_bins(cross, tier_edges, tag_a, tag_b)
    edge_table = build_sub_bin_edge_table(tier_edges)
    print(f"[2/6] 条件分箱完成：{len(edge_table)} 个子箱边界")

    train_stats = build_combined_stats(cross, "train", tag_a, tag_b)
    oot_stats = build_combined_stats(cross, "oot", tag_a, tag_b)
    print(f"[3/6] 组合格统计完成（Train {len(train_stats)} 格 / OOT {len(oot_stats)} 格）")

    tier_summary = build_tier_sub_bin_summary(train_stats)
    head_tail = build_head_tail_eval(train_stats, oot_stats, cross, tag_a)
    monotonicity = pd.concat([
        check_combined_monotonicity(train_stats, "train"),
        check_combined_monotonicity(oot_stats, "oot"),
    ], ignore_index=True)
    psi_combined = calc_combined_psi(train_stats, oot_stats)
    print(f"[4/6] 档内评估完成：组合分布 PSI={psi_combined:.4f}")

    disc = build_discrimination_comparison(cross, tag_a, tag_b, model_a_cfg, model_b_cfg)
    iv_cmp = build_iv_comparison(cross, tag_a, tag_b, model_a_cfg, model_b_cfg)
    seg_relay = build_segment_relayer_table(cross, tag_a, tag_b)
    strategy_table = build_sub_bin_strategy_table(cross, tag_a, tag_b)
    strategy_grid = build_sub_bin_strategy_grid(cross, tag_a, tag_b)
    print("[5/6] 区分度对比与二维策略模拟完成")

    overview = pd.DataFrame([
        {"section": "样本", "metric": "双分样本量（Train / OOT）", "value": f"{len(train_cross):,} / {int(cross['sample_group'].eq('oot').sum()):,}"},
        {"section": "设计", "metric": "子箱切分方式", "value": f"每个 {tag_a} 档内对 {tag_b} 分等频切 {SUB_BIN_COUNT} 个子箱（Train 学边界、OOT 复用）"},
        {"section": "设计", "metric": "组合格数量", "value": len(train_stats)},
        {"section": "设计", "metric": "组合分布 Train/OOT PSI", "value": psi_combined},
    ])
    appendix = pd.DataFrame([
        {"config_group": "基础配置", "config_name": "SUB_BIN_COUNT", "config_value": SUB_BIN_COUNT},
        {"config_group": "基础配置", "config_name": "ANCHOR_DUE_COL", "config_value": ANCHOR_DUE_COL},
        {"config_group": "基础配置", "config_name": "MIN_CELL_N_FOR_RATE", "config_value": MIN_CELL_N_FOR_RATE},
    ])
    print("[6/6] 总览与附录完成")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "01_总览", [("条件分箱总览", None), (None, overview)])
    write_sheet(wb, "02_条件分箱边界", [("档内子箱边界（Train 学习，OOT 复用）", None), (None, edge_table)])
    write_sheet(wb, "03_组合格统计_Train", [("Train 组合格统计", None), (None, train_stats)])
    write_sheet(wb, "04_组合格统计_OOT", [("OOT 组合格统计", None), (None, oot_stats)])
    write_sheet(wb, "05_档内子箱与头尾评估", [
        ("档内子箱风险与相邻显著性（Train）", None),
        (None, tier_summary),
        ("头尾拉开评估", None),
        (None, head_tail),
    ])
    write_sheet(wb, "06_区分度对比", [
        ("序数 AUC / KS 对比", None),
        (None, disc),
        ("3M30+ IV 对比", None),
        (None, iv_cmp),
        ("组合格单调性", None),
        (None, monotonicity),
        ("现行策略分段内的子箱再分层（应用示意）", None),
        (None, seg_relay),
    ])
    write_sheet(wb, "07_二维策略模拟", [
        ("策略对照（A 档上限 × 子箱上限）", None),
        (None, strategy_table),
        ("AND 接纳网格（A 档上限 × 子箱上限）", None),
        (None, strategy_grid),
    ])
    write_sheet(wb, "08_附录", [("配置参数", None), (None, appendix)])
    wb.save(report_path)
    print(f"完成 => {report_path}（耗时 {time.time() - t0:.1f}s）")
