# -*- coding: utf-8 -*-
"""
两模型交叉分析：mlt 主风险模型 × 价值模型（worthiness）

以 binning_mlt_cnt.py（mlt 笔数版）与 binning_worthiness_cnt.py（价值模型）各自的最终风险档为轴，
对同时存在两个模型分的样本做 7×7 交叉矩阵分析，并评估组合评分与二维策略效果。

核心内容（对应日志 1/7 ~ 7/7）：
1. 分别复用两个模型管线（数据加载 → Train/OOT 切分 → 20 等频初分 → 自动合箱）生成各自的
   最终风险档与分数边界，按 application_id 内连接得到 306,149 笔双分样本；
2. 计算两模型分的 Pearson / Spearman 相关性与分档一致性；
3. 生成 Train / OOT 的 7×7 交叉矩阵：每格样本量、占比、1M30+/3M30+ 笔数逾期率与金额逾期率、
   3M30+ Lift，附行列边际；
4. 条件增量分析：每个 mlt 档位内价值模型档位的风险跨度，以及两模型强分歧格（档位差 ≥ 3）的
   规模与风险；
5. 组合评分效果：z-score 平均、7:3 加权、档位平均、档位取大四种组合分，与两个单模型分对比
   Train/OOT 的 AUC / KS；
6. 二维策略模拟：按（mlt 档位上限, 价值档位上限）的 AND / OR 组合测算自动通过、总接纳与拒绝
   三段流量和风险，并与两个单模型现行策略对照；输出分歧象限人群的规模与风险；
7. 输出 out/binning_cross_strategy_report_YYYYMMDD.xlsx（7 个 sheet）。

运行方式：
    python binning_cross_mlt_wth.py
"""

import time
from itertools import product
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

import numpy as np
import pandas as pd

import binning_mlt_cnt as mlt
import binning_worthiness_cnt as wth
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

OUT_DIR = Path("out")
REPORT_PATH = OUT_DIR / f"binning_cross_strategy_report_{time.strftime('%Y%m%d')}.xlsx"

# 两模型现行单模型策略阈值（档位口径，来源：各自分箱报告的最终方案）。
# mlt：自动通过止于 C 档（阈值 0.0803750459943264）、总接纳止于 E 档（0.1821580944836785）；
# 价值模型：自动通过止于 B 档（0.1362170673263007）、总接纳止于 C 档（0.1863252117841281）。
MLT_AUTO_BIN = 3
MLT_ACCEPT_BIN = 5
WTH_AUTO_BIN = 2
WTH_ACCEPT_BIN = 3

# 组合分权重（mlt 为主、价值模型为辅）。
COMBO_MLT_WEIGHT = 0.7

# 强分歧格：两模型档位差达到该值的交叉格单独汇总。
DISAGREE_RANK_GAP = 3

# 交叉矩阵中样本过少的格不展示风险率（保留规模信息）。
MIN_CELL_N_FOR_RATE = 50


# ============================================================
# 1. 复用两个模型管线生成最终风险档
# ============================================================


def build_model_final_bins(
    module,
    tag: str,
) -> Tuple[pd.DataFrame, pd.DataFrame, List[Tuple[int, int]]]:
    """运行单个模型管线的分箱部分，返回 (最终分档宽表, 最终边界表, 选中方案)。"""
    data = module.load_analysis_data()
    all_data, train, oot = module.split_train_oot(data)
    edges = module.learn_equal_freq_edges(train, module.SCORE_COL, module.INITIAL_BIN_COUNT)
    initial_edges = module.build_initial_edge_table(edges)
    initial_bin_count = len(edges) - 1

    all_binned = module.apply_edges(all_data, module.SCORE_COL, edges, module.INITIAL_BIN_COL)
    train_binned = all_binned.loc[all_binned["sample_group"].eq("train")].copy()
    train_initial_stats = module.calc_complete_initial_stats(train_binned, initial_edges)

    candidates, _steps, _protected = module.build_merge_candidate_score_table(
        train_initial_stats,
        initial_bin_count,
        module.STRATEGY_CONFIG,
    )
    selected_ranges = module.selected_ranges_from_candidate_table(candidates)
    merge_map = module.build_merge_map(selected_ranges, initial_bin_count)
    final_edges = module.build_final_edge_table(initial_edges, merge_map, initial_bin_count)
    final = module.apply_merge_map(all_binned, merge_map)

    print(
        f"[{tag}] 最终方案：{len(final_edges)} 档 "
        f"{module.format_merge_ranges(selected_ranges)}"
    )
    return final, final_edges, selected_ranges


def rename_model_columns(frame: pd.DataFrame, tag: str) -> pd.DataFrame:
    """把两个模型各自的最终档/档位/分数列统一重命名为带前缀的列。"""
    rename = {
        "bin_order": f"{tag}_bin_order",
    }
    return frame.rename(columns=rename)


def build_cross_sample(
    mlt_final: pd.DataFrame,
    wth_final: pd.DataFrame,
) -> pd.DataFrame:
    """按 application_id 内连接两模型最终分档，得到双分样本。"""
    mlt_cols = [
        "application_id",
        "sample_group",
        "application_month",
        "duedate_1m_30",
        "duedate_3m_30",
        "principal",
        "estimate_principal_remaining_mob1",
        "estimate_principal_remaining_mob3",
        "dpd_days_ever_mob1",
        "dpd_days_ever_mob3",
        mlt.FINAL_BIN_COL,
        "bin_order",
        mlt.SCORE_COL,
    ]
    wth_cols = [
        "application_id",
        wth.FINAL_BIN_COL,
        "bin_order",
        wth.SCORE_COL,
    ]

    mlt_side = rename_model_columns(mlt_final[mlt_cols], "mlt").rename(
        columns={
            mlt.FINAL_BIN_COL: "mlt_bin",
            mlt.SCORE_COL: "mlt_score",
        }
    )
    wth_side = rename_model_columns(wth_final[wth_cols], "wth").rename(
        columns={
            wth.FINAL_BIN_COL: "wth_bin",
            wth.SCORE_COL: "wth_score",
        }
    )
    cross = mlt_side.merge(wth_side, on="application_id", how="inner")
    for col in ["mlt_score", "wth_score"]:
        cross[col] = pd.to_numeric(cross[col], errors="coerce")
    return cross


# ============================================================
# 2. 风险指标与矩阵统计
# ============================================================


def rate_of(frame: pd.DataFrame, due_col: str) -> Tuple[int, int, float]:
    """笔数口径：返回 (成熟量, 坏样本量, 逾期率)。"""
    values = pd.to_numeric(frame[due_col], errors="coerce")
    mature_mask = values.isin([0, 1])
    mature = int(mature_mask.sum())
    bad = int(values.eq(1).sum())
    rate = bad / mature if mature else np.nan
    return mature, bad, rate


def amt_rate_of(frame: pd.DataFrame, dpd_col: str, remaining_col: str) -> float:
    """金额口径逾期率：dpd >= 30 样本的剩余本金 / 成熟样本本金敞口。"""
    dpd = pd.to_numeric(frame[dpd_col], errors="coerce")
    principal = pd.to_numeric(frame["principal"], errors="coerce").fillna(0)
    remaining = pd.to_numeric(frame[remaining_col], errors="coerce")
    exposure = float(principal.loc[dpd.notna()].sum())
    bad_amt = float(remaining.loc[dpd.ge(30)].fillna(0).sum())
    return bad_amt / exposure if exposure else np.nan


def group_overall_rates(frame: pd.DataFrame) -> Dict[str, float]:
    """样本组整体四项风险率，用于 Lift 计算。"""
    out = {}
    for prefix, due_col, dpd_col, remaining_col in [
        ("1m30p_cnt", "duedate_1m_30", "dpd_days_ever_mob1", "estimate_principal_remaining_mob1"),
        ("3m30p_cnt", "duedate_3m_30", "dpd_days_ever_mob3", "estimate_principal_remaining_mob3"),
    ]:
        _, _, rate = rate_of(frame, due_col)
        out[f"{prefix}_bad_rate"] = rate
        out[f"{prefix[:-4]}_amt_bad_rate"] = amt_rate_of(frame, dpd_col, remaining_col)
    return out


def build_cross_matrix(
    cross: pd.DataFrame,
    sample_group: str,
    mlt_edges: pd.DataFrame,
    wth_edges: pd.DataFrame,
) -> pd.DataFrame:
    """生成某样本组的 7×7 交叉矩阵（含行列边际）。"""
    group = cross.loc[cross["sample_group"].eq(sample_group)].copy()
    overall = group_overall_rates(group)
    group_total = len(group)

    orders = sorted(
        set(group["mlt_bin_order"].dropna().astype(int))
        & set(group["wth_bin_order"].dropna().astype(int))
    )
    rows: List[Dict] = []
    for mo, wo in product(orders, orders):
        cell = group.loc[group["mlt_bin_order"].eq(mo) & group["wth_bin_order"].eq(wo)]
        n = len(cell)
        _, _, r1 = rate_of(cell, "duedate_1m_30")
        _, _, r3 = rate_of(cell, "duedate_3m_30")
        a1 = amt_rate_of(cell, "dpd_days_ever_mob1", "estimate_principal_remaining_mob1")
        a3 = amt_rate_of(cell, "dpd_days_ever_mob3", "estimate_principal_remaining_mob3")
        rows.append(
            {
                "mlt_bin_order": mo,
                "mlt_bin": _label(mlt_edges, mo),
                "wth_bin_order": wo,
                "wth_bin": _label(wth_edges, wo),
                "n": n,
                "sample_pct": n / group_total,
                "1m30p_cnt_bad_rate": r1,
                "3m30p_cnt_bad_rate": r3,
                "1m30p_amt_bad_rate": a1,
                "3m30p_amt_bad_rate": a3,
                "3m30p_cnt_lift": r3 / overall["3m30p_cnt_bad_rate"]
                if overall["3m30p_cnt_bad_rate"] else np.nan,
            }
        )
    matrix = pd.DataFrame(rows)

    # 行边际（mlt 单模型口径）与列边际（价值单模型口径）。
    for mo in orders:
        row_cell = group.loc[group["mlt_bin_order"].eq(mo)]
        _, _, r3 = rate_of(row_cell, "duedate_3m_30")
        _, _, r1 = rate_of(row_cell, "duedate_1m_30")
        rows.append(
            {
                "mlt_bin_order": mo,
                "mlt_bin": _label(mlt_edges, mo),
                "wth_bin_order": 0,
                "wth_bin": "行边际",
                "n": len(row_cell),
                "sample_pct": len(row_cell) / group_total,
                "1m30p_cnt_bad_rate": r1,
                "3m30p_cnt_bad_rate": r3,
                "1m30p_amt_bad_rate": amt_rate_of(row_cell, "dpd_days_ever_mob1", "estimate_principal_remaining_mob1"),
                "3m30p_amt_bad_rate": amt_rate_of(row_cell, "dpd_days_ever_mob3", "estimate_principal_remaining_mob3"),
                "3m30p_cnt_lift": r3 / overall["3m30p_cnt_bad_rate"]
                if overall["3m30p_cnt_bad_rate"] else np.nan,
            }
        )
    for wo in orders:
        col_cell = group.loc[group["wth_bin_order"].eq(wo)]
        _, _, r3 = rate_of(col_cell, "duedate_3m_30")
        _, _, r1 = rate_of(col_cell, "duedate_1m_30")
        rows.append(
            {
                "mlt_bin_order": 0,
                "mlt_bin": "列边际",
                "wth_bin_order": wo,
                "wth_bin": _label(wth_edges, wo),
                "n": len(col_cell),
                "sample_pct": len(col_cell) / group_total,
                "1m30p_cnt_bad_rate": r1,
                "3m30p_cnt_bad_rate": r3,
                "1m30p_amt_bad_rate": amt_rate_of(col_cell, "dpd_days_ever_mob1", "estimate_principal_remaining_mob1"),
                "3m30p_amt_bad_rate": amt_rate_of(col_cell, "dpd_days_ever_mob3", "estimate_principal_remaining_mob3"),
                "3m30p_cnt_lift": r3 / overall["3m30p_cnt_bad_rate"]
                if overall["3m30p_cnt_bad_rate"] else np.nan,
            }
        )
    rows.append(
        {
            "mlt_bin_order": 0,
            "mlt_bin": "整体",
            "wth_bin_order": 0,
            "wth_bin": "整体",
            "n": group_total,
            "sample_pct": 1.0,
            "1m30p_cnt_bad_rate": overall["1m30p_cnt_bad_rate"],
            "3m30p_cnt_bad_rate": overall["3m30p_cnt_bad_rate"],
            "1m30p_amt_bad_rate": overall["1m30p_amt_bad_rate"],
            "3m30p_amt_bad_rate": overall["3m30p_amt_bad_rate"],
            "3m30p_cnt_lift": 1.0,
        }
    )
    return pd.DataFrame(rows)


def _label(edges: pd.DataFrame, order: int) -> str:
    hit = edges.loc[edges["final_bin_order"].eq(order)]
    return str(hit.iloc[0]["final_bin_col_label"]) if not hit.empty else str(order)


def add_bin_labels(edges: pd.DataFrame, bin_col: str) -> pd.DataFrame:
    """给最终边界表补一个统一标签列（final_bin_col_label）。"""
    result = edges.copy()
    result["final_bin_col_label"] = result[bin_col]
    return result


# ============================================================
# 3. 条件增量分析
# ============================================================


def build_conditional_analysis(cross: pd.DataFrame) -> pd.DataFrame:
    """每个 mlt 档位内价值模型档位的风险跨度，与强分歧格汇总。"""
    group = cross.loc[cross["sample_group"].eq("train")].copy()
    group_total = len(group)
    overall = group_overall_rates(group)
    rows: List[Dict] = []

    # 行内跨度：固定 mlt 档，看价值档 A→G 的风险变化。
    for mo in sorted(group["mlt_bin_order"].dropna().unique().astype(int)):
        row_cell = group.loc[group["mlt_bin_order"].eq(mo)]
        _, _, row_rate = rate_of(row_cell, "duedate_3m_30")
        col_rates = {}
        for wo in sorted(group["wth_bin_order"].dropna().unique().astype(int)):
            cell = row_cell.loc[row_cell["wth_bin_order"].eq(wo)]
            if len(cell) >= MIN_CELL_N_FOR_RATE:
                _, _, r3 = rate_of(cell, "duedate_3m_30")
                col_rates[wo] = r3
        vals = sorted(col_rates.values())
        rows.append(
            {
                "dimension": "mlt 行内（价值模型增量）",
                "anchor_bin_order": mo,
                "anchor_bin": chr(ord("A") + mo - 1),
                "anchor_n": len(row_cell),
                "anchor_3m30p_cnt_bad_rate": row_rate,
                "other_min_rate": vals[0] if vals else np.nan,
                "other_max_rate": vals[-1] if vals else np.nan,
                "other_spread": (vals[-1] - vals[0]) if vals else np.nan,
            }
        )

    # 列内跨度：固定价值档，看 mlt 档 A→G 的风险变化。
    for wo in sorted(group["wth_bin_order"].dropna().unique().astype(int)):
        col_cell = group.loc[group["wth_bin_order"].eq(wo)]
        _, _, col_rate = rate_of(col_cell, "duedate_3m_30")
        row_rates = {}
        for mo in sorted(group["mlt_bin_order"].dropna().unique().astype(int)):
            cell = col_cell.loc[col_cell["mlt_bin_order"].eq(mo)]
            if len(cell) >= MIN_CELL_N_FOR_RATE:
                _, _, r3 = rate_of(cell, "duedate_3m_30")
                row_rates[mo] = r3
        vals = sorted(row_rates.values())
        rows.append(
            {
                "dimension": "wth 列内（mlt 增量）",
                "anchor_bin_order": wo,
                "anchor_bin": chr(ord("A") + wo - 1),
                "anchor_n": len(col_cell),
                "anchor_3m30p_cnt_bad_rate": col_rate,
                "other_min_rate": vals[0] if vals else np.nan,
                "other_max_rate": vals[-1] if vals else np.nan,
                "other_spread": (vals[-1] - vals[0]) if vals else np.nan,
            }
        )

    # 强分歧格：两模型档位差 ≥ DISAGREE_RANK_GAP。
    rank_gap = (
        group["mlt_bin_order"].astype(int) - group["wth_bin_order"].astype(int)
    ).abs()
    disagree = group.loc[rank_gap.ge(DISAGREE_RANK_GAP)].copy()
    agree = group.loc[rank_gap.le(1)].copy()
    for label, frame in [("强分歧（档位差 ≥ 3）", disagree), ("近似一致（档位差 ≤ 1）", agree)]:
        _, _, r3 = rate_of(frame, "duedate_3m_30")
        _, _, r1 = rate_of(frame, "duedate_1m_30")
        rows.append(
            {
                "dimension": "分档一致性",
                "anchor_bin_order": np.nan,
                "anchor_bin": label,
                "anchor_n": len(frame),
                "anchor_sample_pct": len(frame) / group_total,
                "anchor_3m30p_cnt_bad_rate": r3,
                "other_1m30p_cnt_bad_rate": r1,
                "other_min_rate": np.nan,
                "other_max_rate": np.nan,
                "other_spread": np.nan,
            }
        )

    # 按分歧方向细分：mlt 低风险 / 价值高风险 与 价值低风险 / mlt 高风险。
    for label, mask in [
        ("mlt 低风险 & 价值高风险（mlt ≤ 3 & wth ≥ 5）",
         group["mlt_bin_order"].le(3) & group["wth_bin_order"].ge(5)),
        ("价值低风险 & mlt 高风险（wth ≤ 3 & mlt ≥ 5）",
         group["wth_bin_order"].le(3) & group["mlt_bin_order"].ge(5)),
    ]:
        frame = group.loc[mask]
        _, _, r3 = rate_of(frame, "duedate_3m_30")
        _, _, r1 = rate_of(frame, "duedate_1m_30")
        rows.append(
            {
                "dimension": "分歧方向",
                "anchor_bin_order": np.nan,
                "anchor_bin": label,
                "anchor_n": len(frame),
                "anchor_sample_pct": len(frame) / group_total,
                "anchor_3m30p_cnt_bad_rate": r3,
                "other_1m30p_cnt_bad_rate": r1,
                "other_min_rate": np.nan,
                "other_max_rate": np.nan,
                "other_spread": np.nan,
            }
        )

    return pd.DataFrame(rows)


# ============================================================
# 4. 组合评分效果
# ============================================================


def calc_auc_ks(score: pd.Series, label: pd.Series) -> Tuple[float, float]:
    """秩和口径 AUC 与 KS（不依赖 sklearn）。"""
    mask = score.notna() & label.isin([0, 1])
    s = score.loc[mask].to_numpy(dtype="float64")
    y = label.loc[mask].to_numpy(dtype="float64")
    n_bad = int(y.sum())
    n_good = int((y == 0).sum())
    if n_bad == 0 or n_good == 0:
        return np.nan, np.nan
    order = np.argsort(s, kind="mergesort")
    y_sorted = y[order]
    ranks = np.where(y_sorted == 1)[0] + 1
    auc = (float(ranks.sum()) - n_bad * (n_bad + 1) / 2) / (n_bad * n_good)
    bad_cdf = np.cumsum(y_sorted) / n_bad
    good_cdf = np.cumsum(1 - y_sorted) / n_good
    ks = float(np.max(np.abs(bad_cdf - good_cdf)))
    return auc, ks


def build_combined_scores(cross: pd.DataFrame) -> pd.DataFrame:
    """z-score 标准化两模型分并构造组合分。"""
    result = cross.copy()
    train_mask = result["sample_group"].eq("train")
    for col in ["mlt_score", "wth_score"]:
        mean = result.loc[train_mask, col].mean()
        std = result.loc[train_mask, col].std()
        result[f"z_{col}"] = (result[col] - mean) / std

    result["combo_avg"] = (result["z_mlt_score"] + result["z_wth_score"]) / 2
    result["combo_w7030"] = (
        COMBO_MLT_WEIGHT * result["z_mlt_score"]
        + (1 - COMBO_MLT_WEIGHT) * result["z_wth_score"]
    )
    result["combo_rank_avg"] = (
        result["mlt_bin_order"].astype(float) + result["wth_bin_order"].astype(float)
    ) / 2
    result["combo_rank_max"] = result[
        ["mlt_bin_order", "wth_bin_order"]
    ].max(axis=1).astype(float)
    return result


def build_score_performance(cross: pd.DataFrame) -> pd.DataFrame:
    """单模型分与组合分的 Train/OOT AUC / KS 对比。"""
    score_cols = [
        ("mlt 单模型分", "mlt_score"),
        ("价值模型单模型分", "wth_score"),
        ("组合分：z 平均", "combo_avg"),
        ("组合分：7:3 加权", "combo_w7030"),
        ("组合分：档位平均", "combo_rank_avg"),
        ("组合分：档位取大", "combo_rank_max"),
    ]
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        for score_name, score_col in score_cols:
            for label_name, label_col in [
                ("1M30+", "duedate_1m_30"),
                ("3M30+", "duedate_3m_30"),
            ]:
                mask = frame[label_col].isin([0, 1]) & frame[score_col].notna()
                n = int(mask.sum())
                bad = int(frame.loc[mask, label_col].eq(1).sum())
                auc, ks = calc_auc_ks(frame[score_col], frame[label_col])
                rows.append(
                    {
                        "sample_group": group_name,
                        "score": score_name,
                        "label": label_name,
                        "n": n,
                        "bad_cnt": bad,
                        "bad_rate": bad / n if n else np.nan,
                        "auc": auc,
                        "ks": ks,
                    }
                )
    return pd.DataFrame(rows)


# ============================================================
# 5. 二维策略模拟
# ============================================================


def segment_mask(
    cross: pd.DataFrame,
    mlt_auto: int,
    wth_auto: int,
    mlt_accept: int,
    wth_accept: int,
    logic: str,
) -> Dict[str, pd.Series]:
    """按二维阈值与 AND/OR 逻辑切分自动通过 / 总接纳 / 拒绝。"""
    mlt_o = cross["mlt_bin_order"].astype(int)
    wth_o = cross["wth_bin_order"].astype(int)
    combine = np.logical_and if logic == "AND" else np.logical_or
    auto = combine(mlt_o.le(mlt_auto), wth_o.le(wth_auto))
    accept = combine(mlt_o.le(mlt_accept), wth_o.le(wth_accept))
    return {
        "auto": auto,
        "accept": accept,
        "reject": ~accept,
        "manual": accept & ~auto,
    }


def simulate_policy(
    cross: pd.DataFrame,
    name: str,
    logic: str,
    mlt_auto: int,
    wth_auto: int,
    mlt_accept: int,
    wth_accept: int,
) -> Dict[str, object]:
    """测算一套二维策略在 Train / OOT 的三段流量与风险。"""
    masks = segment_mask(cross, mlt_auto, wth_auto, mlt_accept, wth_accept, logic)
    row = {
        "policy": name,
        "logic": logic,
        "mlt_auto_bin": mlt_auto,
        "wth_auto_bin": wth_auto,
        "mlt_accept_bin": mlt_accept,
        "wth_accept_bin": wth_accept,
    }
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        total = len(frame)
        for seg in ["auto", "manual", "accept", "reject"]:
            seg_frame = frame.loc[masks[seg]]
            _, _, r1 = rate_of(seg_frame, "duedate_1m_30")
            _, _, r3 = rate_of(seg_frame, "duedate_3m_30")
            row[f"{group_name}_{seg}_rate"] = len(seg_frame) / total
            row[f"{group_name}_{seg}_1m30p"] = r1
            row[f"{group_name}_{seg}_3m30p"] = r3
    return row


def build_policy_table(cross: pd.DataFrame) -> pd.DataFrame:
    """两模型单模型现行策略 + 二维 AND / OR 组合策略对照。"""
    rows: List[Dict] = []
    rows.append(
        simulate_policy(
            cross,
            "mlt 单模型（现行）",
            "AND",
            MLT_AUTO_BIN,
            7,
            MLT_ACCEPT_BIN,
            7,
        )
    )
    rows.append(
        simulate_policy(
            cross,
            "价值模型单模型（现行）",
            "AND",
            7,
            WTH_AUTO_BIN,
            7,
            WTH_ACCEPT_BIN,
        )
    )
    rows.append(
        simulate_policy(
            cross,
            "二维组合（AND）",
            "AND",
            MLT_AUTO_BIN,
            WTH_AUTO_BIN,
            MLT_ACCEPT_BIN,
            WTH_ACCEPT_BIN,
        )
    )
    rows.append(
        simulate_policy(
            cross,
            "二维组合（OR）",
            "OR",
            MLT_AUTO_BIN,
            WTH_AUTO_BIN,
            MLT_ACCEPT_BIN,
            WTH_ACCEPT_BIN,
        )
    )
    return pd.DataFrame(rows)


def build_accept_grid(cross: pd.DataFrame) -> pd.DataFrame:
    """AND 逻辑下（mlt 档上限 × 价值档上限）接纳网格：Train 接纳率与 3M30+ 风险。"""
    rows: List[Dict] = []
    for mlt_cut in range(2, 8):
        for wth_cut in range(2, 8):
            row = simulate_policy(cross, "grid", "AND", mlt_cut, wth_cut, mlt_cut, wth_cut)
            rows.append(
                {
                    "mlt_accept_bin": mlt_cut,
                    "wth_accept_bin": wth_cut,
                    "train_accept_rate": row["train_accept_rate"],
                    "train_accept_3m30p": row["train_accept_3m30p"],
                    "oot_accept_rate": row["oot_accept_rate"],
                    "oot_accept_3m30p": row["oot_accept_3m30p"],
                    "is_current_and": mlt_cut == MLT_ACCEPT_BIN and wth_cut == WTH_ACCEPT_BIN,
                }
            )
    return pd.DataFrame(rows)


def build_quadrant_table(cross: pd.DataFrame) -> pd.DataFrame:
    """按现行接纳阈值（mlt ≤ E、价值 ≤ C）划分四个象限人群。"""
    rows: List[Dict] = []
    for group_name in ["train", "oot"]:
        frame = cross.loc[cross["sample_group"].eq(group_name)]
        total = len(frame)
        quadrants = [
            ("双低（mlt ≤ E 且 wth ≤ C）", frame["mlt_bin_order"].le(MLT_ACCEPT_BIN) & frame["wth_bin_order"].le(WTH_ACCEPT_BIN)),
            ("仅 mlt 低（mlt ≤ E 且 wth > C）", frame["mlt_bin_order"].le(MLT_ACCEPT_BIN) & frame["wth_bin_order"].gt(WTH_ACCEPT_BIN)),
            ("仅价值低（wth ≤ C 且 mlt > E）", frame["wth_bin_order"].le(WTH_ACCEPT_BIN) & frame["mlt_bin_order"].gt(MLT_ACCEPT_BIN)),
            ("双高（mlt > E 且 wth > C）", frame["mlt_bin_order"].gt(MLT_ACCEPT_BIN) & frame["wth_bin_order"].gt(WTH_ACCEPT_BIN)),
        ]
        for label, mask in quadrants:
            seg = frame.loc[mask]
            _, _, r1 = rate_of(seg, "duedate_1m_30")
            _, _, r3 = rate_of(seg, "duedate_3m_30")
            rows.append(
                {
                    "sample_group": group_name,
                    "quadrant": label,
                    "n": len(seg),
                    "sample_pct": len(seg) / total,
                    "1m30p_cnt_bad_rate": r1,
                    "3m30p_cnt_bad_rate": r3,
                }
            )
    return pd.DataFrame(rows)


# ============================================================
# 6. Excel 输出
# ============================================================


def write_sheet(
    wb: Workbook,
    sheet_name: str,
    frames: Sequence[Tuple[Optional[str], Optional[pd.DataFrame]]],
    freeze: str = "A2",
) -> None:
    """把若干 DataFrame 依次写入一个 sheet（标题行可选）。"""
    ws = wb.create_sheet(sheet_name)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    rate_fmt = "0.00%"
    row_idx = 1
    for title, frame in frames:
        if title:
            ws.cell(row=row_idx, column=1, value=title).font = Font(bold=True, size=12)
            row_idx += 1
        if frame is None:
            continue
        for col_idx, col_name in enumerate(frame.columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        row_idx += 1
        for _, record in frame.iterrows():
            for col_idx, col_name in enumerate(frame.columns, start=1):
                value = record[col_name]
                if pd.isna(value):
                    value = None
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if isinstance(value, float):
                    col_key = str(col_name)
                    if (
                        "rate" in col_key
                        or "pct" in col_key
                        or "1m30p" in col_key
                        or "3m30p" in col_key
                    ):
                        cell.number_format = rate_fmt
                    elif "auc" in col_key or "ks" in col_key or "lift" in col_key:
                        cell.number_format = "0.0000"
                    elif "corr" in col_key:
                        cell.number_format = "0.0000"
                elif isinstance(value, int):
                    cell.number_format = "#,##0"
            row_idx += 1
        row_idx += 1
    ws.freeze_panes = freeze
    for col_idx in range(1, 10):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def main() -> None:
    t0 = time.time()

    print("=" * 60)
    print("两模型交叉分析：mlt 主风险模型 × 价值模型")
    print("=" * 60)

    # 1) 两个模型各自的最终分档。
    mlt_final, mlt_edges, mlt_ranges = build_model_final_bins(mlt, "mlt")
    wth_final, wth_edges, wth_ranges = build_model_final_bins(wth, "wth")
    mlt_edges = add_bin_labels(mlt_edges, mlt.FINAL_BIN_COL)
    wth_edges = add_bin_labels(wth_edges, wth.FINAL_BIN_COL)

    cross = build_cross_sample(mlt_final, wth_final)
    train_n = int(cross["sample_group"].eq("train").sum())
    oot_n = int(cross["sample_group"].eq("oot").sum())
    print(f"[1/7] 双分样本：{len(cross):,} 笔（Train {train_n:,} / OOT {oot_n:,}）")

    # 2) 相关性。
    train_cross = cross.loc[cross["sample_group"].eq("train")]
    pearson = float(train_cross["mlt_score"].corr(train_cross["wth_score"]))
    spearman = float(train_cross["mlt_score"].corr(train_cross["wth_score"], method="spearman"))
    bin_spearman = float(
        train_cross["mlt_bin_order"].astype(int).corr(
            train_cross["wth_bin_order"].astype(int), method="spearman"
        )
    )
    print(
        f"[2/7] 相关性：Pearson {pearson:.4f} / Spearman {spearman:.4f} "
        f"/ 分档秩相关 {bin_spearman:.4f}"
    )

    # 3) 交叉矩阵。
    train_matrix = build_cross_matrix(cross, "train", mlt_edges, wth_edges)
    oot_matrix = build_cross_matrix(cross, "oot", mlt_edges, wth_edges)
    print(f"[3/7] 交叉矩阵完成（Train {len(train_matrix)} 行 / OOT {len(oot_matrix)} 行）")

    # 4) 条件增量。
    conditional = build_conditional_analysis(cross)
    print(f"[4/7] 条件增量分析完成（{len(conditional)} 行）")

    # 5) 组合评分。
    cross = build_combined_scores(cross)
    score_perf = build_score_performance(cross)
    print("[5/7] 组合评分效果完成")

    # 6) 二维策略模拟。
    policy_table = build_policy_table(cross)
    accept_grid = build_accept_grid(cross)
    quadrant_table = build_quadrant_table(cross)
    print("[6/7] 二维策略模拟完成")

    # 7) 总览与附录。
    mlt_best = score_perf.loc[
        score_perf["score"].eq("mlt 单模型分") & score_perf["label"].eq("3M30+")
    ].set_index("sample_group")
    combo_best = score_perf.loc[
        score_perf["score"].eq("组合分：7:3 加权") & score_perf["label"].eq("3M30+")
    ].set_index("sample_group")
    overview = pd.DataFrame(
        [
            {"section": "样本", "metric": "双分样本量", "value": len(cross)},
            {"section": "样本", "metric": "Train / OOT 样本量", "value": f"{train_n:,} / {oot_n:,}"},
            {"section": "样本", "metric": "双分覆盖率（相对 328,063 完成申请）", "value": len(cross) / 328063},
            {"section": "相关性", "metric": "模型分 Pearson 相关（Train）", "value": pearson},
            {"section": "相关性", "metric": "模型分 Spearman 相关（Train）", "value": spearman},
            {"section": "相关性", "metric": "最终分档秩相关（Train）", "value": bin_spearman},
            {"section": "分档方案", "metric": "mlt 最终方案", "value": mlt.format_merge_ranges(mlt_ranges)},
            {"section": "分档方案", "metric": "价值模型最终方案", "value": wth.format_merge_ranges(wth_ranges)},
            {"section": "组合效果", "metric": "Train 3M30+ AUC（mlt 单模型 / 7:3 组合）", "value": f"{mlt_best.loc['train','auc']:.4f} / {combo_best.loc['train','auc']:.4f}"},
            {"section": "组合效果", "metric": "OOT 3M30+ AUC（mlt 单模型 / 7:3 组合）", "value": f"{mlt_best.loc['oot','auc']:.4f} / {combo_best.loc['oot','auc']:.4f}"},
            {"section": "组合效果", "metric": "Train 3M30+ KS（mlt 单模型 / 7:3 组合）", "value": f"{mlt_best.loc['train','ks']:.4f} / {combo_best.loc['train','ks']:.4f}"},
            {"section": "组合效果", "metric": "OOT 3M30+ KS（mlt 单模型 / 7:3 组合）", "value": f"{mlt_best.loc['oot','ks']:.4f} / {combo_best.loc['oot','ks']:.4f}"},
            {"section": "二维策略", "metric": "现行单模型阈值（mlt / 价值）", "value": f"mlt 自动 ≤{MLT_AUTO_BIN} 档、接纳 ≤{MLT_ACCEPT_BIN} 档；价值 自动 ≤{WTH_AUTO_BIN} 档、接纳 ≤{WTH_ACCEPT_BIN} 档"},
            {"section": "二维策略", "metric": "组合逻辑", "value": "AND（两模型同时达标）与 OR（任一达标）对照"},
        ]
    )
    appendix = pd.DataFrame(
        [
            {"config_group": "基础配置", "config_name": "MLT_AUTO_BIN / MLT_ACCEPT_BIN", "config_value": f"{MLT_AUTO_BIN} / {MLT_ACCEPT_BIN}"},
            {"config_group": "基础配置", "config_name": "WTH_AUTO_BIN / WTH_ACCEPT_BIN", "config_value": f"{WTH_AUTO_BIN} / {WTH_ACCEPT_BIN}"},
            {"config_group": "基础配置", "config_name": "COMBO_MLT_WEIGHT", "config_value": COMBO_MLT_WEIGHT},
            {"config_group": "基础配置", "config_name": "DISAGREE_RANK_GAP", "config_value": DISAGREE_RANK_GAP},
            {"config_group": "基础配置", "config_name": "MIN_CELL_N_FOR_RATE", "config_value": MIN_CELL_N_FOR_RATE},
            {"config_group": "组合分口径", "config_name": "combo_avg", "config_value": "(z_mlt + z_wth) / 2"},
            {"config_group": "组合分口径", "config_name": "combo_w7030", "config_value": "0.7 × z_mlt + 0.3 × z_wth"},
            {"config_group": "组合分口径", "config_name": "combo_rank_avg / combo_rank_max", "config_value": "两模型最终档位序的平均 / 取大"},
        ]
    )
    print("[7/7] 总览与附录完成")

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "01_总览", [("两模型交叉分析总览", None), (None, overview)])
    write_sheet(wb, "02_交叉矩阵_Train", [("Train 7×7 交叉矩阵（行 = mlt 档，列 = 价值档）", None), (None, train_matrix)])
    write_sheet(wb, "03_交叉矩阵_OOT", [("OOT 7×7 交叉矩阵（行 = mlt 档，列 = 价值档）", None), (None, oot_matrix)])
    write_sheet(wb, "04_条件增量分析", [("条件增量与分档一致性", None), (None, conditional)])
    write_sheet(wb, "05_组合评分效果", [("单模型分与组合分 AUC / KS 对比", None), (None, score_perf)])
    write_sheet(wb, "06_二维策略模拟", [
        ("策略对照（单模型现行 vs 二维 AND / OR）", None),
        (None, policy_table),
        ("AND 接纳网格（mlt 档上限 × 价值档上限）", None),
        (None, accept_grid),
        ("现行阈值四象限人群", None),
        (None, quadrant_table),
    ])
    write_sheet(wb, "07_附录", [("配置参数", None), (None, appendix)])

    wb.save(REPORT_PATH)
    print(f"完成 => {REPORT_PATH}（耗时 {time.time() - t0:.1f}s）")


if __name__ == "__main__":
    main()
